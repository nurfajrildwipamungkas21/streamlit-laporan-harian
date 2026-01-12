import json
import streamlit as st
import pandas as pd
from datetime import date, datetime, timedelta
from zoneinfo import ZoneInfo
from pathlib import Path
import time
import gspread
from google.oauth2.service_account import Credentials
import dropbox
from dropbox.exceptions import AuthError, ApiError
from dropbox.sharing import RequestedVisibility, SharedLinkSettings
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import random
import string
import re
import io
import hashlib
import hmac
import base64
import textwrap

from audit_service import log_admin_action, compare_and_get_changes

# =========================================================
# [BARU] SISTEM LOGGING LANGSUNG (ANTI-GAGAL)
# =========================================================
# Ganti fungsi force_audit_log dengan ini


def force_audit_log(actor, action, target_sheet, chat_msg, details_input):
    try:
        SHEET_NAME = "Global_Audit_Log"
        try:
            ws = spreadsheet.worksheet(SHEET_NAME)
        except gspread.WorksheetNotFound:
            # Buat baru jika tidak ada dengan header standar
            ws = spreadsheet.add_worksheet(title=SHEET_NAME, rows=1000, cols=6)
            ws.append_row(["Waktu", "User", "Status", "Target Data", "Chat & Catatan",
                          "Detail Perubahan"], value_input_option="USER_ENTERED")

        # Ambil header untuk tahu urutan kolom
        headers = ws.row_values(1)

        ts = datetime.now(ZoneInfo("Asia/Jakarta")
                          ).strftime("%d-%m-%Y %H:%M:%S")
        final_details = "\n".join([f"• {k}: {v}" for k, v in details_input.items()]) if isinstance(
            details_input, dict) else (str(details_input) if details_input else "-")

        # Data yang akan dimasukkan (Mapping Keyword ke Value)
        payload = {
            "waktu": f"'{ts}",
            "user": str(actor),
            "pelaku": str(actor),
            "status": str(action),
            "aksi": str(action),  # Menangkap "Aksi Dilakukan"
            "target": str(target_sheet),
            "nama data": str(target_sheet),
            "chat": str(chat_msg),
            "catatan": str(chat_msg),
            "alasan": str(chat_msg),
            "detail": str(final_details),
            "rincian": str(final_details)
        }

        # Susun baris baru mengikuti urutan header di GSheet secara dinamis
        row_to_append = [""] * len(headers)
        for i, h in enumerate(headers):
            h_lower = h.lower()
            for key, val in payload.items():
                if key in h_lower:
                    row_to_append[i] = val
                    break

        # Jika baris masih kosong (header tidak cocok), gunakan format default di akhir
        if all(x == "" for x in row_to_append):
            row_to_append = [f"'{ts}", str(actor), str(action), str(
                target_sheet), str(chat_msg), str(final_details)]

        ws.append_row(row_to_append, value_input_option="USER_ENTERED")
        return True
    except Exception as e:
        print(f"⚠️ FORCE LOG ERROR: {e}")
        return False


# =========================================================
# ANCHOR: HELPER APPROVAL (AMBIL DARI CODE KEDUA)
# =========================================================
SHEET_PENDING = "System_Pending_Approval"


def init_pending_db():
    try:
        try:
            ws = spreadsheet.worksheet(SHEET_PENDING)
            headers = ws.row_values(1)
            if "Old Data JSON" not in headers:
                current_cols = ws.col_count
                new_col_idx = len(headers) + 1
                if current_cols < new_col_idx:
                    ws.resize(cols=new_col_idx)
                ws.update_cell(1, new_col_idx, "Old Data JSON")
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(
                title=SHEET_PENDING, rows=1000, cols=7)
            headers = ["Timestamp", "Requestor", "Target Sheet",
                       "Row Index (0-based)", "New Data JSON", "Reason", "Old Data JSON"]
            ws.append_row(headers, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
        return ws
    except Exception as e:
        print(f"Error init_pending_db: {e}")
        return None


def submit_change_request(target_sheet, row_idx_0based, new_df_row, old_df_row, reason, requestor):
    ws = init_pending_db()
    if not ws:
        return False, "DB Error"
    row_dict_new = new_df_row.astype(str).to_dict()
    json_new = json.dumps(row_dict_new)
    row_dict_old = old_df_row.astype(
        str).to_dict() if old_df_row is not None else {}
    json_old = json.dumps(row_dict_old)
    ts = now_ts_str()
    ws.append_row([ts, requestor, target_sheet, row_idx_0based,
                  json_new, reason, json_old], value_input_option="USER_ENTERED")

    diff_log = {}
    for k, v_new in row_dict_new.items():
        v_old = row_dict_old.get(k, "")
        if str(v_new).strip() != str(v_old).strip():
            diff_log[k] = f"{v_old} ➡ {v_new}"
    diff_str = "\n".join(
        [f"{k}: {v}" for k, v in diff_log.items()]) if diff_log else "Re-save data."

    force_audit_log(actor=requestor, action="⏳ PENDING", target_sheet=target_sheet,
                    chat_msg=f"🙋‍♂️ [ADMIN]: {reason}", details_input=diff_str)
    return True, "Permintaan terkirim!"

# =========================================================
# [BARU] HELPER: ADMIN SMART EDITOR (AUTO-APPROVAL)
# =========================================================
# =========================================================
# [REVISI V2] HELPER: ADMIN SMART EDITOR (FIX TITIK RUPIAH)
# =========================================================
def admin_smart_editor_ui(df_data, unique_key, sheet_target_name):
    """
    Komponen UI standar untuk Admin mengedit data tabel.
    [UPDATE]: Menampilkan format 'Rp 2.000.000' (ada titik) sebagai teks agar rapi,
    tapi otomatis dikonversi balik ke Angka saat disimpan.
    """
    if df_data.empty:
        st.info("Data tidak tersedia untuk ditampilkan.")
        return

    # --- 1. PRE-PROCESS: Format Tampilan (Uang jadi String dengan Titik) ---
    df_display = df_data.copy()
    column_configs = {}
    
    # Identifikasi kolom uang untuk diproses
    money_cols = []
    for col in df_display.columns:
        if any(k in col.lower() for k in ["nilai", "nominal", "harga", "total", "kontrak", "sisa", "sepakat", "amount", "bayar"]):
            money_cols.append(col)

    # Fungsi format rupiah lokal
    def _fmt_money(x):
        try:
            # Pastikan hanya angka
            val = parse_rupiah_to_int(x)
            if val is None: return ""
            # Format: Rp 2.000.000 (Titik sebagai pemisah ribuan)
            return "Rp " + "{:,.0f}".format(val).replace(",", ".")
        except:
            return str(x)

    for col in df_display.columns:
        col_lower = col.lower()
        
        # A. KOLOM UANG: Ubah ke String berformat "Rp 2.000.000"
        if col in money_cols:
            df_display[col] = df_display[col].apply(_fmt_money)
            # Kita set sebagai TextColumn agar user bisa edit teksnya langsung
            column_configs[col] = st.column_config.TextColumn(
                col,
                help="Format: Rp 2.000.000 (Edit angkanya saja, sistem akan merapikan)",
                width="medium"
            )

        # B. KOLOM TANGGAL: Tetap Date Picker
        elif any(k in col_lower for k in ["tanggal", "tgl", "date", "event", "tempo", "waktu"]):
            if "timestamp" not in col_lower: 
                try:
                    df_display[col] = pd.to_datetime(df_display[col], errors='coerce')
                    column_configs[col] = st.column_config.DateColumn(
                        col,
                        format="DD/MM/YYYY"
                    )
                except:
                    pass

    # --- 2. Tampilkan Editor Data ---
    edited_df = st.data_editor(
        df_display,
        key=f"editor_{unique_key}",
        use_container_width=True,
        num_rows="fixed",
        column_config=column_configs
    )

    # --- 3. Form Alasan & Tombol Submit ---
    col_reason, col_btn = st.columns([3, 1])
    with col_reason:
        reason = st.text_input("📝 Alasan Perubahan (Wajib diisi):", key=f"reason_{unique_key}", placeholder="Contoh: Koreksi typo, update nilai kontrak...")
    
    with col_btn:
        st.markdown("<br>", unsafe_allow_html=True) 
        is_submit = st.button("📤 Ajukan Perubahan", key=f"btn_{unique_key}", type="primary", use_container_width=True)

    if is_submit:
        if not reason.strip():
            st.error("⚠️ Harap isi alasan perubahan terlebih dahulu!")
        else:
            # --- 4. Logic Deteksi Perubahan & Pembersihan Data ---
            changes_found = []
            
            # Gunakan string compare untuk mendeteksi perubahan visual
            df_old_str = df_display.astype(str).reset_index(drop=True)
            df_new_str = edited_df.astype(str).reset_index(drop=True)

            for i in df_old_str.index:
                row_old = df_old_str.iloc[i]
                row_new = df_new_str.iloc[i]
                
                if not row_old.equals(row_new):
                    # Ambil data baris yang diedit
                    dirty_row = edited_df.iloc[i].copy()
                    
                    # [PENTING] BERSIHKAN KEMBALI FORMAT RUPIAH KE INTEGER
                    # Agar database menerima angka murni (2000000), bukan string "Rp 2.000.000"
                    for m_col in money_cols:
                        if m_col in dirty_row:
                            # Bersihkan "Rp", Titik, Koma -> Jadi Integer
                            raw_val = dirty_row[m_col]
                            clean_val = parse_rupiah_to_int(raw_val)
                            # Jika hasil cleaning valid, simpan. Jika error/kosong, simpan 0 atau biarkan.
                            dirty_row[m_col] = clean_val if clean_val is not None else 0

                    changes_found.append({
                        "row_idx": i,
                        "new_data": dirty_row, # Data yang sudah dibersihkan jadi angka
                        "old_data": df_data.iloc[i] # Data asli (angka) dari database
                    })

            if not changes_found:
                st.warning("Tidak ada perubahan data yang terdeteksi.")
            else:
                # Kirim Request
                success_count = 0
                current_user = st.session_state.get("user_name", "Admin")
                progress_bar = st.progress(0, text="Mengirim permintaan ke Manager...")

                for idx, change in enumerate(changes_found):
                    ok, msg = submit_change_request(
                        target_sheet=sheet_target_name,
                        row_idx_0based=change['row_idx'],
                        new_df_row=change['new_data'],
                        old_df_row=change['old_data'],
                        reason=reason,
                        requestor=current_user
                    )
                    if ok: success_count += 1
                    progress_bar.progress((idx + 1) / len(changes_found))
                
                progress_bar.empty()

                if success_count > 0:
                    st.success(f"✅ Berhasil mengirim {success_count} permintaan perubahan!")
                    time.sleep(2)
                    st.rerun()
                else:
                    st.error("Gagal mengirim permintaan.")

def execute_approval(request_index_0based, action, admin_name="Manager", rejection_note="-"):
    try:
        ws_pending = init_pending_db()
        all_requests = ws_pending.get_all_records()
        if request_index_0based >= len(all_requests):
            return False, "Data tidak ditemukan."
        req = all_requests[request_index_0based]

        if action == "REJECT":
            force_audit_log(actor=admin_name, action="❌ DITOLAK",
                            target_sheet=req["Target Sheet"], chat_msg=f"⛔ [MANAGER]: {rejection_note}", details_input=f"Pengaju: {req['Requestor']}")
            ws_pending.delete_rows(request_index_0based + 2)
            return True, "Ditolak."

        elif action == "APPROVE":
            new_data_dict = json.loads(req["New Data JSON"])
            ws_target = spreadsheet.worksheet(req["Target Sheet"])
            headers = ws_target.row_values(1)
            row_values = [new_data_dict.get(h, "") for h in headers]
            gsheet_row = int(req["Row Index (0-based)"]) + 2
            ws_target.update(range_name=f"A{gsheet_row}", values=[
                             row_values], value_input_option="USER_ENTERED")
            force_audit_log(actor=admin_name, action="✅ SUKSES/ACC",
                            target_sheet=req["Target Sheet"], chat_msg="✅ [MANAGER]: Disetujui.", details_input=f"Pengaju: {req['Requestor']}")
            ws_pending.delete_rows(request_index_0based + 2)
            return True, "Disetujui."
    except Exception as e:
        return False, str(e)


def init_pending_db():
    """Memastikan sheet pending approval ada dengan kolom untuk DATA LAMA."""
    try:
        # Blok Try Dalam (Mencoba ambil worksheet)
        try:
            ws = spreadsheet.worksheet(SHEET_PENDING)
            # Cek apakah header sudah update (punya Old Data JSON)
            headers = ws.row_values(1)
            if "Old Data JSON" not in headers:
                # PERBAIKAN: Resize sheet dulu sebelum update cell di kolom baru
                current_cols = ws.col_count
                new_col_idx = len(headers) + 1
                if current_cols < new_col_idx:
                    ws.resize(cols=new_col_idx)  # Tambah kolom jika kurang

                ws.update_cell(1, new_col_idx, "Old Data JSON")

        except gspread.WorksheetNotFound:
            # Jika tidak ada, buat baru
            ws = spreadsheet.add_worksheet(
                title=SHEET_PENDING, rows=1000, cols=7)
            headers = ["Timestamp", "Requestor", "Target Sheet",
                       "Row Index (0-based)", "New Data JSON", "Reason", "Old Data JSON"]
            ws.append_row(headers, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)

        return ws

    except Exception as e:  # <--- PASTIKAN BAGIAN INI ADA DAN SEJAJAR DENGAN TRY PERTAMA
        # Tampilkan error di terminal untuk debugging jika terjadi lagi
        print(f"Error init_pending_db: {e}")
        return None


def submit_change_request(target_sheet, row_idx_0based, new_df_row, old_df_row, reason, requestor):
    """
    UPDATE: Menggabungkan logic penyimpanan pending data dan pencatatan log (Audit Trail).
    Status: PENDING | Detail: List Perubahan (String) | Alasan: Format Chat Admin.
    """
    # --- 1. Inisialisasi Database Pending ---
    ws = init_pending_db()
    if not ws:
        return False, "DB Error"

    # --- 2. Persiapan Data (Konversi ke JSON) ---
    # Konversi row dataframe baru ke dictionary
    row_dict_new = new_df_row.astype(str).to_dict()
    json_new = json.dumps(row_dict_new)

    # Konversi row dataframe lama ke dictionary
    row_dict_old = old_df_row.astype(
        str).to_dict() if old_df_row is not None else {}
    json_old = json.dumps(row_dict_old)

    ts = now_ts_str()

    # --- 3. Simpan ke System_Pending_Approval ---
    # (Data ini wajib disimpan agar Manager bisa melihat data asli vs baru saat approval)
    ws.append_row(
        [ts, requestor, target_sheet, row_idx_0based, json_new, reason, json_old],
        value_input_option="USER_ENTERED"
    )

    # --- 4. Hitung Perbedaan (Diff Logic) ---
    diff_log = {}
    for k, v_new in row_dict_new.items():
        v_old = row_dict_old.get(k, "")
        # Normalisasi string agar tidak false alarm (spasi, dll)
        if str(v_new).strip() != str(v_old).strip():
            diff_log[k] = f"{v_old} ➡ {v_new}"

    # Flatten Dictionary ke String (agar muncul rapi di kolom 'Detail Perubahan')
    if not diff_log:
        diff_str = "Tidak ada perubahan data terdeteksi (Re-save)."
    else:
        # Join setiap item dengan enter (\n) agar rapi list ke bawah
        diff_str = "\n".join([f"{k}: {v}" for k, v in diff_log.items()])

    # --- 5. Format Chat & Catat Log (Revisi dari Code Kedua) ---

    # Format Chat Admin agar lebih interaktif di UI
    final_chat = f"🙋‍♂️ [ADMIN]: {reason}" if reason else "🙋‍♂️ [ADMIN]: Request Update Data."

    # Panggil fungsi logging yang baru
    force_audit_log(
        actor=requestor,
        action="⏳ PENDING",       # Status Jelas
        target_sheet=target_sheet,
        chat_msg=final_chat,       # Masuk ke kolom "Chat & Catatan"
        details_input=diff_str     # Masuk ke kolom "Detail Perubahan"
    )

    return True, "Permintaan terkirim & Log tercatat!"


def get_pending_approvals():
    """Fungsi untuk Manager mengambil semua daftar request yang pending."""
    ws = init_pending_db()
    if not ws:
        return []
    return ws.get_all_records()


def execute_approval(request_index_0based, action, admin_name="Manager", rejection_note="-"):
    """
    Eksekusi Approval dengan perbaikan Logging agar kolom Detail & Chat terisi lengkap.
    """
    try:
        ws_pending = init_pending_db()
        if not ws_pending:
            return False, "DB Error: Sheet Pending tidak ditemukan."

        all_requests = ws_pending.get_all_records()
        if request_index_0based >= len(all_requests):
            return False, "Data tidak ditemukan (mungkin sudah diproses)."

        req = all_requests[request_index_0based]
        target_sheet_name = req["Target Sheet"]
        row_target_idx = int(req["Row Index (0-based)"])
        requestor_name = req.get("Requestor", "Unknown")

        # --- [FIX] MENYUSUN ULANG DETAIL PERUBAHAN (DIFF) ---
        # Kita baca ulang JSON dari request pending agar log Manager memiliki detail data
        diff_str_log = ""
        try:
            raw_old = req.get("Old Data JSON", "{}")
            raw_new = req.get("New Data JSON", "{}")
            old_d = json.loads(raw_old) if raw_old else {}
            new_d = json.loads(raw_new) if raw_new else {}

            diff_list = []
            for k, v in new_d.items():
                old_v = old_d.get(k, "")
                if str(old_v).strip() != str(v).strip():
                    diff_list.append(f"• {k}: '{old_v}' ➡ '{v}'")

            diff_str_log = "\n".join(
                diff_list) if diff_list else "Re-save (Tanpa Perubahan Nilai)."
        except:
            diff_str_log = "Detail perubahan tidak terbaca."

        # --- ACTION: REJECT (DITOLAK) ---
        if action == "REJECT":
            final_reason = str(rejection_note).strip()
            # Jika admin tidak menulis alasan, beri default
            if not final_reason or final_reason in ["-", ""]:
                final_reason = "Data perlu direvisi."

            # LOG BARU: Menggunakan parameter 'chat_msg' dan 'details_input'
            force_audit_log(
                actor=admin_name,
                action="❌ DITOLAK",
                target_sheet=target_sheet_name,
                chat_msg=f"⛔ [MANAGER]: {final_reason}",
                details_input=f"Pengaju: {requestor_name}\n(Data dikembalikan ke Admin)"
            )

            # Hapus dari daftar pending karena sudah diproses
            ws_pending.delete_rows(request_index_0based + 2)
            return True, f"DITOLAK. Alasan: {final_reason}"

        # --- ACTION: APPROVE (DI ACC) ---
        elif action == "APPROVE":
            # 1. EKSEKUSI UPDATE DATA KE SHEET TARGET (Logika Asli)
            new_data_dict = json.loads(req["New Data JSON"])
            ws_target = spreadsheet.worksheet(target_sheet_name)
            headers = ws_target.row_values(1)

            # Mapping data baru sesuai urutan header di sheet target
            row_values = [new_data_dict.get(h, "") for h in headers]

            # Update baris di Google Sheet (Write)
            gsheet_row = row_target_idx + 2
            ws_target.update(range_name=f"A{gsheet_row}", values=[
                             row_values], value_input_option="USER_ENTERED")

            # 2. LOG BARU: Mencatat Sukses dengan detail perubahan
            force_audit_log(
                actor=admin_name,
                action="✅ SUKSES/ACC",
                target_sheet=target_sheet_name,
                chat_msg="✅ [MANAGER]: Disetujui & Data Terupdate.",
                details_input=f"Pengaju: {requestor_name}\n---\n{diff_str_log}"
            )

            # Hapus dari daftar pending karena sudah diproses
            ws_pending.delete_rows(request_index_0based + 2)
            return True, "DISETUJUI & Database Terupdate."

    except Exception as e:
        return False, f"System Error: {e}"


# --- BAGIAN IMPORT OPTIONAL LIBS JANGAN DIHAPUS (Excel/AgGrid/Plotly) ---
# Bagian ini dipertahankan dari Code Pertama untuk menjaga kompatibilitas arsitektur
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
    HAS_AGGRID = True
except ImportError:
    HAS_AGGRID = False

try:
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False


# =========================================================
# PAGE CONFIG
# =========================================================
APP_TITLE = "Sales & Marketing Action Center"
st.set_page_config(
    page_title=APP_TITLE,
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =========================================================
# SYSTEM LOGIN OTP VIA EMAIL
# =========================================================


def send_email_otp(target_email, otp_code):
    """Mengirim kode OTP ke email target menggunakan SMTP Gmail"""
    smtp_config = st.secrets["smtp"]
    sender_email = smtp_config["sender_email"]
    sender_password = smtp_config["sender_password"]

    subject = "Kode Login - Sales Action Center"
    body = f"""
    <html>
      <body style="font-family: Arial, sans-serif;">
        <h2 style="color: #2e7d32;">Sales & Marketing Action Center</h2>
        <p>Halo,</p>
        <p>Gunakan kode berikut untuk masuk ke aplikasi:</p>
        <div style="background-color: #f1f8e9; padding: 15px; border-radius: 8px; display: inline-block;">
            <h1 style="color: #1b5e20; letter-spacing: 5px; margin: 0;">{otp_code}</h1>
        </div>
        <p>Kode ini berlaku untuk satu kali login. Jangan berikan kepada siapapun.</p>
        <hr>
        <small>Pesan otomatis dari Sistem Laporan Harian.</small>
      </body>
    </html>
    """

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = target_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))

    try:
        # Menggunakan SSL (Port 465)
        with smtplib.SMTP_SSL(smtp_config["smtp_server"], smtp_config["smtp_port"]) as server:
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, target_email, msg.as_string())
        return True
    except Exception as e:
        st.error(f"Gagal kirim email: {e}")
        return False


def generate_otp():
    return ''.join(random.choices(string.digits, k=6))

# =========================================================
# SYSTEM LOGIN (MODIFIED: Direct Staff Access)
# =========================================================

# =========================================================
# SYSTEM LOGIN (MODIFIED: Direct Staff Access & Role Check)
# =========================================================


def login_page():
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("<h1 style='text-align: center;'>🔐 Access Portal</h1>",
                unsafe_allow_html=True)
    st.markdown(
        f"<p style='text-align: center;'>{APP_TITLE}</p>", unsafe_allow_html=True)
    st.divider()

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        # MEMBUAT TABS: Staff (Langsung) vs Admin (OTP)
        tab_staff, tab_admin = st.tabs(["🚀 Akses Staff", "🛡️ Login Admin"])

        # --- TAB 1: AKSES STAFF (LANGSUNG) ---
        with tab_staff:
            st.markdown("### 👋 Halo, Team!")
            st.info("Klik tombol di bawah untuk masuk dan mulai membuat laporan.")

            if st.button("Masuk Aplikasi (Staff)", type="primary", use_container_width=True):
                # SET SESSION STAFF (GENERIC)
                st.session_state["logged_in"] = True
                st.session_state["user_email"] = "staff_entry"
                # Nama spesifik nanti dipilih di dalam form
                st.session_state["user_name"] = "Staff Member"
                st.session_state["user_role"] = "staff"
                # KUNCI: Staff tidak bisa akses dashboard admin
                st.session_state["is_admin"] = False

                st.success("Berhasil masuk! Mengalihkan...")
                time.sleep(0.5)
                st.rerun()

        # --- TAB 2: LOGIN ADMIN (EMAIL & OTP) ---
        with tab_admin:
            # Step 1: Input Email
            if st.session_state.get("otp_step", 1) == 1:
                with st.form("email_form"):
                    st.caption("Khusus Admin & Manager (via Email Terdaftar)")
                    email_input = st.text_input("Email Address")

                    if st.form_submit_button("Kirim Kode OTP", use_container_width=True):
                        # Cek whitelist di secrets.toml
                        users_db = st.secrets.get("users", {})
                        # Normalisasi email (hilangkan spasi & lowercase)
                        email_clean = email_input.strip().lower()

                        if email_clean in users_db:
                            otp = generate_otp()
                            # Kirim Email OTP
                            if send_email_otp(email_clean, otp):
                                st.session_state["generated_otp"] = otp
                                st.session_state["temp_email"] = email_clean
                                st.session_state["otp_step"] = 2
                                st.success("OTP Terkirim ke email!")
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error(
                                    "Gagal kirim email (Cek Config SMTP).")
                        else:
                            st.error(
                                "⛔ Akses Ditolak: Email tidak terdaftar sebagai Admin/Manager.")

            # Step 2: Input OTP
            elif st.session_state.get("otp_step") == 2:
                st.info(
                    f"Kode dikirim ke: **{st.session_state['temp_email']}**")

                with st.form("otp_form"):
                    otp_input = st.text_input(
                        "Kode OTP (6 Digit)", max_chars=6)
                    c_back, c_ok = st.columns(2)

                    # Tombol Kembali
                    if c_back.form_submit_button("⬅️ Ganti Email"):
                        st.session_state["otp_step"] = 1
                        st.rerun()

                    # Tombol Verifikasi
                    if c_ok.form_submit_button("Verifikasi ✅", type="primary"):
                        if otp_input == st.session_state["generated_otp"]:
                            # LOGIN ADMIN/MANAGER SUKSES
                            email_fix = st.session_state["temp_email"]
                            user_info = st.secrets["users"][email_fix]

                            st.session_state["logged_in"] = True
                            st.session_state["user_email"] = email_fix
                            st.session_state["user_name"] = user_info["name"]

                            # --- DETEKSI ROLE ---
                            # Ambil role dari secrets.toml, default ke 'staff' jika tidak ada
                            role_str = str(user_info.get(
                                "role", "staff")).lower()
                            st.session_state["user_role"] = role_str

                            # Tentukan Flag Admin: True jika role adalah 'admin' ATAU 'manager'
                            # Ini memberikan akses ke menu Dashboard Admin untuk kedua role tersebut
                            if role_str in ["admin", "manager"]:
                                st.session_state["is_admin"] = True
                            else:
                                st.session_state["is_admin"] = False

                            st.success(
                                f"Login Berhasil! Selamat datang {role_str.upper()}.")
                            time.sleep(0.5)
                            st.rerun()
                        else:
                            st.error("Kode OTP Salah.")


# =========================================================
# HELPER: DATABASE STAFF (GOOGLE SHEET)
# =========================================================
SHEET_USERS = "Config_Users"


def init_user_db():
    """Memastikan sheet user ada."""
    try:
        try:
            ws = spreadsheet.worksheet(SHEET_USERS)
        except gspread.WorksheetNotFound:
            ws = spreadsheet.add_worksheet(title=SHEET_USERS, rows=100, cols=4)
            ws.append_row(["Username", "Password", "Nama", "Role"],
                          value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
        return ws
    except Exception:
        return None


def check_staff_login(username, password):
    """Cek login untuk staff biasa via GSheet."""
    ws = init_user_db()
    if not ws:
        return None

    # Ambil semua data
    records = ws.get_all_records()
    for user in records:
        u_db = str(user.get("Username", "")).strip()
        p_db = str(user.get("Password", "")).strip()
        if u_db == username and p_db == password:
            return user
    return None


def add_staff_account(username, password, nama):
    """Admin menambah akun staff."""
    ws = init_user_db()
    if not ws:
        return False, "DB Error"

    existing_users = ws.col_values(1)
    if username in existing_users:
        return False, "Username sudah dipakai!"

    ws.append_row([username, password, nama, "staff"],
                  value_input_option="USER_ENTERED")
    return True, "Akun berhasil dibuat."


def update_staff_account(username_lama, new_password=None, new_name=None):
    """Fitur Edit Akun Staff (Ganti Password / Nama)."""
    ws = init_user_db()
    if not ws:
        return False, "DB Error"

    try:
        cell = ws.find(username_lama)
        row = cell.row
        if new_password and new_password.strip():
            ws.update_cell(row, 2, new_password)  # Kolom 2 = Password
        if new_name and new_name.strip():
            ws.update_cell(row, 3, new_name)  # Kolom 3 = Nama
        return True, f"Data user {username_lama} berhasil diperbarui."
    except gspread.exceptions.CellNotFound:
        return False, "Username tidak ditemukan."
    except Exception as e:
        return False, str(e)


def delete_staff_account(username):
    """Admin menghapus akun staff."""
    ws = init_user_db()
    if not ws:
        return False, "DB Error"

    try:
        cell = ws.find(username)
        ws.delete_rows(cell.row)
        return True, f"User {username} dihapus."
    except gspread.exceptions.CellNotFound:
        return False, "Username tidak ditemukan."


# =========================================================
# MAIN FLOW CHECK
# =========================================================
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

if not st.session_state["logged_in"]:
    login_page()
    st.stop()  # Berhenti disini jika belum login

# =========================================================
# USER INFO SETELAH LOGIN (Variabel Global)
# =========================================================
# Variabel ini akan dipakai di seluruh aplikasi
user_email = st.session_state["user_email"]
user_name = st.session_state["user_name"]
user_role = st.session_state["user_role"]

# =========================================================
# OPTIONAL LIBS (Excel Export / AgGrid / Plotly)
# =========================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
    HAS_AGGRID = True
except ImportError:
    HAS_AGGRID = False

try:
    import plotly.express as px
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False


# =========================================================
# PAGE CONFIG
# =========================================================
APP_TITLE = "Sales & Marketing Action Center"
st.set_page_config(
    page_title=APP_TITLE,
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded"
)

# =========================================================
# GLOBAL STYLE (SpaceX x Muhammadiyah — Elegant, International)
# =========================================================


def inject_global_css():
    st.markdown(
        """
        <style>
        @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600;700;800&display=swap');

        :root{
            --bg0:#020805;
            --bg1:#04110b;
            --bg2:#062015;

            --cardA: rgba(255,255,255,0.06);
            --cardB: rgba(255,255,255,0.045);
            --border: rgba(255,255,255,0.10);

            --text: rgba(255,255,255,0.92);
            --muted: rgba(255,255,255,0.70);

            --green:#16a34a;
            --green2:#22c55e;
            --teal:#14b8a6;
            --gold:#facc15;
            --amber:#f59e0b;
            --danger:#ef4444;

            /* Beri tahu browser bahwa UI ini dark theme */
            color-scheme: dark;
        }

        /* ---------- App background ---------- */
        .stApp {
            background:
                radial-gradient(circle at 14% 12%, rgba(22, 163, 74, 0.20) 0%, rgba(22, 163, 74, 0.0) 46%),
                radial-gradient(circle at 84% 14%, rgba(250, 204, 21, 0.16) 0%, rgba(250, 204, 21, 0.0) 42%),
                radial-gradient(circle at 18% 92%, rgba(20, 184, 166, 0.12) 0%, rgba(20, 184, 166, 0.0) 40%),
                linear-gradient(180deg, var(--bg0) 0%, var(--bg1) 55%, var(--bg2) 100%);
            color: var(--text);
        }

        /* Subtle starfield overlay (Space vibe) */
        .stApp::before {
            content: "";
            position: fixed;
            inset: 0;
            pointer-events: none;
            background:
                radial-gradient(rgba(255,255,255,0.18) 0.8px, transparent 0.8px);
            background-size: 68px 68px;
            opacity: 0.10;
            -webkit-mask-image: radial-gradient(circle at 50% 15%, rgba(0,0,0,1) 0%, rgba(0,0,0,0.0) 70%);
            mask-image: radial-gradient(circle at 50% 15%, rgba(0,0,0,1) 0%, rgba(0,0,0,0.0) 70%);
        }

        /* Hide Streamlit default UI chrome (we use custom header) */
        #MainMenu {visibility: hidden;}
        footer {visibility: hidden;}
        header {visibility: hidden;}

        /* Typography */
        h1, h2, h3, h4, h5, h6, p, label, span, div {
            font-family: "Space Grotesk", ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial, "Helvetica Neue", "Noto Sans", "Liberation Sans", sans-serif;
        }

        /* =========================
        Text selection (blok teks)
        ========================= */
        .stApp ::selection{
            color: #ffffff !important;
            background: rgba(22,163,74,0.35) !important;
        }
        .stApp ::-moz-selection{
            color: #ffffff !important;
            background: rgba(22,163,74,0.35) !important;
        }

        /* Sidebar polish (SpaceX-like) */
        section[data-testid="stSidebar"] > div {
            background: linear-gradient(180deg, rgba(0,0,0,0.92) 0%, rgba(3,10,6,0.92) 60%, rgba(4,16,11,0.92) 100%);
            border-right: 1px solid rgba(255,255,255,0.10);
        }
        section[data-testid="stSidebar"] * {
            color: var(--text) !important;
        }
        section[data-testid="stSidebar"] hr {
            border-color: rgba(255,255,255,0.10);
        }

        /* Card styling for containers with border=True */
        div[data-testid="stVerticalBlockBorderWrapper"] > div {
            background: linear-gradient(180deg, var(--cardA) 0%, var(--cardB) 100%);
            border: 1px solid var(--border);
            border-radius: 18px;
            padding: 1.05rem 1.05rem 0.75rem 1.05rem;
            box-shadow: 0 16px 46px rgba(0,0,0,0.42);
            backdrop-filter: blur(10px);
        }

        /* Buttons */
        .stButton>button, .stDownloadButton>button {
            border-radius: 12px !important;
            border: 1px solid rgba(255,255,255,0.14) !important;
            background: rgba(255,255,255,0.05) !important;
            color: var(--text) !important;
            transition: all 0.15s ease-in-out;
        }
        .stButton>button:hover, .stDownloadButton>button:hover {
            transform: translateY(-1px);
            border-color: rgba(250,204,21,0.35) !important;
            background: rgba(255,255,255,0.08) !important;
        }

        /* Primary button (type=primary) */
        button[kind="primary"] {
            background: linear-gradient(135deg, rgba(22,163,74,0.95), rgba(245,158,11,0.92)) !important;
            color: rgba(6, 26, 17, 0.95) !important;
            border: none !important;
        }
        button[kind="primary"]:hover {
            filter: brightness(1.05);
        }

        /* Inputs */
        .stTextInput input, .stTextArea textarea, .stNumberInput input {
            border-radius: 12px !important;
        }
        .stDateInput input {
            border-radius: 12px !important;
        }
        .stSelectbox div[data-baseweb="select"] > div {
            border-radius: 12px !important;
        }

        /* Dataframes / tables */
        div[data-testid="stDataFrame"] {
            border-radius: 14px;
            overflow: hidden;
            border: 1px solid rgba(255,255,255,0.10);
        }

        /* =========================
           HERO HEADER (Custom)
           ========================= */
        .sx-hero{
            position: relative;
            border-radius: 20px;
            border: 1px solid rgba(255,255,255,0.12);
            overflow: hidden;
            padding: 18px 18px;
            background:
                radial-gradient(circle at 50% 0%, rgba(255,255,255,0.08) 0%, rgba(255,255,255,0.0) 52%),
                linear-gradient(90deg, rgba(0,0,0,0.55) 0%, rgba(0,0,0,0.25) 50%, rgba(0,0,0,0.55) 100%);
            box-shadow: 0 18px 60px rgba(0,0,0,0.45);
        }
        .sx-hero::before{
            content:"";
            position:absolute;
            inset:0;
            background-image: var(--hero-bg);
            background-repeat:no-repeat;
            background-position: var(--hero-bg-pos, 50% 72%);
            background-size: var(--hero-bg-size, 140%);
            opacity: 0.28;
            filter: saturate(1.05) contrast(1.08);
            pointer-events:none;
        }
        .sx-holding-logo{
            display:block;
            margin: 0 auto 10px auto;
            width: clamp(90px, 10vw, 140px);
            height: auto;
            opacity: 0.95;
            filter: drop-shadow(0 10px 22px rgba(0,0,0,0.45));
        }
        .sx-hero::after{
            content:"";
            position:absolute;
            inset:0;
            background:
                linear-gradient(180deg, rgba(2,8,5,0.15) 0%, rgba(2,8,5,0.52) 100%);
            pointer-events:none;
        }

        .sx-hero-grid{
            position: relative;
            display: grid;
            grid-template-columns: 240px 1fr 240px;
            align-items: center;
            gap: 14px;
        }

        .sx-hero-grid > * { min-width: 0; }

        @media (max-width: 1100px){
            .sx-hero-grid{ grid-template-columns: 200px 1fr 200px; }
        }
        @media (max-width: 860px){
            .sx-hero-grid{ grid-template-columns: 1fr; text-align:center; }
        }

        *, *::before, *::after { box-sizing: border-box; }

        .sx-logo-card{
            background: rgba(255,255,255,0.92);
            border: 1px solid rgba(0,0,0,0.06);
            border-radius: 16px;
            width: 100%;
            max-width: 240px;
            height: clamp(120px, 12vw, 160px);
            padding: 10px;
            display:flex;
            align-items:center;
            justify-content:center;
            box-shadow: 0 10px 26px rgba(0,0,0,0.28);
        }

        .sx-logo-card img{
            width: 100%;
            height: 100%;
            max-width: 220px;
            max-height: 100%;
            object-fit: contain;
            object-position: center;
            display: block;
        }

        .sx-hero-center{
            text-align: center;
        }
        .sx-title{
            font-size: 2.05rem;
            font-weight: 800;
            line-height: 1.12;
            letter-spacing: 0.06em;
            text-transform: uppercase;
            margin: 0;
        }
        .sx-subrow{
            margin-top: 0.45rem;
            display:flex;
            gap: 0.55rem;
            flex-wrap: wrap;
            justify-content: center;
            align-items: center;
            color: rgba(255,255,255,0.78);
            font-size: 0.95rem;
        }
        .sx-pill{
            display:inline-flex;
            align-items:center;
            gap: 0.35rem;
            padding: 0.22rem 0.60rem;
            border-radius: 999px;
            border: 1px solid rgba(255,255,255,0.14);
            background: rgba(255,255,255,0.06);
            color: rgba(255,255,255,0.88);
            font-size: 0.80rem;
        }
        .sx-pill.on{
            border-color: rgba(34,197,94,0.55);
            box-shadow: 0 0 0 2px rgba(34,197,94,0.10) inset;
        }
        .sx-pill.off{
            border-color: rgba(239,68,68,0.55);
            box-shadow: 0 0 0 2px rgba(239,68,68,0.10) inset;
        }
        .sx-dot{
            width: 8px; height: 8px; border-radius: 999px; display:inline-block;
            background: rgba(255,255,255,0.55);
        }
        .sx-pill.on .sx-dot{ background: rgba(34,197,94,0.95); }
        .sx-pill.off .sx-dot{ background: rgba(239,68,68,0.95); }

        /* =========================
           Sidebar Nav (SpaceX-like)
           ========================= */
        .sx-nav{
            margin-top: 0.25rem;
        }
        .sx-nav button{
            width: 100% !important;
            text-align: left !important;
            border-radius: 12px !important;
            padding: 0.60rem 0.80rem !important;
            text-transform: uppercase !important;
            letter-spacing: 0.10em !important;
            font-size: 0.78rem !important;
        }
        .sx-nav button[kind="primary"]{
            background: linear-gradient(90deg, rgba(22,163,74,0.95), rgba(245,158,11,0.90)) !important;
            color: rgba(6,26,17,0.95) !important;
        }

        .sx-section-title{
            font-size: 0.82rem;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            color: rgba(255,255,255,0.70);
        }

        /* ==================================================
           MOBILE ONLY (<=768px) - tidak mengubah desktop
           ================================================== */
        @media (max-width: 768px){
          /* Sidebar disembunyikan di HP */
          section[data-testid="stSidebar"] { display: none !important; }

          /* Padding konten + ruang untuk bottom nav */
          .block-container {
            padding-left: 1rem !important;
            padding-right: 1rem !important;
            padding-bottom: 80px !important; /* biar konten tidak ketutup bottom nav */
          }

          /* Hero dibuat lebih ringkas */
          .sx-title { font-size: 1.35rem !important; }
          .sx-hero-grid { grid-template-columns: 1fr !important; }

          /* Logo kiri/kanan dimatikan di HP biar tidak makan tempat */
          .sx-logo-card { display:none !important; }

          .mobile-bottom-nav{
            position: fixed;
            left: 0; right: 0; bottom: 0;
            padding: 10px 12px;
            background: rgba(0,0,0,0.75);
            border-top: 1px solid rgba(255,255,255,0.12);
            display: flex;
            justify-content: space-around;
            gap: 8px;
            z-index: 9999;
            backdrop-filter: blur(10px);
          }
          .mobile-bottom-nav a{
            text-decoration:none;
            color: rgba(255,255,255,0.92);
            padding: 8px 10px;
            border-radius: 12px;
            border: 1px solid rgba(255,255,255,0.12);
            background: rgba(255,255,255,0.06);
            font-size: 14px;
          }

          /* Kurangi efek blur di HP (card form Closing Deal) */
          div[data-testid="stVerticalBlockBorderWrapper"] > div {
            backdrop-filter: none !important;
            background: linear-gradient(
                180deg,
                rgba(6, 36, 22, 0.96),
                rgba(5, 25, 17, 0.98)
            ) !important;
          }
        }

        /* =========================================
           PATCH KONTRAS TEKS & LOGO (MOBILE + DESKTOP)
           ========================================= */

        /* 1. Warna label & teks kecil di dalam form Closing Deal */
        div[data-testid="stForm"] label,
        div[data-testid="stForm"] p {
            color: rgba(255, 255, 255, 0.9) !important;
        }

        /* 2. Warna teks judul field di dalam kartu form (jaga-jaga) */
        div[data-testid="stVerticalBlockBorderWrapper"] label,
        div[data-testid="stVerticalBlockBorderWrapper"] p {
            color: rgba(255, 255, 255, 0.88) !important;
        }

        /* 3. Biar icon / logo tidak nyaru di navbar / header custom */
        .sx-nav button,
        .sx-nav svg,
        .sx-nav span {
            color: #ffffff !important;
            fill: #ffffff !important;
        }

        /* =========================================
           PATCH LANJUTAN – KONTRAS TEKS DI DALAM CARD
           (Riwayat Closing, dst)
           ========================================= */

        /* Semua teks di dalam card ber-border */
        div[data-testid="stVerticalBlockBorderWrapper"],
        div[data-testid="stVerticalBlockBorderWrapper"] p,
        div[data-testid="stVerticalBlockBorderWrapper"] span,
        div[data-testid="stVerticalBlockBorderWrapper"] small,
        div[data-testid="stVerticalBlockBorderWrapper"] li {
            color: rgba(255, 255, 255, 0.90) !important;
        }

        /* Teks yang berasal dari st.markdown / st.write */
        div[data-testid="stMarkdown"],
        div[data-testid="stMarkdown"] p,
        div[data-testid="stMarkdown"] span,
        div[data-testid="stMarkdown"] li,
        div[data-testid="stMarkdown"] small,
        div[data-testid="stMarkdownContainer"],
        div[data-testid="stMarkdownContainer"] p,
        div[data-testid="stMarkdownContainer"] span,
        div[data-testid="stMarkdownContainer"] li,
        div[data-testid="stMarkdownContainer"] small {
            color: rgba(255, 255, 255, 0.90) !important;
        }

        /* =========================================
           FIX KONTRAS METRIC (Total Nilai, Overdue, dll)
           ========================================= */

        /* Container metric */
        div[data-testid="stMetric"] {
            color: var(--text) !important;
        }

        /* Label kecil di atas angka */
        div[data-testid="stMetricLabel"],
        div[data-testid="stMetric"] label {
            color: rgba(255,255,255,0.80) !important;
            font-weight: 500 !important;
        }

        /* Angka besar (nilai utama metric) */
        div[data-testid="stMetricValue"] {
            color: var(--gold) !important;  /* ganti ke var(--text) kalau mau putih */
            font-weight: 700 !important;
        }

        /* Delta metric (jika dipakai) */
        div[data-testid="stMetricDelta"] {
            color: var(--green2) !important;
            font-weight: 600 !important;
        }

        /* =========================================
           LOADING SPINNER OVERLAY (FIXED & FULLSCREEN)
           ========================================= */
        /* Container utama spinner: dibuat memenuhi satu layar penuh */
        div[data-testid="stSpinner"] {
            position: fixed !important;
            top: 0 !important;
            left: 0 !important;
            width: 100vw !important;
            height: 100vh !important;
            z-index: 999999 !important; /* Pastikan di paling depan */

            /* Background Gelap Transparan (Glassmorphism) */
            background: rgba(0, 0, 0, 0.85) !important;
            backdrop-filter: blur(8px); /* Efek blur latar belakang */

            /* Posisi konten di tengah */
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            gap: 20px;

            /* Reset style bawaan yang mengganggu */
            transform: none !important;
            border: none !important;
            box-shadow: none !important;
        }

        /* Teks pesan loading (misal: "Sedang menyimpan...") */
        div[data-testid="stSpinner"] > div {
            color: #ffffff !important;
            font-size: 1.1rem !important;
            font-weight: 500 !important;
            letter-spacing: 0.05em;
            text-shadow: 0 2px 4px rgba(0,0,0,0.5);
        }

        /* Icon Lingkaran Berputar (Spinner) */
        /* Target elemen SVG atau div lingkaran di dalam spinner */
        div[data-testid="stSpinner"] > div > div {
            border-top-color: var(--gold) !important;    /* Warna Emas */
            border-right-color: var(--green) !important; /* Warna Hijau */
            border-bottom-color: var(--gold) !important; /* Warna Emas */
            border-left-color: transparent !important;
            width: 3.5rem !important;  /* Ukuran icon lebih besar */
            height: 3.5rem !important;
            border-width: 4px !important; /* Ketebalan garis */
        }

        </style>
        """,
        unsafe_allow_html=True
    )


inject_global_css()


# =========================================================
# COMPAT HELPERS (toast / link button)
# =========================================================
def ui_toast(message: str, icon=None):
    """Streamlit toast (fallback ke success jika tidak tersedia)."""
    if hasattr(st, "toast"):
        try:
            st.toast(message, icon=icon)
            return
        except Exception:
            pass
    st.success(message)


# =========================================================
# CONSTANTS
# =========================================================
NAMA_GOOGLE_SHEET = "Laporan Kegiatan Harian"
FOLDER_DROPBOX = "/Laporan_Kegiatan_Harian"

# Sheet Names
SHEET_CONFIG_NAMA = "Config_Staf"
SHEET_TARGET_TEAM = "Target_Team_Checklist"
SHEET_TARGET_INDIVIDU = "Target_Individu_Checklist"
SHEET_CONFIG_TEAM = "Config_Team"
SHEET_CLOSING_DEAL = "Closing_Deal"
SHEET_PEMBAYARAN = "Pembayaran_DP"
SHEET_PRESENSI = "Presensi_Kehadiran"
PRESENSI_COLUMNS = ["Timestamp", "Nama", "Tipe Absen", "Hari",
                    "Tanggal", "Bulan", "Tahun", "Waktu","Link Foto"]


def init_presensi_db():
    """Memastikan sheet presensi tersedia dan header sesuai standar terbaru (termasuk Link Foto)."""
    try:
        try:
            # 1. Mencoba membuka worksheet yang sudah ada
            ws = spreadsheet.worksheet(SHEET_PRESENSI)
        except gspread.WorksheetNotFound:
            # 2. Jika worksheet benar-benar belum ada, buat baru
            ws = spreadsheet.add_worksheet(
                title=SHEET_PRESENSI, 
                rows=2000, 
                cols=len(PRESENSI_COLUMNS)
            )
            # Masukkan header awal untuk pertama kali
            ws.append_row(PRESENSI_COLUMNS, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
        
        # 3. KUNCI REVISI: Pastikan kolom selalu sinkron. 
        # Jika Anda baru saja menambah "Link Foto" ke PRESENSI_COLUMNS, 
        # fungsi ini akan otomatis menambahkannya di Google Sheets Anda.
        ensure_headers(ws, PRESENSI_COLUMNS)
        
        return ws
    except Exception as e:
        # Menampilkan log error di terminal agar mudah dilacak jika koneksi bermasalah
        print(f"Error init_presensi_db: {e}")
        return None


def catat_presensi(nama_staf, tipe="Masuk", file_foto=None):
    """
    Logika utama presensi terpadu: 
    Mendukung tipe Masuk/Pulang, Real-time, Upload Dropbox, dan Validasi Alur.
    """
    ws = init_presensi_db()
    if not ws:
        return False, "Database Presensi Error"

    # 1. Ambil Waktu Real-Time (WIB)
    now = datetime.now(TZ_JKT)

    # Mapping Hari Indonesia (Tetap sesuai flow lama Anda)
    hari_map = {
        "Monday": "Senin", "Tuesday": "Selasa", "Wednesday": "Rabu",
        "Thursday": "Kamis", "Friday": "Jumat", "Saturday": "Sabtu", "Sunday": "Minggu"
    }

    ts_full = now.strftime("%d-%m-%Y %H:%M:%S")
    hari = hari_map.get(now.strftime("%A"), now.strftime("%A"))
    tanggal = now.strftime("%d")
    bulan = now.strftime("%B")
    tahun = now.strftime("%Y")
    waktu = now.strftime("%H:%M:%S")
    today_str = now.strftime("%d-%m-%Y")

    # 2. VALIDASI PINTAR (Mencegah double masuk/pulang & urutan yang salah)
    records = ws.get_all_records()
    sudah_masuk = False
    sudah_pulang = False

    for r in records:
        # Cek apakah nama sama dan tanggal hari ini ada di kolom Timestamp
        if str(r.get("Nama")) == nama_staf and today_str in str(r.get("Timestamp")):
            tipe_record = str(r.get("Tipe Absen", "")).strip()
            if tipe_record == "Masuk":
                sudah_masuk = True
            if tipe_record == "Pulang":
                sudah_pulang = True

    # Logika pencegahan
    if tipe == "Masuk":
        if sudah_masuk:
            return False, f"Anda sudah melakukan Presensi Masuk hari ini pada sistem."
    
    elif tipe == "Pulang":
        if not sudah_masuk:
            return False, "Gagal: Anda belum melakukan Presensi Masuk. Silakan Masuk terlebih dahulu."
        if sudah_pulang:
            return False, "Anda sudah melakukan Presensi Pulang hari ini."

    # 3. Handle Upload Foto ke Dropbox
    link_foto = "-" 
    if file_foto is not None:
        if KONEKSI_DROPBOX_BERHASIL:
            # Folder kategori dinamis sesuai tipe: Presensi_Masuk atau Presensi_Pulang
            url = upload_ke_dropbox(file_foto, nama_staf, kategori=f"Presensi_{tipe}")
            if url and url != "-":
                link_foto = url
        else:
            return False, "Gagal simpan: Koneksi Dropbox bermasalah (Foto wajib untuk presensi)."

    # 4. Masukkan Data ke GSheet (Update: Menambahkan kolom 'tipe' di posisi ketiga)
    # Urutan Kolom: Timestamp, Nama, Tipe Absen, Hari, Tanggal, Bulan, Tahun, Waktu, Link Foto
    row = [
        f"'{ts_full}", 
        nama_staf, 
        tipe,        # Masuk ke kolom Tipe Absen
        hari, 
        tanggal, 
        bulan, 
        tahun, 
        waktu, 
        link_foto
    ]
    
    ws.append_row(row, value_input_option="USER_ENTERED")

    return True, f"Berhasil! Presensi {tipe} tercatat pukul {waktu} WIB."


# Kolom laporan harian
COL_TIMESTAMP = "Timestamp"
COL_NAMA = "Nama"
COL_TEMPAT = "Tempat Dikunjungi"
COL_DESKRIPSI = "Deskripsi"
COL_LINK_FOTO = "Link Foto"
COL_LINK_SOSMED = "Link Sosmed"
COL_KESIMPULAN = "Kesimpulan"
COL_KENDALA = "Kendala"
COL_KENDALA_KLIEN = "Kendala Klien"
COL_PENDING = "Next Plan (Pending)"
COL_FEEDBACK = "Feedback Lead"
COL_INTEREST = "Interest (%)"
COL_NAMA_KLIEN = "Nama Klien"
COL_KONTAK_KLIEN = "No HP/WA"

NAMA_KOLOM_STANDAR = [
    COL_TIMESTAMP, COL_NAMA, COL_TEMPAT, COL_DESKRIPSI,
    COL_LINK_FOTO, COL_LINK_SOSMED,
    COL_KESIMPULAN, COL_KENDALA, COL_KENDALA_KLIEN,
    COL_PENDING,
    COL_FEEDBACK,
    COL_INTEREST,
    COL_NAMA_KLIEN,
    COL_KONTAK_KLIEN
]

# Audit columns
COL_TS_UPDATE = "Timestamp Update (Log)"
COL_UPDATED_BY = "Updated By"

# Team config columns
TEAM_COL_NAMA_TEAM = "Nama Team"
TEAM_COL_POSISI = "Posisi"
TEAM_COL_ANGGOTA = "Nama Anggota"
TEAM_COLUMNS = [TEAM_COL_NAMA_TEAM, TEAM_COL_POSISI, TEAM_COL_ANGGOTA]

# =========================================================
# DEFINISI KOLOM DATA (DATABASE STRUCTURE)
# =========================================================

# --- 1. Closing Deal Columns ---
COL_GROUP = "Nama Group"
COL_MARKETING = "Nama Marketing"
COL_TGL_EVENT = "Tanggal Event"
COL_BIDANG = "Bidang"
COL_NILAI_KONTRAK = "Nilai Kontrak"  # disimpan sebagai angka (int)

CLOSING_COLUMNS = [
    COL_GROUP, 
    COL_MARKETING, 
    COL_TGL_EVENT, 
    COL_BIDANG, 
    COL_NILAI_KONTRAK
]

# --- 2. Target / Checklist Columns ---
TEAM_CHECKLIST_COLUMNS = [
    "Misi", "Tgl_Mulai", "Tgl_Selesai", "Status", 
    "Bukti/Catatan", COL_TS_UPDATE, COL_UPDATED_BY
]
INDIV_CHECKLIST_COLUMNS = [
    "Nama", "Target", "Tgl_Mulai", "Tgl_Selesai", "Status", 
    "Bukti/Catatan", COL_TS_UPDATE, COL_UPDATED_BY
]

# --- 3. Smart Pembayaran Columns (Update Khusus) ---
COL_TS_BAYAR = "Timestamp Input"
COL_NILAI_KESEPAKATAN = "Total Nilai Kesepakatan" # [NEW]
COL_JENIS_BAYAR = "Jenis Pembayaran"             # DP, Cicilan, atau Cash
COL_NOMINAL_BAYAR = "Nominal Pembayaran"         # Nominal yang masuk saat ini
COL_TENOR_CICILAN = "Tenor (Bulan)"              # [NEW]
COL_SISA_BAYAR = "Sisa Pembayaran"               # [NEW] Kalkulator Otomatis
COL_JATUH_TEMPO = "Batas Waktu Bayar"
COL_STATUS_BAYAR = "Status Pembayaran"           # Deskripsi status (Lunas/Belum)
COL_BUKTI_BAYAR = "Bukti Pembayaran"
COL_CATATAN_BAYAR = "Catatan"

PAYMENT_COLUMNS = [
    COL_TS_BAYAR,
    COL_GROUP,
    COL_MARKETING,
    COL_TGL_EVENT,
    COL_NILAI_KESEPAKATAN, # Letakkan total di awal agar alur logika jelas
    COL_JENIS_BAYAR,
    COL_NOMINAL_BAYAR,
    COL_TENOR_CICILAN,
    COL_SISA_BAYAR,
    COL_JATUH_TEMPO,
    COL_STATUS_BAYAR,
    COL_BUKTI_BAYAR,
    COL_CATATAN_BAYAR,
    COL_TS_UPDATE,
    COL_UPDATED_BY
]

# --- 4. System Config ---
TZ_JKT = ZoneInfo("Asia/Jakarta")

# Formatting throttling (menghindari API overload saat batch update)
FORMAT_THROTTLE_SECONDS = 300  # 5 minutes

# =========================================================
# MOBILE DETECTION (safe, tidak mengubah desktop)
# =========================================================


def is_mobile_device() -> bool:
    """
    Deteksi via User-Agent. Hanya dipakai untuk membedakan UI HP vs Desktop.
    Jika st.context tidak tersedia, fallback = False (anggap desktop).
    """
    try:
        ua = ""
        if hasattr(st, "context") and hasattr(st.context, "headers"):
            headers = st.context.headers
            ua = (headers.get("user-agent")
                  or headers.get("User-Agent") or "").lower()
        return any(k in ua for k in ["android", "iphone", "ipad", "mobile"])
    except Exception:
        return False


IS_MOBILE = is_mobile_device()


# =========================================================
# SMALL HELPERS
# =========================================================
def now_ts_str() -> str:
    """Timestamp akurat (WIB) untuk semua perubahan."""
    return datetime.now(tz=TZ_JKT).strftime("%d-%m-%Y %H:%M:%S")

# =========================================================
# [MIGRASI] PEMBAYARAN LOGIC HELPERS
# =========================================================
def parse_payment_log_lines(log_text: str):
    log_text = safe_str(log_text, "").strip()
    if not log_text: return []
    raw_lines = [ln.rstrip() for ln in log_text.splitlines() if ln.strip()]
    out = []
    for ln in raw_lines:
        mnum = re.match(r"^\s*\d+\.\s*(.*)$", ln)
        if mnum: ln = mnum.group(1).rstrip()
        m = re.match(r"^\[(.*?)\]\s*\((.*?)\)\s*(.*)$", ln)
        if m:
            ts, actor, rest = m.group(1).strip(), m.group(2).strip(), m.group(3).strip()
            prefix = f"[{ts}] ({actor})"
            if rest:
                parts = [p.strip() for p in rest.split(";") if p.strip()]
                if parts:
                    out.append(f"{prefix} {parts[0]}")
                    for p in parts[1:]: out.append(f" {p}")
                else: out.append(prefix)
            else: out.append(prefix)
        else: out.append(ln)
    return out

def build_numbered_log(lines):
    lines = [str(l).rstrip() for l in (lines or []) if safe_str(l, "").strip()]
    return "\n".join([f"{i}. {line}" for i, line in enumerate(lines, 1)]).strip()

def _fmt_payment_val_for_log(col_name: str, v):
    if col_name == COL_NOMINAL_BAYAR:
        x = parse_rupiah_to_int(v)
        return format_rupiah_display(x) if x is not None else "-"
    if col_name == COL_STATUS_BAYAR:
        return "✅ Dibayar" if normalize_bool(v) else "⏳ Belum"
    if col_name in {COL_JATUH_TEMPO, COL_TGL_EVENT}:
        d = normalize_date(v)
        return d.strftime("%Y-%m-%d") if d else "-"
    s = safe_str(v, "-").replace("\n", " ").strip()
    return s if s else "-"

def append_payment_ts_update(existing_log: str, ts: str, actor: str, changes):
    lines = parse_payment_log_lines(existing_log)
    changes = [safe_str(c, "").strip() for c in (changes or []) if safe_str(c, "").strip()]
    if not changes: return build_numbered_log(lines)
    actor, ts = (safe_str(actor, "-").strip() or "-"), (safe_str(ts, now_ts_str()).strip() or now_ts_str())
    lines.append(f"[{ts}] ({actor}) {changes[0]}")
    for c in changes[1:]: lines.append(f" {c}")
    return build_numbered_log(lines)


def safe_str(x, default="") -> str:
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return default
        s = str(x)
        if s.lower() in {"nan", "none"}:
            return default
        return s
    except Exception:
        return default


def normalize_bool(x) -> bool:
    if isinstance(x, bool):
        return x
    s = safe_str(x, "").strip().upper()
    return True if s == "TRUE" else False


def normalize_date(x):
    """Return datetime.date or None."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return None
    if isinstance(x, date) and not isinstance(x, datetime):
        return x
    try:
        return pd.to_datetime(x, errors="coerce").date()
    except Exception:
        return None


def get_actor_fallback(default="-") -> str:
    """
    Ambil 'actor' (siapa yang mengedit) dari session_state yang tersedia.
    Jika tidak ada, fallback ke default.
    """
    for k in ["pelapor_main", "sidebar_user", "payment_editor_name"]:
        if k in st.session_state and safe_str(st.session_state.get(k), "").strip():
            return safe_str(st.session_state.get(k)).strip()
    return default


def dynamic_column_mapper(df):
    mapping = {}
    # Tambahkan keyword yang mungkin ada di GSheet Anda
    keywords = {
        "Waktu": "Waktu",
        "Pelaku": "User",
        "User": "User",
        "Aksi": "Status",     # Menangkap "Aksi Dilakukan"
        "Status": "Status",   # Menangkap "Status"
        "Nama Data": "Target Data",
        "Target": "Target Data",
        "Sheet": "Target Data",
        "Alasan": "Chat & Catatan",
        "Chat": "Chat & Catatan",
        "Rincian": "Detail Perubahan",
        "Detail": "Detail Perubahan"
    }

    for col in df.columns:
        for key, standard_name in keywords.items():
            if key.lower() in str(col).lower():
                mapping[col] = standard_name
                break

    return df.rename(columns=mapping)

# =========================================================
# [BARU] FALLBACK INSIGHT GENERATOR (MACHINE LEARNING/STATISTIK)
# =========================================================
# =========================================================
# [FINAL REVISI] FALLBACK INSIGHT (MOTIVATIONAL & INCLUSIVE)
# =========================================================
def generate_smart_insight_fallback(df_data, total_laporan):
    """
    Menghasilkan insight otomatis dengan nada motivasi tinggi.
    Fokus: Mengangkat kerja keras tim vs kompetitor luar (Us vs Them).
    """
    try:
        if df_data.empty:
            return "Belum ada laporan masuk. Tim sedang bersiap untuk memulai pergerakan hari ini."

        # 1. Ambil Nama Tim (Inklusif - menyebut semua yang berkontribusi)
        active_names = df_data[COL_NAMA].unique().tolist()
        
        # Format nama agar rapi (A, B, dan C)
        if len(active_names) > 2:
            names_str = ", ".join(active_names[:-1]) + ", dan " + active_names[-1]
        elif len(active_names) == 2:
            names_str = " dan ".join(active_names)
        else:
            names_str = active_names[0]

        # 2. Deteksi Dominasi Kegiatan untuk Konteks
        if "Kategori_Aktivitas" not in df_data.columns:
             df_data["Kategori_Aktivitas"] = df_data[COL_TEMPAT].apply(
                lambda x: "Digital/Kantor" if any(k in str(x) for k in ["Digital", "Ads", "Konten"]) else "Kunjungan Lapangan"
            )
        top_activity = df_data["Kategori_Aktivitas"].mode()[0]

        # 3. Narasi Psikologis (Sesuai Request)
        # Paragraf 1: Fakta Data (Total & Siapa)
        p1 = (
            f"Saat ini, total produktivitas yang terekam telah mencapai **{total_laporan} laporan**. "
            f"Seluruh aktivitas ini merupakan hasil eksekusi langsung dari **semua anggota team Bapak** ({names_str}) yang bergerak sinergis."
        )

        # Paragraf 2: Perbandingan dengan Kompetitor (Motivational Hook)
        p2 = (
            f"Secara performa, nilai aktivitas ini **jauh lebih baik daripada tim lapangan pada perusahaan lain di luar sana**. "
            f"Di saat kompetitor mungkin hari ini masih sibuk berkutat dengan rapat atau baru sekadar merencanakan strategi di atas kertas, "
            f"team Bapak sudah satu langkah di depan dengan eksekusi nyata di lapangan."
        )

        # Paragraf 3: Penutup (Validasi Usaha)
        if "Kunjungan" in top_activity:
            p3 = "Setiap kunjungan yang terjadi hari ini adalah aset data riil yang tidak dimiliki oleh mereka yang hanya menunggu di kantor."
        else:
            p3 = "Konsistensi aktivitas ini adalah fondasi yang kuat untuk memenangkan pasar sebelum kompetitor menyadarinya."

        # Gabungkan
        insight = f"**Analisis Eksekutif (Mode Statistik):**\n\n{p1}\n\n{p2}\n\n{p3}"
        return insight

    except Exception as e:
        return f"Sistem sedang mengkalkulasi data tim... ({str(e)})"

# =========================================================
# [BARU] DYNAMIC UI HELPERS (Anti-Crash & Auto-Type)
# =========================================================

def clean_df_types_dynamically(df: pd.DataFrame) -> pd.DataFrame:
    """
    Versi Perbaikan: Menggunakan datetime64[ns] dan sinkronisasi keyword 
    untuk menghindari StreamlitAPIException akibat ketidakcocokan tipe data.
    """
    df_clean = df.copy()
    for col in df_clean.columns:
        col_lower = col.lower()
        
        # 1. Kolom Numerik: Pastikan murni angka dan isi NaN dengan 0
        # Menambahkan 'tenor' agar konsisten dengan NumberColumn
        if any(key in col_lower for key in ["nilai", "nominal", "sisa", "kontrak", "sepakat", "tenor"]):
            # Konversi rupiah string ke int, lalu paksa ke numeric murni
            df_clean[col] = pd.to_numeric(
                df_clean[col].apply(lambda x: parse_rupiah_to_int(x) if isinstance(x, str) else x), 
                errors='coerce'
            ).fillna(0)
            
        # 2. Kolom Tanggal: Gunakan tipe datetime pandas asli
        elif any(key in col_lower for key in ["tanggal", "tempo", "waktu"]):
            if not any(k in col_lower for k in ["log", "update", "timestamp"]):
                # PERBAIKAN: Hapus .dt.date. 
                # Pandas menyimpan .dt.date sebagai tipe 'object', yang memicu error di st.data_editor.
                # Biarkan tetap bertipe datetime64[ns].
                df_clean[col] = pd.to_datetime(df_clean[col], errors='coerce')
                
    return df_clean

def generate_dynamic_column_config(df):
    """
    Versi Perbaikan: Sinkron dengan fungsi clean agar tipe data kompatibel.
    """
    config = {}
    for col in df.columns:
        col_lower = col.lower()
        
        # Numerik: Gunakan NumberColumn untuk kolom yang sudah di-clean jadi angka
        if any(key in col_lower for key in ["nilai", "nominal", "sisa", "kontrak", "sepakat", "tenor"]):
            config[col] = st.column_config.NumberColumn(col, format="Rp %d", min_value=0)
            
        # Tanggal: Gunakan DateColumn untuk kolom datetime64[ns]
        elif any(key in col_lower for key in ["tanggal", "tempo", "waktu"]):
            if "timestamp" not in col_lower:
                config[col] = st.column_config.DateColumn(col, format="DD/MM/YYYY")
            else:
                config[col] = st.column_config.TextColumn(col, disabled=True)
                
        # Status: Gunakan CheckboxColumn jika data berisi Boolean (True/False)
        elif "status" in col_lower:
            config[col] = st.column_config.CheckboxColumn(col)
            
        else:
            config[col] = st.column_config.TextColumn(col)
            
    return config


# =========================================================
# ADMIN PASSWORD HELPERS
# =========================================================
def verify_admin_password(pwd_input: str) -> bool:
    """
    - Support 2 mode:
      (A) st.secrets["password_admin_hash"] = SHA256 hex dari password
      (B) st.secrets["password_admin"] = password plain (legacy)
    """
    pwd_input = safe_str(pwd_input, "").strip()
    if not pwd_input:
        return False

    # Mode hash (disarankan)
    hash_secret = None
    try:
        hash_secret = st.secrets.get("password_admin_hash", None)
    except Exception:
        hash_secret = None

    if hash_secret and safe_str(hash_secret, "").strip():
        try:
            digest = hashlib.sha256(pwd_input.encode("utf-8")).hexdigest()
            return hmac.compare_digest(digest, safe_str(hash_secret, "").strip())
        except Exception:
            return False

    # Mode plain (legacy)
    plain_secret = None
    try:
        plain_secret = st.secrets.get("password_admin", None)
    except Exception:
        plain_secret = None

    if plain_secret and safe_str(plain_secret, "").strip():
        return hmac.compare_digest(pwd_input, safe_str(plain_secret, "").strip())

    return False


def admin_secret_configured() -> bool:
    try:
        return bool(
            safe_str(st.secrets.get("password_admin_hash", ""), "").strip()
            or safe_str(st.secrets.get("password_admin", ""), "").strip()
        )
    except Exception:
        return False


# =========================================================
# CONNECTIONS
# =========================================================
KONEKSI_GSHEET_BERHASIL = False
KONEKSI_DROPBOX_BERHASIL = False
spreadsheet = None
dbx = None

# 1) Google Sheets
try:
    if "gcp_service_account" in st.secrets:
        scopes = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive"
        ]
        creds_dict = dict(st.secrets["gcp_service_account"])
        if "private_key" in creds_dict:
            creds_dict["private_key"] = creds_dict["private_key"].replace(
                "\\n", "\n")

        creds = Credentials.from_service_account_info(
            creds_dict, scopes=scopes)
        gc = gspread.authorize(creds)
        spreadsheet = gc.open(NAMA_GOOGLE_SHEET)
        KONEKSI_GSHEET_BERHASIL = True
        # =========================================================
        # [BARU] AUTO-CREATE AUDIT SHEET SAAT STARTUP
        # =========================================================
        # Tambahkan blok ini agar sheet otomatis dibuat saat aplikasi dibuka
        from audit_service import ensure_audit_sheet
        try:
            ensure_audit_sheet(spreadsheet)
            # print("Audit sheet ready.") # Opsional untuk debug console
        except Exception as e:
            st.error(
                f"⚠️ Sistem Error: Gagal membuat Sheet Audit otomatis. Pesan: {e}")
            # Ini akan memunculkan kotak merah di layar jika gagal,
            # jadi admin langsung tahu ada yang salah.
    else:
        st.error("GSheet Error: Kredensial tidak ditemukan.")
except Exception as e:
    st.error(f"GSheet Error: {e}")

# 2) Dropbox
try:
    # Kita langsung masukkan data yang didapat dari terminal tadi
    # agar script bisa otomatis refresh token sendiri
    APP_KEY = "6bks8aq249cy8kv"
    APP_SECRET = "2ai0jov47sx4b7y"
    REFRESH_TOKEN = "iFeGPgijH6kAAAAAAAAAAbxwoi6Sr8IYH3KZ1qxENSc_ejlR0p98K2mSUfIKXTo6"

    dbx = dropbox.Dropbox(
        app_key=APP_KEY,
        app_secret=APP_SECRET,
        oauth2_refresh_token=REFRESH_TOKEN
    )
    
    # Tes koneksi
    dbx.users_get_current_account()
    KONEKSI_DROPBOX_BERHASIL = True
    
except AuthError:
    st.error("Dropbox Error: Token Autentikasi tidak valid atau sudah dicabut.")
except Exception as e:
    st.error(f"Dropbox Error: {e}")
    

# === Konfigurasi AI Robust (Tiruan Proyek Telesales) ===
SDK = "new"
try:
    from google import genai as genai_new
    from google.genai import types as types_new
except Exception:
    SDK = "legacy"
    import google.generativeai as genai_legacy

# === Konfigurasi AI Robust (Tiruan Proyek Telesales) ===
SDK = "new"
try:
    from google import genai as genai_new
    from google.genai import types as types_new
except Exception:
    SDK = "legacy"
    import google.generativeai as genai_legacy

# AMBIL DARI SECRETS (SANGAT AMAN)
API_KEY = st.secrets.get("gemini_api_key", "")

# Daftar model cadangan
MODEL_FALLBACKS = ["gemini-2.0-flash", "gemini-1.5-flash"]

if SDK == "new":
    client_ai = genai_new.Client(api_key=API_KEY)
else:
    genai_legacy.configure(api_key=API_KEY)
# Daftar model cadangan agar tidak muncul pesan "berhalangan" jika satu model error
MODEL_FALLBACKS = ["gemini-1.5-flash", "gemini-1.5-pro", "gemini-2.0-flash-exp"]

if SDK == "new":
    client_ai = genai_new.Client(api_key=API_KEY)
else:
    genai_legacy.configure(api_key=API_KEY)


# =========================================================
# RUPIAH PARSER (input bebas -> int Rupiah)
# =========================================================
def parse_rupiah_to_int(value):
    """Parser Rupiah yang lebih pintar."""
    if value is None:
        return None

    # jika sudah numeric
    if isinstance(value, (int, float)) and not pd.isna(value):
        try:
            return int(round(float(value)))
        except Exception:
            return None

    s = str(value).strip()
    if not s:
        return None

    s_lower = s.lower().strip()
    if s_lower in {"nan", "none", "-", "null"}:
        return None

    # hilangkan spasi + penanda mata uang
    s_lower = re.sub(r"\\s+", "", s_lower)
    s_lower = s_lower.replace("idr", "").replace("rp", "")

    # deteksi satuan
    multiplier = 1
    if "miliar" in s_lower or "milyar" in s_lower:
        multiplier = 1_000_000_000
    elif "jt" in s_lower or "juta" in s_lower:
        multiplier = 1_000_000
    elif "rb" in s_lower or "ribu" in s_lower:
        multiplier = 1_000

    # buang kata satuan dari string angka
    s_num = re.sub(r"(miliar|milyar|juta|jt|ribu|rb)", "", s_lower)

    # sisakan digit + pemisah ribuan/desimal
    s_num = re.sub(r"[^0-9.,]", "", s_num)
    if not s_num:
        return None

    def to_float_locale(num_str: str) -> float:
        if "." in num_str and "," in num_str:
            if num_str.rfind(",") > num_str.rfind("."):
                cleaned = num_str.replace(".", "").replace(",", ".")
            else:
                cleaned = num_str.replace(",", "")
            return float(cleaned)

        if "," in num_str:
            if num_str.count(",") > 1:
                return float(num_str.replace(",", ""))
            after = num_str.split(",")[1]
            if len(after) == 3:
                return float(num_str.replace(",", ""))
            return float(num_str.replace(",", "."))

        if "." in num_str:
            if num_str.count(".") > 1:
                return float(num_str.replace(".", ""))
            after = num_str.split(".")[1]
            if len(after) == 3:
                return float(num_str.replace(".", ""))
            return float(num_str)

        return float(num_str)

    try:
        base = to_float_locale(s_num)
    except Exception:
        digits = re.sub(r"\\D", "", s_num)
        return int(digits) if digits else None

    if multiplier != 1:
        if base >= multiplier:
            return int(round(base))
        return int(round(base * multiplier))

    return int(round(base))


def format_rupiah_display(amount) -> str:
    """Hanya untuk display di UI (bukan untuk disimpan)."""
    try:
        if amount is None or pd.isna(amount):
            return ""
        n = int(amount)
        return "Rp " + f"{n:,}".replace(",", ".")
    except Exception:
        return str(amount)


# =========================================================
# AUDIT LOG HELPERS (PEMBAYARAN)
# =========================================================
def parse_payment_log_lines(log_text: str):
    """
    Normalisasi log lama/baru menjadi list baris TANPA nomor.
    - Kalau log sudah bernomor "1. ..." => nomor dihapus dulu.
    - Kalau format lama pakai ';' dalam satu baris => dipecah jadi multiline.
    - Baris tambahan dalam satu event dibuat indent (diawali spasi).
    """
    log_text = safe_str(log_text, "").strip()
    if not log_text:
        return []

    raw_lines = [ln.rstrip() for ln in log_text.splitlines() if ln.strip()]
    out = []

    for ln in raw_lines:
        # hapus numbering lama kalau ada: "12. ...."
        mnum = re.match(r"^\\s*\\d+\\.\\s*(.*)$", ln)
        if mnum:
            ln = mnum.group(1).rstrip()

        # kalau format: "[ts] (actor) ...."
        m = re.match(r"^\\[(.*?)\\]\\s*\\((.*?)\\)\\s*(.*)$", ln)
        if m:
            ts, actor, rest = m.group(1).strip(), m.group(
                2).strip(), m.group(3).strip()
            prefix = f"[{ts}] ({actor})"

            if rest:
                parts = [p.strip() for p in rest.split(";") if p.strip()]
                if parts:
                    out.append(f"{prefix} {parts[0]}")
                    for p in parts[1:]:
                        out.append(f" {p}")  # indent
                else:
                    out.append(prefix)
            else:
                out.append(prefix)
        else:
            out.append(ln)

    return out


def build_numbered_log(lines):
    """Buat output bernomor 1..N dari list baris (tanpa nomor)."""
    lines = [str(l).rstrip() for l in (lines or []) if safe_str(l, "").strip()]
    return "\\n".join([f"{i}. {line}" for i, line in enumerate(lines, 1)]).strip()


def _fmt_payment_val_for_log(col_name: str, v):
    """Format nilai agar enak dibaca di log."""
    if col_name == COL_NOMINAL_BAYAR:
        x = parse_rupiah_to_int(v)
        return format_rupiah_display(x) if x is not None else "-"
    if col_name == COL_STATUS_BAYAR:
        return "✅ Dibayar" if normalize_bool(v) else "⏳ Belum"
    if col_name in {COL_JATUH_TEMPO, COL_TGL_EVENT}:
        d = normalize_date(v)
        return d.strftime("%Y-%m-%d") if d else "-"
    s = safe_str(v, "-").replace("\\n", " ").strip()
    return s if s else "-"


def append_payment_ts_update(existing_log: str, ts: str, actor: str, changes):
    """
    Append perubahan ke log dengan format rapih & bernomor.
    """
    lines = parse_payment_log_lines(existing_log)
    changes = [safe_str(c, "").strip()
               for c in (changes or []) if safe_str(c, "").strip()]
    if not changes:
        return build_numbered_log(lines)

    actor = safe_str(actor, "-").strip() or "-"
    ts = safe_str(ts, now_ts_str()).strip() or now_ts_str()

    # baris pertama event
    lines.append(f"[{ts}] ({actor}) {changes[0]}")

    # baris selanjutnya: indent (tanpa ulang prefix)
    for c in changes[1:]:
        lines.append(f" {c}")

    return build_numbered_log(lines)


# =========================================================
# UI DISPLAY HELPERS (RUPIAH)
# =========================================================
def payment_df_for_display(df: pd.DataFrame) -> pd.DataFrame:
    """
    Mengubah kolom angka menjadi format string 'Rp 2.000.000' khusus untuk tampilan UI.
    """
    dfv = df.copy()
    if dfv is None or dfv.empty:
        return dfv
    
    # Daftar kolom yang harus diformat Rupiah
    money_cols = [COL_NOMINAL_BAYAR, COL_NILAI_KESEPAKATAN, COL_SISA_BAYAR]
    
    for col in money_cols:
        if col in dfv.columns:
            # Fungsi lambda: Cek validitas angka -> Format Rupiah dengan Titik
            dfv[col] = dfv[col].apply(
                lambda x: "Rp " + "{:,.0f}".format(parse_rupiah_to_int(x)).replace(",", ".") 
                if parse_rupiah_to_int(x) is not None else "Rp 0"
            )
            
    return dfv


def on_change_pay_nominal():
    """Auto-format input nominal ke 'Rp 15.000.000' (untuk UI)."""
    raw = st.session_state.get("pay_nominal", "")
    val = parse_rupiah_to_int(raw)
    if val is not None:
        st.session_state["pay_nominal"] = format_rupiah_display(val)


def reset_payment_form_state():
    """Reset field input pembayaran (agar terasa seperti clear_on_submit)."""
    keys = [
        "pay_group",
        "pay_marketing",
        "pay_event_date",
        "pay_jenis_opt",
        "pay_jenis_custom",
        "pay_nominal",
        "pay_due_date",
        "pay_status",
        "pay_note",
        "pay_file",
    ]
    for k in keys:
        try:
            if k == "pay_event_date":
                st.session_state[k] = datetime.now(tz=TZ_JKT).date()
            elif k == "pay_due_date":
                st.session_state[k] = datetime.now(
                    tz=TZ_JKT).date() + timedelta(days=7)
            elif k == "pay_jenis_opt":
                st.session_state[k] = "Down Payment (DP)"
            elif k == "pay_status":
                st.session_state[k] = False
            else:
                st.session_state[k] = ""
        except Exception:
            pass


# =========================================================
# EXCEL EXPORT
# =========================================================
def df_to_excel_bytes(
    df: pd.DataFrame,
    sheet_name="Sheet1",
    col_widths=None,
    wrap_cols=None,
    right_align_cols=None,
    number_format_cols=None
):
    """Export dataframe ke .xlsx rapi."""
    if not HAS_OPENPYXL:
        return None

    df_export = df.copy()
    df_export = df_export.where(pd.notna(df_export), None)

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]

    for r in dataframe_to_rows(df_export, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill("solid", fgColor="E6E6E6")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"

    wrap_cols = set(wrap_cols or [])
    right_align_cols = set(right_align_cols or [])
    number_format_cols = dict(number_format_cols or {})
    col_widths = dict(col_widths or {})

    cols = list(df_export.columns)

    for i, col_name in enumerate(cols, 1):
        col_letter = get_column_letter(i)

        if col_name in col_widths:
            ws.column_dimensions[col_letter].width = col_widths[col_name]
        else:
            max_len = len(str(col_name))
            try:
                for v in df_export[col_name]:
                    v_str = "" if v is None else str(v)
                    max_len = max(max_len, len(v_str))
            except Exception:
                pass
            ws.column_dimensions[col_letter].width = min(
                max(10, max_len + 2), 60)

        for cell in ws[col_letter][1:]:
            wrap = col_name in wrap_cols
            horiz = "right" if col_name in right_align_cols else "left"
            cell.alignment = Alignment(
                vertical="top", horizontal=horiz, wrap_text=wrap)

            if col_name in number_format_cols:
                cell.number_format = number_format_cols[col_name]

    wb.save(output)
    return output.getvalue()


# =========================================================
# GOOGLE SHEETS FORMATTING
# =========================================================
def _build_currency_number_format_rupiah():
    return {"type": "CURRENCY", "pattern": '"Rp" #,##0'}


def maybe_auto_format_sheet(worksheet, force: bool = False):
    """Throttled formatting: avoid calling heavy formatting too often."""
    try:
        if worksheet is None:
            return
        if "_fmt_sheet_last" not in st.session_state:
            st.session_state["_fmt_sheet_last"] = {}

        now = time.time()
        key = str(getattr(worksheet, "id", "unknown"))
        last = float(st.session_state["_fmt_sheet_last"].get(key, 0))
        if force or (now - last) > FORMAT_THROTTLE_SECONDS:
            auto_format_sheet(worksheet)
            st.session_state["_fmt_sheet_last"][key] = now
    except Exception:
        # Never break app due to formatting.
        pass


def auto_format_sheet(worksheet):
    """Auto-format Google Sheet."""
    try:
        sheet_id = worksheet.id
        all_values = worksheet.get_all_values()
        if not all_values:
            return

        headers = all_values[0]
        data_row_count = len(all_values)
        formatting_row_count = max(worksheet.row_count, data_row_count)

        requests = []
        default_body_format = {
            "verticalAlignment": "TOP", "wrapStrategy": "CLIP"}

        # 1) Reset body base style
        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 1, "endRowIndex": formatting_row_count},
                "cell": {"userEnteredFormat": default_body_format},
                "fields": "userEnteredFormat(verticalAlignment,wrapStrategy)"
            }
        })

        # 2) Column sizing + per-column overrides
        for i, col_name in enumerate(headers):
            col_index = i
            cell_format_override = {}
            width = 100

            long_text_cols = {
                "Misi", "Target", "Deskripsi", "Bukti/Catatan", "Link Foto", "Link Sosmed",
                "Tempat Dikunjungi", "Kesimpulan", "Kendala", "Next Plan (Pending)", "Feedback Lead",
                COL_KENDALA_KLIEN,
                COL_NAMA_KLIEN,
                TEAM_COL_NAMA_TEAM, TEAM_COL_POSISI, TEAM_COL_ANGGOTA,
                COL_GROUP, COL_MARKETING, COL_BIDANG,
                COL_JENIS_BAYAR, COL_BUKTI_BAYAR, COL_CATATAN_BAYAR,
                COL_TS_UPDATE,
            }

            if col_name in long_text_cols:
                width = 360 if col_name == COL_TS_UPDATE else 300
                cell_format_override["wrapStrategy"] = "WRAP"
            elif col_name in {"Tgl_Mulai", "Tgl_Selesai", "Timestamp", COL_TGL_EVENT, COL_JATUH_TEMPO, COL_TS_BAYAR}:
                width = 160 if col_name in {"Timestamp", COL_TS_BAYAR} else 120
                cell_format_override["horizontalAlignment"] = "CENTER"
            elif col_name in {"Status", "Done?", COL_STATUS_BAYAR}:
                width = 130 if col_name == COL_STATUS_BAYAR else 80
                cell_format_override["horizontalAlignment"] = "CENTER"
            elif col_name == "Nama":
                width = 160
            elif col_name == COL_UPDATED_BY:
                width = 160
            elif col_name == COL_INTEREST:
                width = 140
                cell_format_override["horizontalAlignment"] = "CENTER"
            elif col_name == COL_KONTAK_KLIEN:
                width = 150
                cell_format_override["horizontalAlignment"] = "CENTER"
            elif col_name in {COL_NILAI_KONTRAK, COL_NOMINAL_BAYAR}:
                width = 180
                cell_format_override["horizontalAlignment"] = "RIGHT"
                cell_format_override["numberFormat"] = _build_currency_number_format_rupiah(
                )

            # Set width
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": col_index,
                        "endIndex": col_index + 1
                    },
                    "properties": {"pixelSize": width},
                    "fields": "pixelSize"
                }
            })

            # Apply per-column format
            if cell_format_override:
                fields = ",".join(cell_format_override.keys())
                requests.append({
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 1,
                            "endRowIndex": formatting_row_count,
                            "startColumnIndex": col_index,
                            "endColumnIndex": col_index + 1
                        },
                        "cell": {"userEnteredFormat": cell_format_override},
                        "fields": f"userEnteredFormat({fields})"
                    }
                })

        # 3) Header style
        requests.append({
            "repeatCell": {
                "range": {"sheetId": sheet_id, "startRowIndex": 0, "endRowIndex": 1},
                "cell": {"userEnteredFormat": {
                    "textFormat": {"bold": True},
                    "horizontalAlignment": "CENTER",
                    "verticalAlignment": "MIDDLE",
                    "backgroundColor": {"red": 0.90, "green": 0.92, "blue": 0.96},
                    "wrapStrategy": "WRAP"
                }},
                "fields": "userEnteredFormat(textFormat,horizontalAlignment,verticalAlignment,backgroundColor,wrapStrategy)"
            }
        })

        # 4) Freeze header
        requests.append({
            "updateSheetProperties": {
                "properties": {"sheetId": sheet_id, "gridProperties": {"frozenRowCount": 1}},
                "fields": "gridProperties.frozenRowCount"
            }
        })

        worksheet.spreadsheet.batch_update({"requests": requests})
    except Exception as e:
        print(f"Format Error: {e}")


def ensure_headers(worksheet, desired_headers):
    """
    Pastikan header sesuai urutan standar.
    """
    try:
        if worksheet.col_count < len(desired_headers):
            worksheet.resize(cols=len(desired_headers))

        headers = worksheet.row_values(1)
        need_reset = (
            not headers
            or (len(headers) < len(desired_headers))
            or (headers[:len(desired_headers)] != desired_headers)
        )
        if need_reset:
            worksheet.update(range_name="A1", values=[
                             desired_headers], value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(worksheet, force=True)
    except Exception as e:
        print(f"Ensure Header Error: {e}")


# =========================================================
# WORKSHEET GET/CREATE + STAFF LIST
# =========================================================
@st.cache_resource(ttl=600)
def _get_or_create_ws_cached(nama_worksheet: str):
    """Get/create worksheet object (cached)."""
    try:
        ws = spreadsheet.worksheet(nama_worksheet)
        return ws
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(
            title=nama_worksheet, rows=200, cols=len(NAMA_KOLOM_STANDAR))
        ws.append_row(NAMA_KOLOM_STANDAR, value_input_option="USER_ENTERED")
        maybe_auto_format_sheet(ws, force=True)
        return ws
    except Exception:
        return None


def get_or_create_worksheet(nama_worksheet):
    """
    Pastikan header selalu up-to-date.
    """
    ws = _get_or_create_ws_cached(nama_worksheet)
    if ws is not None:
        ensure_headers(ws, NAMA_KOLOM_STANDAR)
    return ws


@st.cache_data(ttl=3600)
def get_daftar_staf_terbaru():
    default_staf = ["Saya"]
    if not KONEKSI_GSHEET_BERHASIL:
        return default_staf

    try:
        try:
            ws = spreadsheet.worksheet(SHEET_CONFIG_NAMA)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=SHEET_CONFIG_NAMA, rows=100, cols=1)
            ws.append_row(["Daftar Nama Staf"],
                          value_input_option="USER_ENTERED")
            ws.append_row(["Saya"], value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
            return default_staf

        nama_list = ws.col_values(1)
        if nama_list and nama_list[0] == "Daftar Nama Staf":
            nama_list.pop(0)

        return nama_list if nama_list else default_staf
    except Exception:
        return default_staf


def hapus_staf_by_name(nama_staf):
    """Menghapus nama staf dari worksheet Config_Staf."""
    try:
        ws = spreadsheet.worksheet(SHEET_CONFIG_NAMA)
        # Cari sel yang berisi nama tersebut
        cell = ws.find(nama_staf)
        if cell:
            ws.delete_rows(cell.row)
            return True, f"Staf '{nama_staf}' berhasil dihapus."
        return False, "Nama staf tidak ditemukan di database."
    except Exception as e:
        return False, f"Gagal menghapus: {e}"


def tambah_staf_baru(nama_baru):
    try:
        try:
            ws = spreadsheet.worksheet(SHEET_CONFIG_NAMA)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=SHEET_CONFIG_NAMA, rows=100, cols=1)

        if nama_baru in ws.col_values(1):
            return False, "Nama sudah ada!"

        ws.append_row([nama_baru], value_input_option="USER_ENTERED")
        # maybe_auto_format_sheet(ws)
        return True, "Berhasil tambah tim!"
    except Exception as e:
        return False, str(e)


# =========================================================
# TEAM CONFIG
# =========================================================
@st.cache_data(ttl=3600)
def load_team_config():
    if not KONEKSI_GSHEET_BERHASIL:
        return pd.DataFrame(columns=TEAM_COLUMNS)

    try:
        try:
            ws = spreadsheet.worksheet(SHEET_CONFIG_TEAM)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=SHEET_CONFIG_TEAM, rows=300, cols=len(TEAM_COLUMNS))
            ws.append_row(TEAM_COLUMNS, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
            return pd.DataFrame(columns=TEAM_COLUMNS)

        data = ws.get_all_records()
        df = pd.DataFrame(data).fillna("")
        for c in TEAM_COLUMNS:
            if c not in df.columns:
                df[c] = ""
        return df[TEAM_COLUMNS].copy()
    except Exception:
        return pd.DataFrame(columns=TEAM_COLUMNS)


def tambah_team_baru(nama_team, posisi, anggota_list):
    if not KONEKSI_GSHEET_BERHASIL:
        return False, "Koneksi GSheet belum aktif."

    try:
        nama_team = str(nama_team).strip()
        posisi = str(posisi).strip()
        anggota_list = [str(a).strip() for a in anggota_list if str(a).strip()]

        if not nama_team or not posisi or not anggota_list:
            return False, "Nama team, posisi, dan minimal 1 anggota wajib diisi."

        try:
            ws = spreadsheet.worksheet(SHEET_CONFIG_TEAM)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=SHEET_CONFIG_TEAM, rows=300, cols=len(TEAM_COLUMNS))
            ws.append_row(TEAM_COLUMNS, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)

        existing = set()
        try:
            for r in ws.get_all_records():
                key = (
                    str(r.get(TEAM_COL_NAMA_TEAM, "")).strip(),
                    str(r.get(TEAM_COL_POSISI, "")).strip(),
                    str(r.get(TEAM_COL_ANGGOTA, "")).strip()
                )
                existing.add(key)
        except Exception:
            pass

        rows_to_add = []
        for anggota in anggota_list:
            key = (nama_team, posisi, anggota)
            if key not in existing:
                rows_to_add.append([nama_team, posisi, anggota])

        if not rows_to_add:
            return False, "Semua anggota sudah terdaftar di team tersebut."

        ws.append_rows(rows_to_add, value_input_option="USER_ENTERED")
        # maybe_auto_format_sheet(ws)
        return True, f"Berhasil tambah team '{nama_team}' ({len(rows_to_add)} anggota)."
    except Exception as e:
        return False, str(e)


# =========================================================
# DROPBOX UPLOAD
# =========================================================
def upload_ke_dropbox(file_obj, nama_staf, kategori="Umum"):
    if not KONEKSI_DROPBOX_BERHASIL or dbx is None:
        return "Koneksi Dropbox Error"

    try:
        file_data = file_obj.getvalue()
        ts = datetime.now(tz=TZ_JKT).strftime("%Y%m%d_%H%M%S")

        clean_filename = "".join(
            [c for c in file_obj.name if c.isalnum() or c in (".", "_")])
        clean_user_folder = "".join(
            [c for c in nama_staf if c.isalnum() or c in (" ", "_")]).replace(" ", "_")
        clean_kategori = "".join(
            [c for c in kategori if c.isalnum() or c in (" ", "_")]).replace(" ", "_")

        path = f"{FOLDER_DROPBOX}/{clean_user_folder}/{clean_kategori}/{ts}_{clean_filename}"
        dbx.files_upload(file_data, path, mode=dropbox.files.WriteMode.add)

        settings = SharedLinkSettings(
            requested_visibility=RequestedVisibility.public)
        try:
            link = dbx.sharing_create_shared_link_with_settings(
                path, settings=settings)
        except ApiError as e:
            if e.error.is_shared_link_already_exists():
                link = dbx.sharing_list_shared_links(
                    path, direct_only=True).links[0]
            else:
                return "-"

        return link.url.replace("?dl=0", "?raw=1")
    except Exception:
        return "-"


# =========================================================
# TARGET / CHECKLIST HELPERS
# =========================================================
def clean_bulk_input(text_input):
    # GUNAKAN \n (satu backslash), bukan \\n
    lines = (text_input or "").split("\n") 
    cleaned_targets = []
    for line in lines:
        # Regex ini akan menghapus angka "1.", "2.", "-", atau "*" di awal baris
        cleaned = re.sub(r"^[\d\.\-\*\s]+", "", line).strip()
        if cleaned:
            cleaned_targets.append(cleaned)
    return cleaned_targets


@st.cache_data(ttl=3600)
def load_checklist(sheet_name, columns):
    try:
        try:
            ws = spreadsheet.worksheet(sheet_name)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=sheet_name, rows=200, cols=len(columns))
            ws.append_row(columns, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
            return pd.DataFrame(columns=columns)

        ensure_headers(ws, columns)

        data = ws.get_all_records()
        df = pd.DataFrame(data).fillna("")

        for col in columns:
            if col not in df.columns:
                if col == "Status":
                    df[col] = False
                else:
                    df[col] = ""

        if "Status" in df.columns:
            df["Status"] = df["Status"].apply(
                lambda x: True if str(x).upper() == "TRUE" else False)

        return df[columns].copy()
    except Exception:
        return pd.DataFrame(columns=columns)


def save_checklist(sheet_name, df, columns):
    try:
        ws = spreadsheet.worksheet(sheet_name)
        ensure_headers(ws, columns)

        ws.clear()

        rows_needed = len(df) + 1
        if ws.row_count < rows_needed:
            ws.resize(rows=rows_needed)

        df_save = df.copy().fillna("")
        for c in columns:
            if c not in df_save.columns:
                df_save[c] = ""

        if "Status" in df_save.columns:
            df_save["Status"] = df_save["Status"].apply(
                lambda x: "TRUE" if bool(x) else "FALSE")

        df_save = df_save[columns].astype(str)
        data_to_save = [df_save.columns.values.tolist()] + \
            df_save.values.tolist()

        ws.update(range_name="A1", values=data_to_save,
                  value_input_option="USER_ENTERED")
        # maybe_auto_format_sheet(ws)
        return True
    except Exception:
        return False


def apply_audit_checklist_changes(df_before: pd.DataFrame, df_after: pd.DataFrame, key_cols, actor: str):
    """Update audit columns hanya untuk baris yang benar-benar berubah."""
    if df_after is None or df_after.empty:
        return df_after

    actor = safe_str(actor, "-").strip() or "-"

    before = df_before.copy() if df_before is not None else pd.DataFrame()
    after = df_after.copy()

    for c in [COL_TS_UPDATE, COL_UPDATED_BY]:
        if c not in after.columns:
            after[c] = ""

    def make_key_row(r):
        return tuple(safe_str(r.get(k, "")).strip() for k in key_cols)

    before_map = {}
    if not before.empty:
        for _, r in before.iterrows():
            before_map[make_key_row(r)] = r.to_dict()

    ts = now_ts_str()
    watched_cols = {"Status", "Bukti/Catatan"}

    for idx, r in after.iterrows():
        k = make_key_row(r)
        prev = before_map.get(k, None)
        if prev is None:
            after.at[idx, COL_TS_UPDATE] = ts
            after.at[idx, COL_UPDATED_BY] = actor
            continue

        changed = False
        for col in watched_cols:
            if col not in after.columns:
                continue
            oldv = prev.get(col, "")
            newv = r.get(col, "")
            if col == "Status":
                if normalize_bool(oldv) != normalize_bool(newv):
                    changed = True
            else:
                if safe_str(oldv, "").strip() != safe_str(newv, "").strip():
                    changed = True

        if changed:
            after.at[idx, COL_TS_UPDATE] = ts
            after.at[idx, COL_UPDATED_BY] = actor

    return after


def add_bulk_targets(sheet_name, base_row_data, targets_list):
    try:
        try:
            ws = spreadsheet.worksheet(sheet_name)
        except Exception:
            return False

        columns = TEAM_CHECKLIST_COLUMNS if sheet_name == SHEET_TARGET_TEAM else INDIV_CHECKLIST_COLUMNS
        ensure_headers(ws, columns)

        actor = get_actor_fallback(default="Admin")
        ts = now_ts_str()

        rows_to_add = []
        for t in targets_list:
            row_vals = list(base_row_data) if base_row_data else []
            new_row = [""] * len(columns)

            for i in range(min(len(row_vals), len(columns))):
                new_row[i] = row_vals[i]

            if sheet_name == SHEET_TARGET_TEAM:
                new_row[0] = t
            elif sheet_name == SHEET_TARGET_INDIVIDU:
                new_row[1] = t

            if COL_TS_UPDATE in columns:
                new_row[columns.index(COL_TS_UPDATE)] = ts
            if COL_UPDATED_BY in columns:
                new_row[columns.index(COL_UPDATED_BY)] = actor

            rows_to_add.append(new_row)

        ws.append_rows(rows_to_add, value_input_option="USER_ENTERED")
        # maybe_auto_format_sheet(ws)
        return True
    except Exception:
        return False


def update_evidence_row(sheet_name, target_name, note, file_obj, user_folder_name, kategori_folder):
    """
    Update bukti/catatan untuk checklist (Team/Individu).
    ✅ Optimasi: gunakan batch_update untuk mengurangi jumlah API call.
    """
    try:
        ws = spreadsheet.worksheet(sheet_name)

        columns = TEAM_CHECKLIST_COLUMNS if sheet_name == SHEET_TARGET_TEAM else INDIV_CHECKLIST_COLUMNS
        ensure_headers(ws, columns)

        df = pd.DataFrame(ws.get_all_records()).fillna("")

        col_target_key = "Misi" if sheet_name == SHEET_TARGET_TEAM else "Target"
        if col_target_key not in df.columns:
            return False, "Kolom kunci error."

        matches = df.index[df[col_target_key] == target_name].tolist()
        if not matches:
            return False, "Target tidak ditemukan."

        row_idx_pandas = matches[0]
        row_idx_gsheet = row_idx_pandas + 2

        link_bukti = ""
        if file_obj:
            link_bukti = upload_ke_dropbox(
                file_obj, user_folder_name, kategori=kategori_folder)

        catatan_lama = str(df.at[row_idx_pandas, "Bukti/Catatan"]
                           ) if "Bukti/Catatan" in df.columns else ""
        if catatan_lama in {"-", "nan"}:
            catatan_lama = ""

        ts_update = now_ts_str()
        actor = safe_str(user_folder_name, "-").strip() or "-"

        update_text = f"[{ts_update}] "
        if note:
            update_text += f"{note}. "
        if link_bukti and link_bukti != "-":
            update_text += f"[FOTO: {link_bukti}]"

        final_note = f"{catatan_lama}\\n{update_text}" if catatan_lama.strip(
        ) else update_text
        final_note = final_note.strip() if final_note.strip() else "-"
        final_note = final_note.strip() if final_note.strip() else "-"

        headers = ws.row_values(1)
        if "Bukti/Catatan" not in headers:
            return False, "Kolom Bukti error."

        updates = []

        # Bukti/Catatan
        col_idx_gsheet = headers.index("Bukti/Catatan") + 1
        cell_address = gspread.utils.rowcol_to_a1(
            row_idx_gsheet, col_idx_gsheet)
        updates.append({"range": cell_address, "values": [[final_note]]})

        # Timestamp Update
        if COL_TS_UPDATE in headers:
            col_ts = headers.index(COL_TS_UPDATE) + 1
            cell_ts = gspread.utils.rowcol_to_a1(row_idx_gsheet, col_ts)
            updates.append({"range": cell_ts, "values": [[ts_update]]})

        # Updated By
        if COL_UPDATED_BY in headers:
            col_by = headers.index(COL_UPDATED_BY) + 1
            cell_by = gspread.utils.rowcol_to_a1(row_idx_gsheet, col_by)
            updates.append({"range": cell_by, "values": [[actor]]})

        ws.batch_update(updates, value_input_option="USER_ENTERED")

        # maybe_auto_format_sheet(ws)
        return True, "Berhasil update!"
    except Exception as e:
        return False, f"Error: {e}"


# =========================================================
# FEEDBACK + DAILY REPORT
# =========================================================
def kirim_feedback_admin(nama_staf, timestamp_key, isi_feedback):
    try:
        ws = spreadsheet.worksheet(nama_staf)

        if ws.col_count < len(NAMA_KOLOM_STANDAR):
            ws.resize(cols=len(NAMA_KOLOM_STANDAR))

        headers = ws.row_values(1)
        if COL_FEEDBACK not in headers:
            ws.update_cell(1, len(headers) + 1, COL_FEEDBACK)
            headers.append(COL_FEEDBACK)
            maybe_auto_format_sheet(ws, force=True)

        all_timestamps = ws.col_values(1)

        def clean_ts(text):
            return "".join(filter(str.isdigit, str(text)))

        target_clean = clean_ts(timestamp_key)
        found_row = None

        for idx, val in enumerate(all_timestamps):
            if clean_ts(val) == target_clean:
                found_row = idx + 1
                break

        if not found_row:
            return False, "Data tidak ditemukan."

        col_idx = headers.index(COL_FEEDBACK) + 1

        ts = now_ts_str()
        actor = get_actor_fallback(default="Admin")
        feedback_text = f"[{ts}] ({actor}) {isi_feedback}"

        ws.update_cell(found_row, col_idx, feedback_text)
        return True, "Feedback terkirim!"
    except Exception as e:
        return False, f"Error: {e}"


def simpan_laporan_harian_batch(list_of_rows, nama_staf):
    try:
        ws = get_or_create_worksheet(nama_staf)
        if ws is None:
            return False

        ensure_headers(ws, NAMA_KOLOM_STANDAR)
        ws.append_rows(list_of_rows, value_input_option="USER_ENTERED")

        # ✅ Optimasi: jangan format tiap submit (throttled)
        # maybe_auto_format_sheet(ws)

        return True
    except Exception as e:
        print(f"Error saving daily report batch: {e}")
        return False


@st.cache_data(ttl=3600)
def get_reminder_pending(nama_staf):
    try:
        ws = get_or_create_worksheet(nama_staf)
        if not ws:
            return None
        all_vals = ws.get_all_records()
        if not all_vals:
            return None
        last_row = all_vals[-1]
        pending_task = last_row.get(COL_PENDING, "")
        if pending_task and str(pending_task).strip() not in {"-", ""}:
            return pending_task
        return None
    except Exception:
        return None


@st.cache_data(ttl=3600)
def load_all_reports(daftar_staf):
    all_data = []
    for nama in daftar_staf:
        try:
            ws = get_or_create_worksheet(nama)
            if ws:
                d = ws.get_all_records()
                if d:
                    all_data.extend(d)
        except Exception:
            pass
    return pd.DataFrame(all_data) if all_data else pd.DataFrame(columns=NAMA_KOLOM_STANDAR)


def render_hybrid_table(df_data, unique_key, main_text_col):
    use_aggrid_attempt = HAS_AGGRID

    if use_aggrid_attempt:
        try:
            df_grid = df_data.copy().reset_index(drop=True)
            gb = GridOptionsBuilder.from_dataframe(df_grid)

            if "Status" in df_grid.columns:
                gb.configure_column("Status", editable=True, width=90)

            if main_text_col in df_grid.columns:
                gb.configure_column(
                    main_text_col, wrapText=True, autoHeight=True, width=400, editable=False)

            if "Bukti/Catatan" in df_grid.columns:
                gb.configure_column(
                    "Bukti/Catatan",
                    wrapText=True,
                    autoHeight=True,
                    editable=True,
                    cellEditor="agLargeTextCellEditor",
                    width=300
                )

            if COL_TS_UPDATE in df_grid.columns:
                gb.configure_column(COL_TS_UPDATE, editable=False, width=420)
            if COL_UPDATED_BY in df_grid.columns:
                gb.configure_column(COL_UPDATED_BY, editable=False, width=160)

            gb.configure_default_column(editable=False)
            gridOptions = gb.build()

            grid_response = AgGrid(
                df_grid,
                gridOptions=gridOptions,
                update_mode=GridUpdateMode.MODEL_CHANGED,
                fit_columns_on_grid_load=True,
                height=420,
                theme="streamlit",
                key=f"aggrid_{unique_key}"
            )
            return pd.DataFrame(grid_response["data"])
        except Exception:
            use_aggrid_attempt = False

    column_config = {}
    if "Status" in df_data.columns:
        column_config["Status"] = st.column_config.CheckboxColumn(
            "Done?", width="small")
    if main_text_col in df_data.columns:
        column_config[main_text_col] = st.column_config.TextColumn(
            main_text_col, disabled=True, width="large")
    if "Bukti/Catatan" in df_data.columns:
        column_config["Bukti/Catatan"] = st.column_config.TextColumn(
            "Bukti/Note", width="large")
    if COL_TS_UPDATE in df_data.columns:
        column_config[COL_TS_UPDATE] = st.column_config.TextColumn(
            COL_TS_UPDATE, disabled=True, width="large")
    if COL_UPDATED_BY in df_data.columns:
        column_config[COL_UPDATED_BY] = st.column_config.TextColumn(
            COL_UPDATED_BY, disabled=True, width="medium")

    return st.data_editor(
        df_data,
        column_config=column_config,
        hide_index=True,
        key=f"editor_native_{unique_key}",
        use_container_width=True
    )


def render_laporan_harian_mobile():
    st.markdown("## 📝 Laporan Harian")

    # tombol balik
    if st.button("⬅️ Kembali ke Beranda", use_container_width=True):
        set_nav("home")

    staff_list = get_daftar_staf_terbaru()

    # tetap pakai key pelapor_main agar actor log tetap konsisten
    nama_pelapor = st.selectbox("Nama Pelapor", staff_list, key="pelapor_main")

    pending_msg = get_reminder_pending(nama_pelapor)
    if pending_msg:
        st.warning(f"🔔 Pending terakhir: **{pending_msg}**")

    tab1, tab2, tab3, tab4 = st.tabs(
        ["📌 Aktivitas", "🏁 Kesimpulan", "📇 Kontak", "✅ Submit"])

    # ===== TAB 1: Aktivitas =====
    with tab1:
        kategori_aktivitas = st.radio(
            "Jenis Aktivitas",
            ["🚗 Sales (Kunjungan Lapangan)", "💻 Digital Marketing / Konten / Ads",
             "📞 Telesales / Follow Up", "🏢 Lainnya"],
            horizontal=False,
            key="m_kategori"
        )
        is_kunjungan = kategori_aktivitas.startswith("🚗")

        if "Digital Marketing" in kategori_aktivitas:
            st.text_input("Link Konten / Ads / Drive (Opsional)",
                          key="m_sosmed")

        if is_kunjungan:
            st.text_input(
                "📍 Nama Klien / Lokasi Kunjungan (Wajib)", key="m_lokasi")
        else:
            st.text_input("Jenis Tugas", value=kategori_aktivitas,
                          disabled=True, key="m_tugas")

        fotos = st.file_uploader(
            "Upload Bukti (opsional)",
            accept_multiple_files=True,
            disabled=not KONEKSI_DROPBOX_BERHASIL,
            key="m_fotos"
        )

        # 1 deskripsi saja agar ringkas (bisa detail per file via expander)
        st.text_area("Deskripsi Aktivitas (Wajib)",
                     height=120, key="m_deskripsi")

        with st.expander("Detail deskripsi per file (opsional)", expanded=False):
            if fotos:
                for i, f in enumerate(fotos):
                    st.text_input(f"Ket. {f.name}", key=f"m_desc_{i}")

    # ===== TAB 2: Kesimpulan =====
    with tab2:
        st.text_area("💡 Kesimpulan hari ini", height=100, key="m_kesimpulan")
        st.text_area("🚧 Kendala internal", height=90, key="m_kendala")
        st.text_area("🧑‍💼 Kendala klien", height=90, key="m_kendala_klien")

    # ===== TAB 3: Kontak =====
    with tab3:
        st.radio(
            "📈 Tingkat Interest",
            ["Under 50% (A)", "50-75% (B)", "75%-100%"],
            horizontal=False,
            key="interest_persen"
        )
        st.text_input("👤 Nama Klien", key="nama_klien_input")
        st.text_input("📞 No HP/WA Klien", key="kontak_klien_input")
        st.text_input("📌 Next Plan / Pending (Reminder Besok)",
                      key="m_pending")

# ===== TAB 4: Submit =====
    with tab4:
        st.caption("Pastikan data sudah benar, lalu submit.")

        if st.button("✅ Submit Laporan", type="primary", use_container_width=True):

            # --- 1. SIAPKAN VARIABEL DATA ---
            kategori_aktivitas = st.session_state.get("m_kategori", "")
            is_kunjungan = str(kategori_aktivitas).startswith("🚗")
            lokasi_input = st.session_state.get(
                "m_lokasi", "") if is_kunjungan else kategori_aktivitas
            main_deskripsi = st.session_state.get("m_deskripsi", "")
            sosmed_link = st.session_state.get(
                "m_sosmed", "") if "Digital Marketing" in str(kategori_aktivitas) else ""
            fotos = st.session_state.get("m_fotos", None)

            # --- 2. VALIDASI INPUT ---
            if is_kunjungan and not str(lokasi_input).strip():
                st.error("Lokasi kunjungan wajib diisi.")
                st.stop()

            if (not fotos) and (not str(main_deskripsi).strip()):
                st.error("Deskripsi wajib diisi.")
                st.stop()

            # --- 3. PERSIAPAN PROGRESS BAR ---
            # Container kosong untuk menaruh loading bar
            progress_placeholder = st.empty()

            # Hitung total langkah (Jumlah Foto + 1 langkah simpan ke Excel/GSheet)
            jml_foto = len(fotos) if fotos else 0
            total_steps = jml_foto + 1
            current_step = 0

            # Tampilkan Bar Awal (0%)
            my_bar = progress_placeholder.progress(
                0, text="🚀 Memulai proses...")

            try:
                # Siapkan data timestamp & string lain
                ts = now_ts_str()
                val_kesimpulan = (st.session_state.get(
                    "m_kesimpulan") or "-").strip() or "-"
                val_kendala = (st.session_state.get(
                    "m_kendala") or "-").strip() or "-"
                val_kendala_klien = (st.session_state.get(
                    "m_kendala_klien") or "-").strip() or "-"
                val_pending = (st.session_state.get(
                    "m_pending") or "-").strip() or "-"
                val_feedback = ""
                val_interest = st.session_state.get("interest_persen") or "-"
                val_nama_klien = (st.session_state.get(
                    "nama_klien_input") or "-").strip() or "-"
                val_kontak_klien = (st.session_state.get(
                    "kontak_klien_input") or "-").strip() or "-"

                rows = []
                final_lokasi = lokasi_input if is_kunjungan else kategori_aktivitas

                # --- 4. PROSES UPLOAD FOTO (LOOPING) ---
                if fotos and KONEKSI_DROPBOX_BERHASIL:
                    for i, f in enumerate(fotos):
                        # Update Persentase Progress Bar
                        # (Contoh: Foto 1 dari 3 => 33%)
                        pct = float(current_step / total_steps)
                        # Pastikan pct tidak lebih dari 1.0
                        if pct > 1.0:
                            pct = 1.0

                        my_bar.progress(
                            pct, text=f"📤 Mengupload foto ke-{i+1} dari {jml_foto}...")

                        # Eksekusi Upload (Berat)
                        url = upload_ke_dropbox(
                            f, nama_pelapor, "Laporan_Harian")

                        # Ambil deskripsi per foto jika ada
                        desc = st.session_state.get(
                            f"m_desc_{i}", "") or main_deskripsi or "-"

                        # Masukkan ke list rows
                        rows.append([
                            ts, nama_pelapor, final_lokasi, desc,
                            url, sosmed_link if sosmed_link else "-",
                            val_kesimpulan, val_kendala, val_kendala_klien,
                            val_pending, val_feedback, val_interest,
                            val_nama_klien, val_kontak_klien
                        ])

                        # Tambah counter langkah
                        current_step += 1
                else:
                    # Jika tidak ada foto, langsung siapkan 1 baris
                    rows.append([
                        ts, nama_pelapor, final_lokasi, main_deskripsi,
                        "-", sosmed_link if sosmed_link else "-",
                        val_kesimpulan, val_kendala, val_kendala_klien,
                        val_pending, val_feedback, val_interest,
                        val_nama_klien, val_kontak_klien
                    ])

                # --- 5. PROSES SIMPAN KE DATABASE (GSHEET) ---
                # Update bar ke langkah terakhir sebelum selesai
                pct_save = float(current_step / total_steps)
                if pct_save > 0.95:
                    pct_save = 0.95  # Biarkan sisa sedikit untuk efek selesai

                my_bar.progress(
                    pct_save, text="💾 Menyimpan data ke Database...")

                # Eksekusi Simpan (Berat)
                ok = simpan_laporan_harian_batch(rows, nama_pelapor)

                # --- 6. FINISHING ---
                # Set bar ke 100%
                my_bar.progress(1.0, text="✅ Selesai!")
                time.sleep(0.8)  # Jeda sebentar agar user lihat status 100%
                progress_placeholder.empty()  # Hapus bar agar bersih

                if ok:
                    st.success(
                        f"✅ Laporan tersimpan! Reminder: **{val_pending}**")
                    ui_toast("Laporan tersimpan!", icon="✅")

                    # Clear cache & Navigasi
                    st.cache_data.clear()
                    time.sleep(1)
                    set_nav("home")
                else:
                    st.error("Gagal menyimpan ke Database (GSheet).")

            except Exception as e:
                # Jika error, hapus bar dan tampilkan error
                progress_placeholder.empty()
                st.error(f"Terjadi kesalahan: {e}")


# =========================================================
# CLOSING DEAL
# =========================================================
@st.cache_data(ttl=3600)
def load_closing_deal():
    if not KONEKSI_GSHEET_BERHASIL:
        return pd.DataFrame(columns=CLOSING_COLUMNS)

    try:
        try:
            ws = spreadsheet.worksheet(SHEET_CLOSING_DEAL)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=SHEET_CLOSING_DEAL, rows=300, cols=len(CLOSING_COLUMNS))
            ws.append_row(CLOSING_COLUMNS, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
            return pd.DataFrame(columns=CLOSING_COLUMNS)

        ensure_headers(ws, CLOSING_COLUMNS)

        data = ws.get_all_records()
        df = pd.DataFrame(data)

        for c in CLOSING_COLUMNS:
            if c not in df.columns:
                df[c] = ""

        if COL_NILAI_KONTRAK in df.columns:
            parsed = df[COL_NILAI_KONTRAK].apply(parse_rupiah_to_int)
            df[COL_NILAI_KONTRAK] = pd.Series(parsed, dtype="Int64")

        for c in [COL_GROUP, COL_MARKETING, COL_TGL_EVENT, COL_BIDANG]:
            if c in df.columns:
                df[c] = df[c].fillna("").astype(str)

        return df[CLOSING_COLUMNS].copy()
    except Exception:
        return pd.DataFrame(columns=CLOSING_COLUMNS)


def tambah_closing_deal(nama_group, nama_marketing, tanggal_event, bidang, nilai_kontrak_input):
    if not KONEKSI_GSHEET_BERHASIL:
        return False, "Koneksi GSheet belum aktif."

    try:
        nama_group = str(nama_group).strip() if nama_group is not None else ""
        nama_marketing = str(nama_marketing).strip(
        ) if nama_marketing is not None else ""
        bidang = str(bidang).strip() if bidang is not None else ""

        if not nama_group:
            nama_group = "-"

        if not nama_marketing or not tanggal_event or not bidang or not str(nilai_kontrak_input).strip():
            return False, "Field wajib: Nama Marketing, Tanggal Event, Bidang, dan Nilai Kontrak."

        nilai_int = parse_rupiah_to_int(nilai_kontrak_input)
        if nilai_int is None:
            return False, "Nilai Kontrak tidak valid. Contoh: 15000000 / 15.000.000 / Rp 15.000.000 / 15jt / 15,5jt"

        try:
            ws = spreadsheet.worksheet(SHEET_CLOSING_DEAL)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=SHEET_CLOSING_DEAL, rows=300, cols=len(CLOSING_COLUMNS))
            ws.append_row(CLOSING_COLUMNS, value_input_option="USER_ENTERED")

        ensure_headers(ws, CLOSING_COLUMNS)

        tgl_str = tanggal_event.strftime(
            "%Y-%m-%d") if hasattr(tanggal_event, "strftime") else str(tanggal_event)

        ws.append_row([nama_group, nama_marketing, tgl_str, bidang, int(
            nilai_int)], value_input_option="USER_ENTERED")

        # maybe_auto_format_sheet(ws)
        return True, "Closing deal berhasil disimpan!"
    except Exception as e:
        return False, str(e)


# =========================================================
# PEMBAYARAN
# =========================================================
@st.cache_data(ttl=3600)
def load_pembayaran_dp():
    if not KONEKSI_GSHEET_BERHASIL:
        return pd.DataFrame(columns=PAYMENT_COLUMNS)

    try:
        try:
            ws = spreadsheet.worksheet(SHEET_PEMBAYARAN)
        except Exception:
            ws = spreadsheet.add_worksheet(
                title=SHEET_PEMBAYARAN, rows=500, cols=len(PAYMENT_COLUMNS))
            ws.append_row(PAYMENT_COLUMNS, value_input_option="USER_ENTERED")
            maybe_auto_format_sheet(ws, force=True)
            return pd.DataFrame(columns=PAYMENT_COLUMNS)

        ensure_headers(ws, PAYMENT_COLUMNS)

        data = ws.get_all_records()
        df = pd.DataFrame(data)

        for c in PAYMENT_COLUMNS:
            if c not in df.columns:
                df[c] = ""

        numeric_targets = [COL_NOMINAL_BAYAR, COL_NILAI_KESEPAKATAN, COL_SISA_BAYAR]
        for col in numeric_targets:
            if col in df.columns:
                df[col] = df[col].apply(lambda x: parse_rupiah_to_int(x) if isinstance(x, str) else x)
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

        if COL_NILAI_KESEPAKATAN in df.columns and COL_NOMINAL_BAYAR in df.columns:
            df[COL_SISA_BAYAR] = df[COL_NILAI_KESEPAKATAN] - df[COL_NOMINAL_BAYAR]
            df[COL_SISA_BAYAR] = df[COL_SISA_BAYAR].apply(lambda x: x if x > 0 else 0)

        if COL_STATUS_BAYAR in df.columns:
            df[COL_STATUS_BAYAR] = df[COL_STATUS_BAYAR].apply(
                lambda x: True if str(x).strip().upper() == "TRUE" else False)

        if COL_JATUH_TEMPO in df.columns:
            def smart_date_parser(x):
                s = str(x).strip()
                if not s or s.lower() in ["nan", "none", "-", ""]:
                    return pd.NaT
                try:
                    return pd.to_datetime(s, format="%Y-%m-%d").date()
                except:
                    try:
                        return pd.to_datetime(s, dayfirst=True).date()
                    except:
                        return pd.NaT
            
            df[COL_JATUH_TEMPO] = df[COL_JATUH_TEMPO].apply(smart_date_parser)

        text_cols = [COL_TS_BAYAR, COL_GROUP, COL_MARKETING, COL_TGL_EVENT, COL_JENIS_BAYAR,
                     COL_BUKTI_BAYAR, COL_CATATAN_BAYAR, COL_TS_UPDATE, COL_UPDATED_BY]
        for c in text_cols:
            if c in df.columns:
                df[c] = df[c].fillna("").astype(str)

        if COL_TS_UPDATE in df.columns:
            df[COL_TS_UPDATE] = df[COL_TS_UPDATE].apply(
                lambda x: build_numbered_log(parse_payment_log_lines(x)))

        return df[PAYMENT_COLUMNS].copy()

    except Exception as e:
        print(f"Error load_pembayaran_dp: {e}")
        return pd.DataFrame(columns=PAYMENT_COLUMNS)

def save_pembayaran_dp(df: pd.DataFrame) -> bool:
    try:
        ws = spreadsheet.worksheet(SHEET_PEMBAYARAN)
        ensure_headers(ws, PAYMENT_COLUMNS)

        ws.clear()

        rows_needed = len(df) + 1
        if ws.row_count < rows_needed:
            ws.resize(rows=rows_needed)

        df_save = df.copy()

        for c in PAYMENT_COLUMNS:
            if c not in df_save.columns:
                df_save[c] = ""

        df_save[COL_STATUS_BAYAR] = df_save[COL_STATUS_BAYAR].apply(
            lambda x: "TRUE" if bool(x) else "FALSE")

        def _to_int_or_blank(x):
            if x is None or pd.isna(x):
                return ""
            val = parse_rupiah_to_int(x)
            return "" if val is None else int(val)

        df_save[COL_NOMINAL_BAYAR] = df_save[COL_NOMINAL_BAYAR].apply(
            _to_int_or_blank)

        def _fmt_date(d):
            if d is None or pd.isna(d):
                return ""
            if hasattr(d, "strftime"):
                return d.strftime("%Y-%m-%d")
            s = str(d).strip()
            return s if s and s.lower() not in {"nan", "none"} else ""

        df_save[COL_JATUH_TEMPO] = df_save[COL_JATUH_TEMPO].apply(_fmt_date)

        df_save[COL_TS_UPDATE] = df_save[COL_TS_UPDATE].apply(
            lambda x: build_numbered_log(parse_payment_log_lines(x)))
        df_save[COL_UPDATED_BY] = df_save[COL_UPDATED_BY].apply(
            lambda x: safe_str(x, "-").strip() or "-")

        df_save = df_save[PAYMENT_COLUMNS].fillna("")
        data_to_save = [df_save.columns.values.tolist()] + \
            df_save.values.tolist()

        ws.update(range_name="A1", values=data_to_save,
                  value_input_option="USER_ENTERED")
        # maybe_auto_format_sheet(ws)
        return True
    except Exception:
        return False


def apply_audit_payments_changes(df_before: pd.DataFrame, df_after: pd.DataFrame, actor: str):
    """Update Timestamp Update (Log) & Updated By hanya untuk baris yang berubah."""
    if df_after is None or df_after.empty:
        return df_after

    actor = safe_str(actor, "-").strip() or "-"
    before = df_before.copy() if df_before is not None else pd.DataFrame()
    after = df_after.copy()

    for c in [COL_TS_UPDATE, COL_UPDATED_BY]:
        if c not in after.columns:
            after[c] = ""

    if before.empty or COL_TS_BAYAR not in before.columns or COL_TS_BAYAR not in after.columns:
        ts = now_ts_str()
        for i in range(len(after)):
            oldlog = after.at[i,
                              COL_TS_UPDATE] if COL_TS_UPDATE in after.columns else ""
            after.at[i, COL_TS_UPDATE] = append_payment_ts_update(
                oldlog, ts, actor, ["Data diperbarui (fallback)"])
            after.at[i, COL_UPDATED_BY] = actor
        return after

    before_idx = before.set_index(COL_TS_BAYAR, drop=False)
    after_idx = after.set_index(COL_TS_BAYAR, drop=False)

    watched_cols = [
        COL_JENIS_BAYAR,
        COL_NOMINAL_BAYAR,
        COL_JATUH_TEMPO,
        COL_STATUS_BAYAR,
        COL_BUKTI_BAYAR,
        COL_CATATAN_BAYAR,
    ]

    ts = now_ts_str()

    for key, row in after_idx.iterrows():
        if key not in before_idx.index:
            oldlog = safe_str(row.get(COL_TS_UPDATE, ""), "")
            if not safe_str(oldlog, "").strip():
                oldlog = build_numbered_log(
                    [safe_str(row.get(COL_TS_BAYAR, ts), ts)])
            after_idx.at[key, COL_TS_UPDATE] = oldlog
            after_idx.at[key, COL_UPDATED_BY] = actor
            continue

        prev = before_idx.loc[key]
        if isinstance(prev, pd.DataFrame):
            prev = prev.iloc[0]

        changes = []

        for col in watched_cols:
            if col not in after_idx.columns or col not in before_idx.columns:
                continue

            oldv = prev[col]
            newv = row[col]

            if col == COL_STATUS_BAYAR:
                if normalize_bool(oldv) != normalize_bool(newv):
                    changes.append(
                        f"Status Pembayaran: {_fmt_payment_val_for_log(col, oldv)} → {_fmt_payment_val_for_log(col, newv)}")
            elif col == COL_JATUH_TEMPO:
                if normalize_date(oldv) != normalize_date(newv):
                    changes.append(
                        f"Jatuh Tempo: {_fmt_payment_val_for_log(col, oldv)} → {_fmt_payment_val_for_log(col, newv)}")
            elif col == COL_NOMINAL_BAYAR:
                if parse_rupiah_to_int(oldv) != parse_rupiah_to_int(newv):
                    changes.append(
                        f"Nominal: {_fmt_payment_val_for_log(col, oldv)} → {_fmt_payment_val_for_log(col, newv)}")
            else:
                if safe_str(oldv, "").strip() != safe_str(newv, "").strip():
                    changes.append(
                        f"{col}: {_fmt_payment_val_for_log(col, oldv)} → {_fmt_payment_val_for_log(col, newv)}")

        if changes:
            oldlog = safe_str(prev.get(COL_TS_UPDATE, ""), "")
            newlog = append_payment_ts_update(oldlog, ts, actor, changes)
            after_idx.at[key, COL_TS_UPDATE] = newlog
            after_idx.at[key, COL_UPDATED_BY] = actor

    return after_idx.reset_index(drop=True)


def tambah_pembayaran_dp(nama_group, nama_marketing, tgl_event, jenis_bayar, nominal_input, total_sepakat_input, tenor, jatuh_tempo, bukti_file, catatan):
    """
    Menambah record pembayaran dengan sistem Smart Balance Tracking dan kalkulator cicilan transparan.
    """
    if not KONEKSI_GSHEET_BERHASIL: 
        return False, "Sistem Error: Koneksi Google Sheets tidak aktif."

    try:
        group = str(nama_group).strip() if nama_group else "-"
        marketing = str(nama_marketing).strip() if nama_marketing else "Unknown"
        catatan_clean = str(catatan).strip() if catatan else "-"
        
        nom_bayar = parse_rupiah_to_int(nominal_input) or 0
        total_sepakat = parse_rupiah_to_int(total_sepakat_input) or 0
        tenor_val = int(tenor) if tenor else 0
        
        if total_sepakat <= 0:
            return False, "Input Gagal: Total nilai kesepakatan harus diisi dengan benar."

        sisa_bayar = total_sepakat - nom_bayar
        
        info_cicilan = ""
        if tenor_val > 0 and sisa_bayar > 0:
            nilai_per_cicilan = sisa_bayar / tenor_val
            info_cicilan = f" | Cicilan: {format_rupiah_display(nilai_per_cicilan)} x{tenor_val} term"

        if sisa_bayar <= 0:
            status_fix = "✅ Lunas"
            if jenis_bayar == "Cash": 
                status_fix += " (Cash)"
        else:
            if jenis_bayar == "Down Payment (DP)":
                status_fix = f"⏳ DP (Sisa: {format_rupiah_display(sisa_bayar)}){info_cicilan}"
            elif jenis_bayar == "Cicilan":
                status_fix = f"💳 Cicilan (Sisa: {format_rupiah_display(sisa_bayar)}){info_cicilan}"
            else:
                status_fix = f"⚠️ Belum Lunas (Sisa: {format_rupiah_display(sisa_bayar)}){info_cicilan}"

        link_bukti = "-"
        if bukti_file and KONEKSI_DROPBOX_BERHASIL:
            link_bukti = upload_ke_dropbox(bukti_file, marketing, kategori="Bukti_Pembayaran")

        ts_in = now_ts_str()
        
        fmt_tgl_event = tgl_event.strftime("%Y-%m-%d") if hasattr(tgl_event, "strftime") else str(tgl_event)
        fmt_jatuh_tempo = jatuh_tempo.strftime("%Y-%m-%d") if hasattr(jatuh_tempo, "strftime") else str(jatuh_tempo)

        log_entry = f"[{ts_in}] Input Baru: {jenis_bayar}{info_cicilan}"

        row = [
            ts_in,
            group,
            marketing,
            fmt_tgl_event,
            total_sepakat,
            jenis_bayar,
            nom_bayar,
            tenor_val,
            sisa_bayar,
            fmt_jatuh_tempo,
            status_fix,
            link_bukti,
            catatan_clean,
            build_numbered_log([log_entry]),
            marketing
        ]

        ws = spreadsheet.worksheet(SHEET_PEMBAYARAN)
        ensure_headers(ws, PAYMENT_COLUMNS)
        
        ws.append_row(row, value_input_option="USER_ENTERED")
        
        msg_feedback = f"Pembayaran berhasil disimpan! "
        if sisa_bayar > 0:
            msg_feedback += f"Sisa tagihan: {format_rupiah_display(sisa_bayar)} dengan rincian {info_cicilan.replace(' | ', '')}."
        else:
            msg_feedback += "Status: LUNAS."

        return True, msg_feedback

    except Exception as e:
        return False, f"System Error: {str(e)}"


def build_alert_pembayaran(df: pd.DataFrame, days_due_soon: int = 3):
    if df is None or df.empty:
        return pd.DataFrame(columns=PAYMENT_COLUMNS), pd.DataFrame(columns=PAYMENT_COLUMNS)

    today = datetime.now(tz=TZ_JKT).date()
    df_alert = df.copy()

    if COL_JATUH_TEMPO in df_alert.columns:
        def ensure_date_obj(x):
            if isinstance(x, (datetime, pd.Timestamp)):
                return x.date()
            if isinstance(x, date):
                return x
            return pd.NaT
        
        df_alert[COL_JATUH_TEMPO] = df_alert[COL_JATUH_TEMPO].apply(ensure_date_obj)

    if COL_SISA_BAYAR in df_alert.columns:
        df_alert[COL_SISA_BAYAR] = pd.to_numeric(df_alert[COL_SISA_BAYAR], errors='coerce').fillna(0)
    else:
        df_alert[COL_SISA_BAYAR] = 0

    mask_aktif = (
        (df_alert[COL_SISA_BAYAR] > 100) & 
        (pd.notna(df_alert[COL_JATUH_TEMPO]))
    )
    
    if COL_STATUS_BAYAR in df_alert.columns:
        mask_aktif = mask_aktif & (df_alert[COL_STATUS_BAYAR] == False)

    df_tagihan_aktif = df_alert[mask_aktif].copy()

    if df_tagihan_aktif.empty:
        return pd.DataFrame(columns=df.columns), pd.DataFrame(columns=df.columns)

    overdue = df_tagihan_aktif[df_tagihan_aktif[COL_JATUH_TEMPO] <= today].copy()
    
    limit_date = today + timedelta(days=days_due_soon)
    due_soon = df_tagihan_aktif[
        (df_tagihan_aktif[COL_JATUH_TEMPO] > today) & 
        (df_tagihan_aktif[COL_JATUH_TEMPO] <= limit_date)
    ].copy()

    return overdue, due_soon


def update_bukti_pembayaran_by_index(row_index_0based: int, file_obj, nama_marketing: str, actor: str = "-"):
    if not KONEKSI_GSHEET_BERHASIL:
        return False, "Koneksi GSheet belum aktif."
    if not KONEKSI_DROPBOX_BERHASIL:
        return False, "Dropbox non-aktif. Upload bukti dimatikan."
    if file_obj is None:
        return False, "File bukti belum dipilih."

    try:
        ws = spreadsheet.worksheet(SHEET_PEMBAYARAN)
        ensure_headers(ws, PAYMENT_COLUMNS)

        link = upload_ke_dropbox(
            file_obj, nama_marketing or "Unknown", kategori="Bukti_Pembayaran")
        if not link or link == "-":
            return False, "Gagal upload ke Dropbox."

        headers = ws.row_values(1)
        row_gsheet = row_index_0based + 2

        if COL_BUKTI_BAYAR not in headers:
            return False, "Kolom 'Bukti Pembayaran' tidak ditemukan."
        col_bukti = headers.index(COL_BUKTI_BAYAR) + 1

        old_bukti = ""
        try:
            old_bukti = ws.cell(row_gsheet, col_bukti).value
        except Exception:
            old_bukti = ""

        cell_bukti = gspread.utils.rowcol_to_a1(row_gsheet, col_bukti)

        ts = now_ts_str()
        actor_final = safe_str(actor, "-").strip() or "-"

        updates = [{"range": cell_bukti, "values": [[link]]}]

        if COL_TS_UPDATE in headers:
            col_ts = headers.index(COL_TS_UPDATE) + 1
            old_log = ""
            try:
                old_log = ws.cell(row_gsheet, col_ts).value
            except Exception:
                old_log = ""
            new_log = append_payment_ts_update(
                old_log,
                ts,
                actor_final,
                [f"{COL_BUKTI_BAYAR}: {_fmt_payment_val_for_log(COL_BUKTI_BAYAR, old_bukti)} → {_fmt_payment_val_for_log(COL_BUKTI_BAYAR, link)}"]
            )
            cell_ts = gspread.utils.rowcol_to_a1(row_gsheet, col_ts)
            updates.append({"range": cell_ts, "values": [[new_log]]})

        if COL_UPDATED_BY in headers:
            col_by = headers.index(COL_UPDATED_BY) + 1
            cell_by = gspread.utils.rowcol_to_a1(row_gsheet, col_by)
            updates.append({"range": cell_by, "values": [[actor_final]]})

        ws.batch_update(updates, value_input_option="USER_ENTERED")
        # maybe_auto_format_sheet(ws)
        return True, "Bukti pembayaran berhasil di-update!"
    except Exception as e:
        return False, f"Error: {e}"


# =========================================================
# HEADER (LOGO LEFT/RIGHT + HOLDING BACKGROUND)
# =========================================================
ASSET_DIR = Path(__file__).parent / "assets"
LOGO_LEFT = ASSET_DIR / "log EO.png"
LOGO_RIGHT = ASSET_DIR / "logo traine.png"

# Logo holding tetap dipakai, tapi jadi logo mandiri di atas judul
LOGO_HOLDING = ASSET_DIR / "Logo-holding.png"

# Background hero diganti jadi sportarium
HERO_BG = ASSET_DIR / "sportarium.jpg"


def _img_to_base64(path: Path) -> str:
    try:
        if path and path.exists():
            return base64.b64encode(path.read_bytes()).decode("utf-8")
        return ""
    except Exception:
        return ""
    

# =========================================================
# [MIGRASI] CORE DATABASE & AUDIT PAYMENTS
# =========================================================
def apply_audit_payments_changes(df_before: pd.DataFrame, df_after: pd.DataFrame, actor: str):
    actor = safe_str(actor, "-").strip() or "-"
    before = df_before.copy() if df_before is not None else pd.DataFrame()
    after = df_after.copy()
    for c in [COL_TS_UPDATE, COL_UPDATED_BY]:
        if c not in after.columns: after[c] = ""
    
    if before.empty:
        ts = now_ts_str()
        for i in range(len(after)):
            after.at[i, COL_TS_UPDATE] = build_numbered_log([ts])
            after.at[i, COL_UPDATED_BY] = actor
        return after

    before_idx = before.set_index(COL_TS_BAYAR, drop=False)
    after_idx = after.set_index(COL_TS_BAYAR, drop=False)
    watched_cols = [COL_JENIS_BAYAR, COL_NOMINAL_BAYAR, COL_JATUH_TEMPO, COL_STATUS_BAYAR, COL_BUKTI_BAYAR, COL_CATATAN_BAYAR]
    ts = now_ts_str()

    for key, row in after_idx.iterrows():
        if key not in before_idx.index: continue
        prev = before_idx.loc[key]
        if isinstance(prev, pd.DataFrame): prev = prev.iloc[0]
        changes = []
        for col in watched_cols:
            if col not in after_idx.columns: continue
            oldv, newv = prev[col], row[col]
            if col == COL_STATUS_BAYAR:
                if normalize_bool(oldv) != normalize_bool(newv):
                    changes.append(f"Status: {_fmt_payment_val_for_log(col, oldv)} → {_fmt_payment_val_for_log(col, newv)}")
            elif col == COL_NOMINAL_BAYAR:
                if parse_rupiah_to_int(oldv) != parse_rupiah_to_int(newv):
                    changes.append(f"Nominal: {_fmt_payment_val_for_log(col, oldv)} → {_fmt_payment_val_for_log(col, newv)}")
            elif col == COL_JATUH_TEMPO:
                if normalize_date(oldv) != normalize_date(newv):
                    changes.append(f"Jatuh Tempo: {_fmt_payment_val_for_log(col, oldv)} → {_fmt_payment_val_for_log(col, newv)}")
            else:
                if safe_str(oldv).strip() != safe_str(newv).strip():
                    changes.append(f"{col}: {oldv} → {newv}")

        if changes:
            oldlog = safe_str(prev.get(COL_TS_UPDATE, ""), "")
            after_idx.at[key, COL_TS_UPDATE] = append_payment_ts_update(oldlog, ts, actor, changes)
            after_idx.at[key, COL_UPDATED_BY] = actor
    return after_idx.reset_index(drop=True)


def render_header():
    ts_now = datetime.now(tz=TZ_JKT).strftime("%d %B %Y %H:%M:%S")

    left_b64 = _img_to_base64(LOGO_LEFT)
    right_b64 = _img_to_base64(LOGO_RIGHT)
    holding_b64 = _img_to_base64(LOGO_HOLDING)  # Logo UMB
    bg_b64 = _img_to_base64(HERO_BG)

    g_on = bool(KONEKSI_GSHEET_BERHASIL)
    d_on = bool(KONEKSI_DROPBOX_BERHASIL)

    def pill(label: str, on: bool):
        cls = "sx-pill on" if on else "sx-pill off"
        return f"<span class='{cls}'><span class='sx-dot'></span>{label}</span>"

    # Style background hero (Sportarium)
    hero_style = (
        f"--hero-bg: url('data:image/jpeg;base64,{bg_b64}'); "
        f"--hero-bg-pos: 50% 72%; "
        f"--hero-bg-size: 140%;"
    ) if bg_b64 else "--hero-bg: none;"

    # Logo Kiri & Kanan (Mentari Sejuk)
    left_html = f"<img src='data:image/png;base64,{left_b64}' alt='Logo EO' />" if left_b64 else ""
    right_html = f"<img src='data:image/png;base64,{right_b64}' alt='Logo Training' />" if right_b64 else ""

    # --- BAGIAN BARU: Logo Holding di Paling Atas ---
    # Kita buat div terpisah di luar card utama
    top_logo_html = ""
    if holding_b64:
        top_logo_html = f"""
        <div style="display: flex; justify-content: center; margin-bottom: 25px; padding-top: 10px;">
            <img src='data:image/png;base64,{holding_b64}'
                 alt='Holding Logo'
                 style="height: 100px; width: auto; object-fit: contain; filter: drop-shadow(0 5px 15px rgba(0,0,0,0.5));" />
        </div>
        """

    # Susunan HTML: Logo Atas -> Baru kemudian Hero Card
    html = f"""
{top_logo_html}
<div class="sx-hero" style="{hero_style}">
<div class="sx-hero-grid">
<div class="sx-logo-card">{left_html}</div>
<div class="sx-hero-center">
<div class="sx-title">🚀 {APP_TITLE}</div>
<div class="sx-subrow">
<span>Realtime: {ts_now}</span>
{pill('GSheet: ON' if g_on else 'GSheet: OFF', g_on)}
{pill('Dropbox: ON' if d_on else 'Dropbox: OFF', d_on)}
</div>
</div>
<div class="sx-logo-card">{right_html}</div>
</div>
</div>
    """

    st.markdown(html, unsafe_allow_html=True)


def render_section_watermark():
    """
    Menampilkan watermark Sportarium di bagian bawah halaman/tab.
    Menggunakan file HERO_BG (sportarium.jpg) dengan style CSS .sx-section-watermark.
    """
    # Menggunakan aset global HERO_BG yang sudah didefinisikan di atas
    if not HERO_BG or not HERO_BG.exists():
        return

    b64 = _img_to_base64(HERO_BG)
    if not b64:
        return

    # Render HTML dengan class CSS yang sudah ada di inject_global_css
    html = f"""
    <div class="sx-section-watermark">
        <img src="data:image/jpeg;base64,{b64}" alt="Sportarium Watermark" />
    </div>
    """
    st.markdown(html, unsafe_allow_html=True)


def render_home_mobile():
    st.markdown("## 🧭 Menu Utama")
    st.caption("Pilih fitur seperti shortcut ala aplikasi mobile.")

    features = [
        {"key": "report",  "icon": "📝", "title": "Laporan Harian",
            "sub": "Input aktivitas + reminder"},
        {"key": "kpi",     "icon": "🎯", "title": "Target & KPI",
            "sub": "Checklist team & individu"},
        {"key": "closing", "icon": "🤝", "title": "Closing Deal",
            "sub": "Catat deal + export"},
        {"key": "payment", "icon": "💳", "title": "Pembayaran",
            "sub": "DP/Termin/Pelunasan + jatuh tempo"},
        {"key": "log",     "icon": "📜", "title": "Global Audit Log",
            "sub": "Riwayat perubahan data (Super Admin)"},
        {"key": "admin",   "icon": "🔐", "title": "Akses Admin",
            "sub": "Dashboard + kontrol (butuh login)"},
    ]

    cols = st.columns(2, gap="medium")
    for i, f in enumerate(features):
        with cols[i % 2]:
            with st.container(border=True):
                st.markdown(f"### {f['icon']} {f['title']}")
                st.caption(f["sub"])
                if st.button("Buka", use_container_width=True, key=f"home_open_{f['key']}"):
                    set_nav(f["key"])


# =========================================================
# APP UI
# =========================================================
if not KONEKSI_GSHEET_BERHASIL:
    st.error("Database Error.")
    st.stop()

# Small banner for Dropbox status
if not KONEKSI_DROPBOX_BERHASIL:
    st.warning("⚠️ Dropbox non-aktif. Fitur upload foto/bukti dimatikan.")

# =========================================================
# ROUTER NAV (untuk mobile ala "Facebook shortcut")
# =========================================================
HOME_NAV = "🏠 Beranda"

# Update: Menambahkan entry 'presensi' ke dalam Mapping
NAV_MAP = {
    "home": HOME_NAV,
    "presensi": "📅 Presensi",
    "report": "📝 Laporan Harian",
    "kpi": "🎯 Target & KPI",
    "closing": "🤝 Closing Deal",
    "payment": "💳 Pembayaran",
    "log": "📜 Global Audit Log",
    "admin": "📊 Dashboard Admin",
}


def _get_query_nav():
    try:
        # streamlit baru
        if hasattr(st, "query_params"):
            v = st.query_params.get("nav", None)
            if isinstance(v, (list, tuple)):
                return v[0] if v else None
            return v
        # streamlit lama
        qp = st.experimental_get_query_params()
        return (qp.get("nav", [None])[0])
    except Exception:
        return None


def set_nav(nav_key: str):
    nav_key = nav_key if nav_key in NAV_MAP else "home"
    try:
        if hasattr(st, "query_params"):
            st.query_params["nav"] = [nav_key]
        else:
            st.experimental_set_query_params(nav=nav_key)
    except Exception:
        pass
    st.session_state["menu_nav"] = NAV_MAP[nav_key]
    st.rerun()


# Session defaults
if "is_admin" not in st.session_state:
    st.session_state["is_admin"] = False

if "menu_nav" not in st.session_state:
    # Mobile masuk Beranda, Desktop tetap ke Laporan Harian
    st.session_state["menu_nav"] = HOME_NAV if IS_MOBILE else "📝 Laporan Harian"

# Sinkronkan kalau URL ada ?nav=...
nav_from_url = _get_query_nav()
if nav_from_url in NAV_MAP:
    st.session_state["menu_nav"] = NAV_MAP[nav_from_url]

# Render header
render_header()

# MOBILE: tampilkan Beranda sebagai landing page
menu_nav = st.session_state.get(
    "menu_nav", HOME_NAV if IS_MOBILE else "📝 Laporan Harian")

if IS_MOBILE and menu_nav == HOME_NAV:
    render_home_mobile()
    st.stop()

# =========================================================
# SIDEBAR (SpaceX-inspired)
# =========================================================
with st.sidebar:
    if st.button("🔄 Refresh Data", type="primary", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    st.markdown("<div class='sx-section-title'>Navigation</div>",
                unsafe_allow_html=True)

    # Update: Menambahkan "📅 Presensi" di daftar menu utama sidebar
    menu_items = [
        "📅 Presensi",
        "📝 Laporan Harian",
        "🎯 Target & KPI",
        "🤝 Closing Deal",
        "💳 Pembayaran",
        "📜 Global Audit Log",
    ]

    if st.session_state.get("is_admin"):
        menu_items.append("📊 Dashboard Admin")

    # SpaceX-like nav buttons
    st.markdown("<div class='sx-nav'>", unsafe_allow_html=True)
    for i, item in enumerate(menu_items):
        active = (st.session_state.get("menu_nav") == item)
        btype = "primary" if active else "secondary"
        if st.button(item, use_container_width=True, type=btype, key=f"nav_{i}"):
            st.session_state["menu_nav"] = item
            # Sync URL query param saat menu diklik
            nav_k = [k for k, v in NAV_MAP.items() if v == item]
            if nav_k:
                try:
                    if hasattr(st, "query_params"):
                        st.query_params["nav"] = nav_k[0]
                    else:
                        st.experimental_set_query_params(nav=nav_k[0])
                except:
                    pass
            st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)
    # -----------------------------------------------------------
    # PROFIL USER (OTP LOGIN)
    # -----------------------------------------------------------
    st.divider()

    col_p1, col_p2 = st.columns([1, 3])

    with col_p1:
        # Icon default karena OTP tidak ambil foto profil Google
        st.markdown("👤")

    with col_p2:
        st.caption("Login sebagai:")
        st.markdown(f"**{st.session_state.get('user_name', 'User')}**")

        role_now = st.session_state.get("user_role", "user")
        role_color = "red" if role_now == "admin" else "blue"
        st.markdown(f":{role_color}[{role_now.upper()}]")

    # Tombol Logout Manual (Reset State)
    if st.button("🚪 Sign Out / Logout", use_container_width=True):
        # Reset semua variabel sesi yang penting
        st.session_state["logged_in"] = False
        st.session_state["user_email"] = None
        st.session_state["user_name"] = None
        st.session_state["user_role"] = None
        st.session_state["is_admin"] = False

        # Reset step OTP agar kembali ke input email saat login ulang
        st.session_state["otp_step"] = 1
        st.session_state["temp_email"] = ""
        st.session_state["generated_otp"] = ""

        st.rerun()

    st.divider()

    # Quick stats (lightweight)
    try:
        df_pay_sidebar = load_pembayaran_dp()
        overdue_s, due_soon_s = build_alert_pembayaran(
            df_pay_sidebar, days_due_soon=3) if not df_pay_sidebar.empty else (pd.DataFrame(), pd.DataFrame())
        st.markdown("<div class='sx-section-title'>Quick Stats</div>",
                    unsafe_allow_html=True)
        st.metric("Overdue Payment", int(len(overdue_s))
                  if overdue_s is not None else 0)
        st.metric("Due ≤ 3 hari", int(len(due_soon_s))
                  if due_soon_s is not None else 0)
    except Exception:
        pass

    st.divider()
    st.caption("Tip: navigasi ala SpaceX → ringkas, jelas, fokus.")


menu_nav = st.session_state.get("menu_nav", "📝 Laporan Harian")

menu_nav = st.session_state.get("menu_nav", "📝 Laporan Harian")

# [MULAI KODE TAMBAHAN: FIX NAVIGASI MOBILE]
# Ini akan memunculkan tombol Back & Menu Bawah untuk Closing, KPI, Payment, dll.
if IS_MOBILE and menu_nav != "📝 Laporan Harian":
    # 1. Tombol Kembali ke Beranda
    if st.button("⬅️ Kembali ke Beranda", use_container_width=True, key="global_mobile_back"):
        set_nav("home")

    # 2. Bottom Navigation Bar (Menu Bawah)
    # Perbaikan: Menambahkan link nav=log dan merapikan tag HTML
    st.markdown("""
    <div class="mobile-bottom-nav">
      <a href="?nav=home">🏠</a>
      <a href="?nav=report">📝</a>
      <a href="?nav=kpi">🎯</a>
      <a href="?nav=closing">🤝</a>
      <a href="?nav=payment">💳</a>
      <a href="?nav=log">📜</a>
    </div>
    """, unsafe_allow_html=True)

    st.divider()


# =========================================================
# FUNGSI RENDER MOBILE PER FITUR (BARU)
# =========================================================
def render_kpi_mobile():
    st.markdown("### 🎯 Target & KPI (Full Mobile)")

    # Gunakan Tabs seperti Desktop agar fitur lengkap
    tab1, tab2, tab3 = st.tabs(["🏆 Team", "⚡ Individu", "⚙️ Admin"])

    # --- TAB 1: TEAM ---
    with tab1:
        st.caption("Checklist & Upload Bukti Team")
        df_team = load_checklist(SHEET_TARGET_TEAM, TEAM_CHECKLIST_COLUMNS)

        if not df_team.empty:
            # 1. Editor (Bisa Edit Status/Text)
            edited_team = render_hybrid_table(df_team, "mob_team_tbl", "Misi")

            # Tombol Simpan
            if st.button("💾 Simpan Perubahan (Team)", use_container_width=True, key="mob_btn_save_team"):
                actor = get_actor_fallback(default="Admin")
                final_df = apply_audit_checklist_changes(
                    df_team, edited_team, ["Misi"], actor)
                if save_checklist(SHEET_TARGET_TEAM, final_df, TEAM_CHECKLIST_COLUMNS):
                    st.success("Tersimpan!")
                    st.rerun()

            st.divider()

            # 2. Upload Bukti (Fitur Desktop dibawa ke HP)
            with st.expander("📂 Upload Bukti / Catatan"):
                sel_misi = st.selectbox(
                    "Pilih Misi", df_team["Misi"].unique(), key="mob_sel_misi")
                note_misi = st.text_area("Catatan", key="mob_note_misi")
                file_misi = st.file_uploader("File", key="mob_file_misi")

                if st.button("Update Bukti", use_container_width=True, key="mob_upd_team"):
                    actor = get_actor_fallback()
                    res, msg = update_evidence_row(
                        SHEET_TARGET_TEAM, sel_misi, note_misi, file_misi, actor, "Team")
                    if res:
                        st.success("Updated!")
                        st.rerun()
                    else:
                        st.error(msg)
        else:
            st.info("Belum ada target team.")

    # --- TAB 2: INDIVIDU ---
    with tab2:
        st.caption("Target Individu")
        staff = get_daftar_staf_terbaru()
        filter_nama = st.selectbox("Filter Nama:", staff, key="mob_indiv_filter")

        df_indiv_all = load_checklist(SHEET_TARGET_INDIVIDU, INDIV_CHECKLIST_COLUMNS)
        df_user = df_indiv_all[df_indiv_all["Nama"] == filter_nama]

        if not df_user.empty:
            # --- TAMBAHKAN LOGIKA PROGRES DI SINI ---
            total_target = len(df_user)
            # Menghitung jumlah 'TRUE' pada kolom Status
            jumlah_selesai = df_user["Status"].sum() 
            
            # Hitung persentase
            persentase = (jumlah_selesai / total_target) if total_target > 0 else 0
            
            # Tampilkan Progress Bar yang Estetik
            st.markdown(f"### 📈 Progres Kerja: {int(persentase * 100)}%")
            st.progress(persentase)
            st.write(f"Selesai: **{jumlah_selesai}** dari **{total_target}** tugas.")
            st.divider()
            # --- END LOGIKA PROGRES ---

            edited_indiv = render_hybrid_table(df_user, f"mob_indiv_{filter_nama}", "Target")

            if st.button(f"💾 Simpan ({filter_nama})", use_container_width=True, key="mob_save_indiv"):
                df_merged = df_indiv_all.copy()
                df_merged.update(edited_indiv)
                final_df = apply_audit_checklist_changes(
                    df_indiv_all, df_merged, ["Nama", "Target"], filter_nama)
                save_checklist(SHEET_TARGET_INDIVIDU, final_df,
                               INDIV_CHECKLIST_COLUMNS)
                st.success("Tersimpan!")
                st.rerun()

            # Upload Bukti Individu
            with st.expander(f"📂 Update Bukti ({filter_nama})"):
                pilih_target = st.selectbox(
                    "Target:", df_user["Target"].tolist(), key="mob_sel_indiv")
                note_target = st.text_area("Catatan", key="mob_note_indiv")
                file_target = st.file_uploader("File", key="mob_file_indiv")
                if st.button("Update Pribadi", use_container_width=True, key="mob_upd_indiv"):
                    res, msg = update_evidence_row(
                        SHEET_TARGET_INDIVIDU, pilih_target, note_target, file_target, filter_nama, "Individu")
                    if res:
                        st.success("Updated!")
                        st.rerun()
                    else:
                        st.error(msg)
        else:
            st.info("Kosong.")

    # --- TAB 3: ADMIN (Fitur Tambah Target) ---
    with tab3:
        st.markdown("#### ➕ Tambah Target Baru")
        jenis_t = st.radio(
            "Jenis", ["Team", "Individu"], horizontal=True, key="mob_jenis_target")

        with st.form("mob_add_kpi"):
            target_text = st.text_area("Isi Target (1 per baris)", height=100)
            c1, c2 = st.columns(2)
            t_mulai = c1.date_input("Mulai", value=datetime.now())
            t_selesai = c2.date_input(
                "Selesai", value=datetime.now()+timedelta(days=30))

            nama_target = ""
            if jenis_t == "Individu":
                nama_target = st.selectbox(
                    "Staf:", get_daftar_staf_terbaru(), key="mob_add_staf_target")

            if st.form_submit_button("Tambah Target", use_container_width=True):
                targets = clean_bulk_input(target_text)
                sheet = SHEET_TARGET_TEAM if jenis_t == "Team" else SHEET_TARGET_INDIVIDU
                base = ["", str(t_mulai), str(t_selesai), "FALSE", "-"]
                if jenis_t == "Individu":
                    base = [nama_target] + base

                if add_bulk_targets(sheet, base, targets):
                    st.success("Berhasil!")
                    st.rerun()
                else:
                    st.error("Gagal.")


def render_closing_mobile():
    st.markdown("### 🤝 Closing Deal (Full Mobile)")

    # Form Input Tetap Sama
    with st.expander("➕ Input Deal Baru", expanded=False):
        with st.form("mob_form_closing"):
            cd_group = st.text_input("Nama Group (Opsional)")
            cd_marketing = st.selectbox(
                "Nama Marketing", get_daftar_staf_terbaru())
            cd_tgl = st.date_input("Tanggal Event")
            cd_bidang = st.text_input("Bidang", placeholder="F&B / Wedding")
            cd_nilai = st.text_input("Nilai (Rp)", placeholder="Contoh: 15jt")

            if st.form_submit_button("Simpan Deal", type="primary", use_container_width=True):
                res, msg = tambah_closing_deal(
                    cd_group, cd_marketing, cd_tgl, cd_bidang, cd_nilai)
                if res:
                    st.success(msg)
                    st.cache_data.clear()
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(msg)

    st.divider()
    st.markdown("#### 📋 Riwayat Lengkap & Download")

    df_cd = load_closing_deal()

    if not df_cd.empty:
        # 1. Tampilkan Statistik Singkat
        tot = df_cd[COL_NILAI_KONTRAK].sum(
        ) if COL_NILAI_KONTRAK in df_cd.columns else 0
        st.metric("Total Closing", format_rupiah_display(tot))

        # 2. Tampilkan Semua Data (Tanpa batasan .head)
        st.dataframe(df_cd, use_container_width=True, hide_index=True)

        # 3. Fitur Download (Excel & CSV) - Diaktifkan di Mobile
        c1, c2 = st.columns(2)
        with c1:
            if HAS_OPENPYXL:
                xb = df_to_excel_bytes(df_cd, sheet_name="Closing")
                if xb:
                    st.download_button("⬇️ Excel", data=xb, file_name="closing_mob.xlsx",
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                       use_container_width=True)
        with c2:
            csv = df_cd.to_csv(index=False).encode('utf-8')
            st.download_button("⬇️ CSV", data=csv, file_name="closing_mob.csv",
                               mime="text/csv", use_container_width=True)

        # 4. Grafik (Jika ada Plotly)
        if HAS_PLOTLY:
            with st.expander("📊 Lihat Grafik Performance"):
                try:
                    df_plot = df_cd.copy()
                    df_plot[COL_NILAI_KONTRAK] = df_plot[COL_NILAI_KONTRAK].fillna(
                        0).astype(int)
                    fig = px.bar(df_plot, x=COL_MARKETING, y=COL_NILAI_KONTRAK, color=COL_BIDANG,
                                 title="Total per Marketing")
                    st.plotly_chart(fig, use_container_width=True)
                except:
                    pass
    else:
        st.info("Belum ada data.")


def render_payment_mobile():
    st.markdown("### 💳 Pembayaran (Full Mobile)")
    
    # =========================================================
    # 1. FORM INPUT BARU
    # =========================================================
    with st.expander("➕ Input Pembayaran Baru", expanded=False):
        with st.form("mob_form_pay"):
            p_group = st.text_input("Group (Opsional)")
            # Menggunakan daftar staf terbaru agar konsisten dengan Config_Staf
            p_marketing = st.selectbox("Marketing", get_daftar_staf_terbaru())
            p_nominal = st.text_input("Nominal (Rp)", placeholder="Contoh: 15.000.000 atau 15jt")
            p_jenis = st.selectbox("Jenis", ["Down Payment (DP)", "Termin", "Pelunasan"])
            p_jatuh_tempo = st.date_input("Batas Waktu Bayar", value=datetime.now(tz=TZ_JKT).date() + timedelta(days=7))
            p_status = st.checkbox("Sudah Dibayar?")
            p_bukti = st.file_uploader("Upload Bukti Transfer", disabled=not KONEKSI_DROPBOX_BERHASIL)
            
            if st.form_submit_button("Simpan Pembayaran", type="primary", use_container_width=True):
                with st.spinner("Menyimpan data..."):
                    res, msg = tambah_pembayaran_dp(
                        p_group, p_marketing, datetime.now(tz=TZ_JKT), 
                        p_jenis, p_nominal, p_jatuh_tempo, p_status, p_bukti, "-"
                    )
                    if res:
                        st.success(msg)
                        st.cache_data.clear()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(msg)

    st.divider()

    # =========================================================
    # 2. LOAD DATA & SISTEM ALERT
    # =========================================================
    df_pay = load_pembayaran_dp()

    if not df_pay.empty:
        # Sistem Peringatan (Overdue & Due Soon)
        overdue, due_soon = build_alert_pembayaran(df_pay)

        # [UPDATE] TABEL RINCIAN UNTUK FOLLOW UP (MOBILE)
        
        # 1. ALERT OVERDUE (MERAH)
        if not overdue.empty:
            st.error(f"⛔ **{len(overdue)} TAGIHAN OVERDUE!**")
            # Expander otomatis terbuka (expanded=True) agar langsung terlihat
            with st.expander("📄 LIHAT DATA & KONTAK (Klik)", expanded=True):
                # Ambil kolom penting: Marketing, Klien, Sisa, Catatan
                df_ov_mob = overdue[[COL_MARKETING, COL_GROUP, COL_SISA_BAYAR, COL_CATATAN_BAYAR]].copy()
                # Format Rupiah
                df_ov_mob[COL_SISA_BAYAR] = df_ov_mob[COL_SISA_BAYAR].apply(format_rupiah_display)
                
                st.dataframe(
                    df_ov_mob,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        COL_MARKETING: st.column_config.TextColumn("Sales", width="small"),
                        COL_GROUP: st.column_config.TextColumn("Klien"),
                        COL_SISA_BAYAR: st.column_config.TextColumn("Sisa"),
                        COL_CATATAN_BAYAR: st.column_config.TextColumn("Kontak/WA")
                    }
                )

        # 2. ALERT JATUH TEMPO DEKAT (KUNING)
        if not due_soon.empty:
            st.warning(f"⚠️ **{len(due_soon)} Jatuh Tempo Dekat (≤3 Hari)**")
            with st.expander("📄 Lihat Detail", expanded=False):
                df_ds_mob = due_soon[[COL_GROUP, COL_SISA_BAYAR, COL_JATUH_TEMPO]].copy()
                df_ds_mob[COL_SISA_BAYAR] = df_ds_mob[COL_SISA_BAYAR].apply(format_rupiah_display)
                
                st.dataframe(
                    df_ds_mob,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        COL_GROUP: "Klien",
                        COL_SISA_BAYAR: "Sisa",
                        COL_JATUH_TEMPO: st.column_config.DateColumn("Tgl", format="DD/MM")
                    }
                )

        # =========================================================
        # 3. EDITOR DATA (Audit Log Otomatis)
        # =========================================================

        st.markdown("#### 📋 Edit Data & Cek Status")
        st.caption("Ubah status 'Lunas' atau 'Jatuh Tempo' langsung di tabel bawah ini.")

        # [UPDATE] 1. Format Data untuk Tampilan (Ada Titik & Rp)
        # Menggunakan helper payment_df_for_display yang sudah diperbarui
        df_view = payment_df_for_display(df_pay)
        
        # [UPDATE] 2. Konfigurasi Kolom
        # Kita set kolom uang sebagai TextColumn agar format "Rp 200.000" tidak diubah balik oleh Streamlit
        # Kolom lain (Status, Tanggal, dll) tetap menggunakan konfigurasi lama
        column_configs = {
            COL_STATUS_BAYAR: st.column_config.CheckboxColumn("Lunas?", width="small"),
            COL_JATUH_TEMPO: st.column_config.DateColumn("Jatuh Tempo", format="DD/MM/YYYY"),
            COL_BUKTI_BAYAR: st.column_config.LinkColumn("Bukti"),
            COL_TS_UPDATE: st.column_config.TextColumn("Riwayat Perubahan (Log)", disabled=True),
            # Kolom Uang (Disabled karena ini view mobile, edit status/tanggal saja)
            COL_NOMINAL_BAYAR: st.column_config.TextColumn("Nominal", disabled=True),
            COL_NILAI_KESEPAKATAN: st.column_config.TextColumn("Total Deal", disabled=True),
            COL_SISA_BAYAR: st.column_config.TextColumn("Sisa", disabled=True),
        }

        # Sesuai Code Lama: Batasi kolom yang boleh diubah staf via HP
        editable_cols = [COL_STATUS_BAYAR, COL_JATUH_TEMPO, COL_CATATAN_BAYAR]
        disabled_cols = [c for c in df_view.columns if c not in editable_cols]

        edited_pay_mob = st.data_editor(
            df_view,
            column_config=column_configs,
            disabled=disabled_cols,
            hide_index=True,
            use_container_width=True,
            key="editor_pay_mobile_final"
        )

        # Tombol Simpan Perubahan dengan Logic Deteksi Perubahan (Diff)
        if st.button("💾 Simpan Perubahan Data", type="primary", use_container_width=True):
            with st.spinner("Memproses perubahan & mencatat audit log..."):
                # Actor diambil dari sesi login (Staff/Admin)
                actor_name = st.session_state.get("user_name", "Mobile User")
                
                # [UPDATE] 3. Cleaning Data Sebelum Disimpan
                # Karena tampilan menggunakan Text (Rp ...), kita harus kembalikan ke Integer
                # agar saat dibandingkan dengan database asli (df_pay) tidak dianggap berbeda semua.
                
                df_clean_edit = edited_pay_mob.copy()
                
                # Daftar kolom uang yang perlu dibersihkan kembali menjadi angka
                cols_to_clean = [COL_NOMINAL_BAYAR, COL_NILAI_KESEPAKATAN, COL_SISA_BAYAR]
                for c in cols_to_clean:
                    if c in df_clean_edit.columns:
                        df_clean_edit[c] = df_clean_edit[c].apply(parse_rupiah_to_int)

                # Membandingkan data lama (df_pay) vs data baru yang sudah dibersihkan (df_clean_edit)
                # Menggunakan helper apply_audit_payments_changes dari kode lama
                final_df = apply_audit_payments_changes(df_pay, df_clean_edit, actor=actor_name)
                
                if save_pembayaran_dp(final_df):
                    st.success("✅ Perubahan database berhasil disimpan!")
                    st.cache_data.clear()
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error("❌ Gagal menyimpan ke Database GSheet.")

        st.divider()

        # =========================================================
        # 4. FITUR UPLOAD BUKTI SUSULAN
        # =========================================================
        with st.expander("📎 Upload Bukti (Susulan)", expanded=False):
            st.caption("Gunakan ini untuk menambah/mengganti foto bukti transfer.")
            df_pay_reset = df_pay.reset_index(drop=True)
            
            # Membuat list pilihan data agar user tidak salah pilih baris
            options = [
                f"{i+1}. {r[COL_MARKETING]} | {r[COL_GROUP]} ({format_rupiah_display(r[COL_NOMINAL_BAYAR])})" 
                for i, r in df_pay_reset.iterrows()
            ]
            
            sel_idx = st.selectbox("Pilih Data Pembayaran:", range(len(options)), 
                                  format_func=lambda x: options[x], key="mob_sel_susulan")

            file_susulan = st.file_uploader("Pilih File Bukti Baru", key="mob_file_susulan")

            if st.button("⬆️ Update Foto Bukti", use_container_width=True):
                if file_susulan:
                    marketing_name = df_pay_reset.iloc[sel_idx][COL_MARKETING]
                    actor_now = st.session_state.get("user_name", "Mobile User")
                    
                    ok, msg = update_bukti_pembayaran_by_index(sel_idx, file_susulan, marketing_name, actor=actor_now)
                    if ok:
                        st.success("✅ Bukti berhasil di-update!")
                        st.cache_data.clear()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(msg)
                else:
                    st.warning("Silakan pilih file terlebih dahulu.")
    else:
        st.info("Belum ada data pembayaran yang tercatat.")


def render_admin_mobile():
    st.markdown("### 🔐 Admin Dashboard (Full Mobile)")

    # 1. Cek Login
    if not st.session_state["is_admin"]:
        pwd = st.text_input(
            "Password Admin", type="password", key="mob_adm_pwd")
        if st.button("Login", use_container_width=True, key="mob_adm_login"):
            if verify_admin_password(pwd):
                st.session_state["is_admin"] = True
                st.rerun()
            else:
                st.error("Password salah.")
        return  # Stop disini kalau belum login

    # 2. Jika Sudah Login -> Tampilkan Dashboard Penuh
    if st.button("🔓 Logout", use_container_width=True, key="mob_adm_logout"):
        st.session_state["is_admin"] = False
        st.rerun()

    # --- LOADING DATA ---
    staff_list = get_daftar_staf_terbaru()
    df_all = load_all_reports(staff_list)

    if not df_all.empty:
        try:
            df_all[COL_TIMESTAMP] = pd.to_datetime(
                df_all[COL_TIMESTAMP], format="%d-%m-%Y %H:%M:%S", errors="coerce")
            df_all["Tgl"] = df_all[COL_TIMESTAMP].dt.date
            # Helper kategori sederhana
            df_all["Kat"] = df_all[COL_TEMPAT].apply(lambda x: "Digital" if any(
                k in str(x) for k in ["Digital", "Ads", "Konten"]) else "Sales")
        except:
            pass

# TABS NAVIGATION MOBILE
    tab_prod, tab_leads, tab_data, tab_cfg = st.tabs(["📈 Grafik", "🧲 Leads", "📦 Data", "⚙️ Config"])

    with tab_prod:
        st.caption("Analisa Kinerja")
        if not df_all.empty:
            days = st.selectbox("Hari Terakhir:", [7, 30, 90], key="mob_adm_days")
            start_d = datetime.now(tz=TZ_JKT).date() - timedelta(days=days)
            df_f = df_all[df_all["Tgl"] >= start_d].copy()
            st.metric("Total Laporan", len(df_f))
            
            report_counts = df_f[COL_NAMA].value_counts()
            st.bar_chart(report_counts)

            st.divider()
            st.markdown("#### 🤖 AI / Machine Learning Management Insight")
            
            with st.spinner("Asisten Pak Nugroho sedang meninjau kinerja tim..."):
                try:
                    # Penyiapan Data
                    staf_stats_str = json.dumps(report_counts.to_dict(), indent=2)
                    
                    full_prompt = f"""
                    [CONTEXT_DATA]
                    Nama Pemimpin: Pak Nugroho
                    Total Laporan Masuk: {len(df_f)}
                    Statistik Per Staf: {staf_stats_str}

                    [SYSTEM_INSTRUCTION]
                    Kamu adalah asisten kepercayaan Pak Nugroho. Gunakan bahasa Indonesia yang santun, cerdas, namun tetap membumi agar mudah dipahami. 

                    PANDUAN PENULISAN:
                    1. Gunakan bahasa yang awam tapi berwibawa. Jangan gunakan istilah teknis yang terlalu berat dan jangan gunakan simbol em-dash atau sejenisnya.
                    2. JANGAN pernah menyebutkan target angka spesifik seperti 48 kunjungan.
                    3. Gunakan Analisis Perbandingan Kompetitor: Jelaskan bahwa saat sales di perusahaan lain mungkin hari ini masih sibuk di dalam kantor, terjebak urusan kertas, atau baru sekadar merencanakan jadwal, tim Pak Nugroho sudah mengambil langkah nyata di lapangan.
                    4. Gunakan Teori Keunggulan Awal: Tekankan bahwa satu laporan di awal waktu jauh lebih berharga daripada banyak laporan yang terlambat, karena ini adalah data nyata tentang kondisi pasar saat ini yang bisa langsung Bapak ambil kebijakannya.
                    5. Jika volume laporan sedikit, jelaskan dengan teori Kualitas di Atas Kuantitas: Sampaikan bahwa tim sedang melakukan pendekatan yang sangat mendalam ke klien besar, sehingga interaksinya lebih berkualitas daripada sekadar kunjungan formalitas.
                    6. Berikan apresiasi kepada staf yang sudah mengirim laporan (sebutkan namanya) sebagai bukti bahwa mereka lebih tanggap dan gesit dibanding rata-rata sales di luar sana.
                    7. JANGAN PERNAH mengaku sebagai AI atau Gemini. Tunjukkan empati dan semangat tinggi untuk mendukung visi Pak Nugroho.

                    [TASK]
                    Berikan analisis kinerja tim Sales kepada Pak Nugroho secara naratif dan kreatif berdasarkan data yang ada.
                    """

                    # Eksekusi Pemanggilan (Meniru logika Desktop)
                    ai_reply = ""
                    for model_name in MODEL_FALLBACKS:
                        try:
                            if SDK == "new":
                                resp = client_ai.models.generate_content(model=model_name, contents=full_prompt)
                                ai_reply = resp.text
                            else:
                                model = genai_legacy.GenerativeModel(model_name)
                                resp = model.generate_content(full_prompt)
                                ai_reply = resp.text
                            if ai_reply: break
                        except:
                            continue

                    if ai_reply:
                        st.info(ai_reply)
                    else:
                        # --- FALLBACK MECHANISM AKTIF ---
                        fallback_msg = generate_smart_insight_fallback(df_f, len(df_f))
                        st.warning(f"⚠️ Kuota AI sedang penuh. Beralih ke analisis statistik otomatis.\n\n{fallback_msg}")
                        
                except Exception as e:
                    # Jika terjadi error sistem, tetap jalankan fallback
                    fallback_msg = generate_smart_insight_fallback(df_f, len(df_f))
                    st.warning(f"⚠️ AI Maintenance. Fallback Insight:\n\n{fallback_msg}")
        else:
            st.info("Belum ada data laporan.")

    with tab_leads:
        st.caption("Filter & Download Leads")
        sel_int = st.selectbox("Interest:", ["Under 50% (A)", "50-75% (B)", "75%-100%"], key="mob_adm_int")
        if not df_all.empty and COL_INTEREST in df_all.columns:
            df_leads = df_all[df_all[COL_INTEREST].astype(str).str.strip() == sel_int]
            st.dataframe(df_leads[[COL_NAMA_KLIEN, COL_KONTAK_KLIEN]], use_container_width=True)
            if HAS_OPENPYXL:
                xb = df_to_excel_bytes(df_leads, sheet_name="Leads")
                if xb:
                    st.download_button("⬇️ Excel Leads", data=xb, file_name=f"leads_{sel_int}.xlsx", use_container_width=True)

    with tab_data:
        st.caption("Master Data Laporan")
        if st.button("Refresh Data", use_container_width=True, key="mob_ref_data"):
            st.cache_data.clear()
            st.rerun()
        st.dataframe(df_all, use_container_width=True)

    with tab_cfg:
        st.markdown("#### 👥 Kelola Personel (Staf)")
        with st.form("mob_add_staff"):
            st.markdown("➕ **Tambah Staf Baru**")
            new_st = st.text_input("Nama Staf", placeholder="Ketik nama baru...")
            if st.form_submit_button("Simpan Staf", use_container_width=True):
                if new_st.strip():
                    ok, msg = tambah_staf_baru(new_st)
                    if ok:
                        st.success("Berhasil ditambahkan!")
                        st.cache_data.clear()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(msg)
                else:
                    st.error("Nama tidak boleh kosong.")
        st.markdown("---") 
        st.markdown("#### 🗑️ Hapus Staf")
        st.caption("Menghapus nama dari daftar pelapor.")
        staff_now = get_daftar_staf_terbaru()
        hapus_select = st.selectbox("Pilih staf yang akan dihapus:", ["-- Pilih Staf --"] + staff_now, key="mob_del_st")
        confirm_del = st.checkbox("Konfirmasi penghapusan permanen", key="mob_del_confirm")
        if st.button("🔥 Konfirmasi Hapus", type="primary", use_container_width=True, key="mob_btn_del"):
            if hapus_select == "-- Pilih Staf --":
                st.error("Pilih nama staf terlebih dahulu!")
            elif not confirm_del:
                st.error("Silakan centang kotak konfirmasi penghapusan.")
            else:
                with st.spinner("Menghapus..."):
                    ok, m = hapus_staf_by_name(hapus_select)
                    if ok:
                        force_audit_log(actor=st.session_state.get("user_name", "Admin Mobile"), action="❌ DELETE USER", target_sheet="Config_Staf", chat_msg=f"Menghapus staf via HP: {hapus_select}", details_input=f"User {hapus_select} telah dihapus dari sistem mobile.")
                        st.success(f"Staf {hapus_select} Berhasil dihapus!")
                        st.cache_data.clear()
                        time.sleep(1.5)
                        st.rerun()
                    else:
                        st.error(m)

            # --- SUB-BAGIAN: HAPUS STAF ---
            st.markdown("#### 🗑️ Hapus Staf")
            st.caption("Menghapus nama dari daftar pelapor.")

            staff_now = get_daftar_staf_terbaru()
            hapus_select = st.selectbox("Pilih staf yang akan dihapus:", [
                                        "-- Pilih Staf --"] + staff_now, key="mob_del_st")

            confirm_del = st.checkbox(
                "Konfirmasi penghapusan permanen", key="mob_del_confirm")

            if st.button("🔥 Konfirmasi Hapus", type="primary", use_container_width=True, key="mob_btn_del"):
                if hapus_select == "-- Pilih Staf --":
                    st.error("Pilih nama staf terlebih dahulu!")
                elif not confirm_del:
                    st.error("Silakan centang kotak konfirmasi penghapusan.")
                else:
                    with st.spinner("Menghapus..."):
                        ok, m = hapus_staf_by_name(hapus_select)
                        if ok:
                            force_audit_log(
                                actor=st.session_state.get(
                                    "user_name", "Admin Mobile"),
                                action="❌ DELETE USER",
                                target_sheet="Config_Staf",
                                chat_msg=f"Menghapus staf via HP: {hapus_select}",
                                details_input=f"User {hapus_select} telah dihapus dari sistem mobile."
                            )
                            st.success(f"Staf {hapus_select} Berhasil dihapus!")
                            st.cache_data.clear()
                            time.sleep(1.5)
                            st.rerun()
                        else:
                            st.error(m)


def render_audit_mobile():
    st.markdown("### 📜 Global Audit Log (Mobile)")
    st.caption("Rekaman jejak perubahan data admin.")

    from audit_service import load_audit_log

    if st.button("🔄 Refresh", use_container_width=True, key="mob_refresh_log"):
        st.cache_data.clear()
        st.rerun()

    df_raw = load_audit_log(spreadsheet)

    if not df_raw.empty:
        # Gunakan mapper dinamis agar kolom terdeteksi otomatis
        df_log = dynamic_column_mapper(df_raw)

        # Sortir data terbaru
        try:
            df_log["Waktu"] = pd.to_datetime(df_log["Waktu"], errors="coerce")
            df_log = df_log.sort_values(by="Waktu", ascending=False)
        except:
            pass

        st.markdown("#### 🕒 10 Aktivitas Terakhir")

        for i, row in df_log.head(10).iterrows():
            with st.container(border=True):
                # Gunakan .get() agar aman jika kolom tetap tidak terdeteksi
                st.markdown(f"**{row.get('User', '-')}**")
                st.caption(
                    f"📅 {row.get('Waktu', '-')} | Status: {row.get('Status', '-')}")
                st.text(f"Data: {row.get('Target Data', '-')}")

                chat_val = row.get('Chat & Catatan', '-')
                if chat_val not in ["-", ""]:
                    st.info(f"📝 {chat_val}")

                with st.expander("Lihat Detail"):
                    st.code(row.get('Detail Perubahan', '-'), language="text")
    else:
        st.info("Belum ada data log.")

# =========================================================
# MAIN ROUTER LOGIC (REVISI TOTAL)
# =========================================================


# =========================================================
# MAIN ROUTER LOGIC: IMPLEMENTASI SELURUH FITUR
# =========================================================

# --- 1. HALAMAN PRESENSI (REAL-TIME + MASUK/PULANG + PHOTO UPLOAD) ---
if menu_nav == "📅 Presensi":
    st.markdown("## 📅 Presensi Kehadiran Real-Time")
    st.caption(
        "Pilih Nama, Tipe Absen (Masuk/Pulang), dan lampirkan foto selfie. Waktu akan tercatat otomatis oleh sistem (WIB).")

    with st.container(border=True):
        staff_list = get_daftar_staf_terbaru()
        pilih_nama = st.selectbox(
            "Pilih Nama Anda:", ["-- Pilih Nama --"] + staff_list, key="presensi_name_sel")

        # --- PILIHAN TIPE ABSEN ---
        # Membuat pilihan Masuk atau Pulang dengan Radio Button horizontal
        col_absen1, col_absen2 = st.columns(2)
        with col_absen1:
            tipe_absen = st.radio("Tipe Presensi:", ["Masuk", "Pulang"], horizontal=True, key="tipe_absen_radio")
        
        # Penentuan Icon dinamis berdasarkan pilihan
        icon_absen = "🚀" if tipe_absen == "Masuk" else "🏠"
        
        # --- FITUR UPLOAD FOTO ---
        input_foto = st.file_uploader(f"Ambil Foto Selfie {tipe_absen} (Kamera/Galeri)", 
                                     type=['png', 'jpg', 'jpeg'],
                                     help=f"Gunakan kamera HP untuk selfie saat jam {tipe_absen.lower()}.")
        
        if input_foto:
            # Menampilkan preview foto kecil
            st.image(input_foto, caption=f"Preview Selfie {tipe_absen}", width=150)

        # --- INFO WAKTU & TOMBOL KIRIM ---
        waktu_skrg = datetime.now(TZ_JKT)
        st.info(
            f"🕒 Waktu Sistem Saat Ini: **{waktu_skrg.strftime('%A, %d %B %Y - %H:%M:%S')} WIB**")

        # Tombol Kirim dengan label dinamis (Kirim Presensi Masuk / Kirim Presensi Pulang)
        if st.button(f"{icon_absen} Kirim Presensi {tipe_absen}", type="primary", use_container_width=True):
            if pilih_nama == "-- Pilih Nama --":
                st.error("Silakan pilih nama terlebih dahulu!")
            elif input_foto is None:
                st.error(f"Wajib melampirkan foto untuk presensi {tipe_absen.lower()}!")
            else:
                with st.spinner(f"Mencatat {tipe_absen.lower()} & mengupload foto..."):
                    # Memanggil fungsi catat_presensi yang sudah mendukung parameter 'tipe'
                    ok, msg = catat_presensi(pilih_nama, tipe=tipe_absen, file_foto=input_foto)
                    
                    if ok:
                        st.success(msg)
                        # Mencatat aktivitas ke Global Audit Log
                        force_audit_log(
                            actor=pilih_nama,
                            action=f"✅ {tipe_absen.upper()}",
                            target_sheet="Presensi_Kehadiran",
                            chat_msg=f"Presensi {tipe_absen} sukses.",
                            details_input=f"Jam: {waktu_skrg.strftime('%H:%M:%S')} | Bukti foto terlampir"
                        )
                        time.sleep(2)
                        st.rerun()
                    else:
                        # Pesan error jika gagal validasi (misal: belum masuk sudah mau pulang)
                        st.error(msg)

    st.divider()
    st.markdown("### 📋 Kehadiran Hari Ini")
    
    # Menampilkan riwayat kehadiran hari ini di bawah form agar staf tahu statusnya
    ws_p = init_presensi_db()
    if ws_p:
        data_p = ws_p.get_all_records()
        if data_p:
            df_p = pd.DataFrame(data_p)
            
            # Filter data khusus hari ini agar tampilan ringkas
            tgl_hari_ini = waktu_skrg.strftime("%d")
            bln_hari_ini = waktu_skrg.strftime("%B")
            thn_hari_ini = waktu_skrg.strftime("%Y")
            
            # Pastikan kolom Tanggal, Bulan, Tahun tersedia sesuai PRESENSI_COLUMNS
            df_today = df_p[
                (df_p['Tanggal'].astype(str) == tgl_hari_ini) & 
                (df_p['Bulan'] == bln_hari_ini) &
                (df_p['Tahun'].astype(str) == thn_hari_ini)
            ]
            
            if not df_today.empty:
                # Konfigurasi tabel agar Link Foto bisa langsung diklik
                st.dataframe(
                    df_today, 
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Link Foto": st.column_config.LinkColumn("📸 Lihat Foto"),
                        "Tipe Absen": st.column_config.TextColumn("Status", width="small"),
                        "Waktu": st.column_config.TextColumn("Jam", width="small")
                    }
                )
            else:
                st.info("Belum ada data kehadiran hari ini.")

# --- 2. HALAMAN LAPORAN HARIAN ---
elif menu_nav == "📝 Laporan Harian":
    if IS_MOBILE:
        render_laporan_harian_mobile()
    else:
        st.markdown("## 📝 Laporan Kegiatan Harian")
        c1, c2 = st.columns([1, 2])
        with c1:
            pelapor = st.selectbox(
                "Nama Pelapor", get_daftar_staf_terbaru(), key="pelapor_desk")
        with c2:
            pending = get_reminder_pending(pelapor)
            if pending:
                st.warning(f"🔔 Reminder Pending: {pending}")

        with st.container(border=True):
            with st.form("daily_report_desk", clear_on_submit=False):
                st.markdown("### 📌 Detail Aktivitas")
                col_kiri, col_kanan = st.columns(2)
                with col_kiri:
                    kategori = st.radio(
                        "Kategori", ["🚗 Sales Lapangan", "💻 Digital/Kantor", "📞 Telesales", "🏢 Lainnya"])
                    lokasi = st.text_input(
                        "Lokasi / Nama Klien / Jenis Tugas", placeholder="Wajib diisi...")
                    deskripsi = st.text_area("Deskripsi Detail", height=150)
                    foto = st.file_uploader(
                        "Upload Bukti", accept_multiple_files=True, disabled=not KONEKSI_DROPBOX_BERHASIL)
                with col_kanan:
                    st.markdown("### 📊 Hasil & Follow Up")
                    kesimpulan = st.text_area("Kesimpulan / Hasil", height=80)
                    kendala = st.text_area(
                        "Kendala Internal/Lapangan", height=60)
                    kendala_klien = st.text_area("Kendala dari Sisi Klien", height=60, placeholder="Misal: Budget belum turun, owner sedang keluar kota...")
                    next_plan = st.text_input("Next Plan / Pending (Reminder)")
                    st.markdown("### 👤 Data Klien")
                    cl_nama = st.text_input("Nama Klien")
                    cl_kontak = st.text_input("No HP/WA")
                    cl_interest = st.selectbox(
                        "Interest Level", ["-", "Under 50%", "50-75%", "75-100%"])
                st.divider()
                if st.form_submit_button("✅ KIRIM LAPORAN", type="primary", use_container_width=True):
                    if not lokasi or not deskripsi:
                        st.error("Lokasi dan Deskripsi wajib diisi!")
                    else:
                        with st.spinner("Mengirim laporan..."):
                            ts = now_ts_str()
                            final_link = "-"
                            if foto and KONEKSI_DROPBOX_BERHASIL:
                                links = [upload_ke_dropbox(
                                    f, pelapor, "Laporan_Harian") for f in foto]
                                final_link = ", ".join(links)
                            row_data = [ts, pelapor, lokasi, deskripsi, final_link, "-", kesimpulan,
                                        kendala,kendala_klien, "-", next_plan, "-", cl_interest, cl_nama, cl_kontak]
                            if simpan_laporan_harian_batch([row_data], pelapor):
                                st.success("Laporan Terkirim!")
                                st.cache_data.clear()
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error("Gagal simpan ke GSheet.")

# --- 3. TARGET & KPI ---
elif menu_nav == "🎯 Target & KPI":
    if IS_MOBILE:
        render_kpi_mobile()
    else:
        st.markdown("## 🎯 Manajemen Target & KPI")
        tab1, tab2, tab3 = st.tabs(
            ["🏆 Target Team", "⚡ Target Individu", "⚙️ Admin Setup"])
        with tab1:
            df_team = load_checklist(SHEET_TARGET_TEAM, TEAM_CHECKLIST_COLUMNS)
            if not df_team.empty:
                edited_team = render_hybrid_table(df_team, "team_desk", "Misi")
                if st.button("💾 Simpan Perubahan Team"):
                    final_df = apply_audit_checklist_changes(
                        df_team, edited_team, ["Misi"], get_actor_fallback())
                    save_checklist(SHEET_TARGET_TEAM, final_df,
                                   TEAM_CHECKLIST_COLUMNS)
                    st.success("Tersimpan!")
                    st.cache_data.clear()
                    st.rerun()
        with tab2:
            st.caption("Monitoring target perorangan.")
            pilih_staf = st.selectbox(
                "Pilih Nama Staf:", get_daftar_staf_terbaru())
            
            df_indiv_all = load_checklist(
                SHEET_TARGET_INDIVIDU, INDIV_CHECKLIST_COLUMNS)
            df_user = df_indiv_all[df_indiv_all["Nama"] == pilih_staf]
            
            if not df_user.empty:
                # ==========================================
                # ANCHOR: LOGIKA PROGRESS BAR (PENTING)
                # ==========================================
                total_target = len(df_user)
                # Menghitung jumlah baris yang statusnya dicentang (True)
                jumlah_selesai = df_user["Status"].sum() 
                persentase = jumlah_selesai / total_target if total_target > 0 else 0
                
                # Tampilan Visual Progres
                st.markdown(f"### 📈 Progres {pilih_staf}: {int(persentase * 100)}%")
                st.progress(persentase)
                st.write(f"✅ **{jumlah_selesai}** selesai dari **{total_target}** target.")
                st.divider()
                # ==========================================

                # Tabel editor untuk mencentang target
                edited_indiv = render_hybrid_table(
                    df_user, f"indiv_{pilih_staf}", "Target")
                
                if st.button(f"💾 Simpan Target {pilih_staf}", use_container_width=True):
                    df_merged = df_indiv_all.copy()
                    
                    # Update data lama dengan data hasil editan tabel
                    df_merged.update(edited_indiv)
                    
                    final_df = apply_audit_checklist_changes(
                        df_indiv_all, df_merged, ["Nama", "Target"], pilih_staf)
                    
                    if save_checklist(SHEET_TARGET_INDIVIDU, final_df, INDIV_CHECKLIST_COLUMNS):
                        st.success(f"Berhasil menyimpan progres {pilih_staf}!")
                        st.cache_data.clear()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error("Gagal menyimpan ke database.")
            else:
                st.info(f"Belum ada target yang ditugaskan untuk {pilih_staf}.")
        with tab3:
            st.markdown("### ➕ Tambah Target Baru")
            jenis_t = st.radio(
                "Jenis Target", ["Team", "Individu"], horizontal=True)
            with st.form("add_kpi_desk"):
                target_text = st.text_area("Isi Target (1 per baris)")
                tgl_m = st.date_input("Mulai", value=datetime.now())
                tgl_s = st.date_input(
                    "Selesai", value=datetime.now()+timedelta(days=30))
                nama_t = st.selectbox(
                    "Untuk Staf:", get_daftar_staf_terbaru()) if jenis_t == "Individu" else ""
                if st.form_submit_button("Tambah Target"):
                    targets = clean_bulk_input(target_text)
                    sheet = SHEET_TARGET_TEAM if jenis_t == "Team" else SHEET_TARGET_INDIVIDU
                    base = ["", str(tgl_m), str(tgl_s), "FALSE", "-"]
                    if jenis_t == "Individu":
                        base = [nama_t] + base
                    if add_bulk_targets(sheet, base, targets):
                        st.success("Berhasil!")
                        st.cache_data.clear()
                        st.rerun()

# --- 4. CLOSING DEAL ---
elif menu_nav == "🤝 Closing Deal":
    if IS_MOBILE:
        render_closing_mobile()
    else:
        st.markdown("## 🤝 Closing Deal")
        with st.container(border=True):
            with st.form("form_closing_desk_full", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                inp_group = c1.text_input("Nama Group (Opsional)")
                inp_marketing = c2.text_input("Nama Marketing")
                inp_tgl_event = c3.date_input(
                    "Tanggal Event", value=datetime.now(tz=TZ_JKT).date())
                inp_bidang = st.text_input("Bidang / Jenis Event")
                inp_nilai = st.text_input("Nilai Kontrak (Rupiah)")
                if st.form_submit_button("✅ Simpan Closing Deal", type="primary", use_container_width=True):
                    res, msg = tambah_closing_deal(
                        inp_group, inp_marketing, inp_tgl_event, inp_bidang, inp_nilai)
                    if res:
                        st.success(msg)
                        st.cache_data.clear()
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(msg)
        st.divider()
        df_cd = load_closing_deal()
        if not df_cd.empty:
            st.dataframe(df_cd, use_container_width=True, hide_index=True)

# --- 5. PEMBAYARAN ---
elif menu_nav == "💳 Pembayaran":
    if IS_MOBILE:
        render_payment_mobile()
    else:
        st.markdown("## 💳 Smart Payment Action Center")
        st.caption("Manajemen pembayaran terpadu dengan kalkulator sisa tagihan dan pelacakan cicilan.")

        # =========================================================
        # 1. SEKSI INPUT: KALKULATOR PEMBAYARAN PINTAR
        # =========================================================
        with st.container(border=True):
            st.markdown("### ➕ Input Pembayaran & Kalkulator Sisa")
            with st.form("form_smart_pay", clear_on_submit=True):
                c1, c2, c3 = st.columns(3)
                with c1:
                    p_marketing = st.selectbox("Nama Marketing", get_daftar_staf_terbaru())
                    p_group = st.text_input("Nama Group / Klien", placeholder="Masukkan nama entitas...")
                    p_total_sepakat = st.text_input("Total Nilai Kesepakatan (Rp)", placeholder="Contoh: 100.000.000")
                
                with c2:
                    p_jenis = st.selectbox("Mekanisme Pembayaran", ["Down Payment (DP)", "Cicilan", "Cash"])
                    p_nom_bayar = st.text_input("Nominal yang Dibayar Sekarang (Rp)")
                    p_tenor = st.number_input("Tenor Cicilan (Bulan)", min_value=0, step=1, help="Isi 0 jika pembayaran Cash/DP sekali bayar")
                
                with c3:
                    p_tgl_event = st.date_input("Tanggal Event", value=datetime.now(tz=TZ_JKT).date())
                    p_due = st.date_input("Batas Waktu Bayar (Jatuh Tempo)", value=datetime.now(tz=TZ_JKT).date() + timedelta(days=7))
                    p_bukti = st.file_uploader("Upload Bukti Transfer (Foto/PDF)")

                p_note = st.text_area("Catatan Tambahan (Opsional)", placeholder="Keterangan bank, nomor referensi, dll.")
                
                if st.form_submit_button("✅ Simpan & Hitung Sisa", type="primary", use_container_width=True):
                    if not p_total_sepakat or not p_nom_bayar:
                        st.error("Gagal: Nilai Kesepakatan dan Nominal Bayar wajib diisi!")
                    else:
                        with st.spinner("Memproses transaksi..."):
                            ok, msg = tambah_pembayaran_dp(
                                p_group, p_marketing, p_tgl_event, p_jenis, 
                                p_nom_bayar, p_total_sepakat, p_tenor, p_due, p_bukti, p_note
                            )
                            if ok:
                                st.success(msg)
                                st.cache_data.clear()
                                time.sleep(2)
                                st.rerun()
                            else:
                                st.error(msg)

        st.divider()

        # =========================================================
        # 2. SEKSI MONITORING: ALERT & DATA EDITOR DINAMIS
        # =========================================================
        st.markdown("### 📋 Monitoring & Riwayat Pembayaran")
        df_pay = load_pembayaran_dp()

        if df_pay.empty:
            st.info("Belum ada data pembayaran yang tersimpan.")
        else:
            # --- Sistem Alert Pintar (Berdasarkan Sisa Bayar) ---
            overdue, due_soon = build_alert_pembayaran(df_pay)
            col_stat1, col_stat2 = st.columns(2)
            
            # Tampilkan Angka Summary
            with col_stat1:
                st.metric("⛔ Overdue (Belum Lunas)", len(overdue))
            with col_stat2:
                st.metric("⚠️ Jatuh Tempo Dekat (≤ 3 Hari)", len(due_soon))

            # [UPDATE] TABEL DETAIL UNTUK ADMIN (DESKTOP)
            # Ditampilkan lebar penuh di bawah angka statistik
            
            if not overdue.empty:
                st.error(f"🚨 **PERHATIAN: Ada {len(overdue)} Tagihan Lewat Jatuh Tempo!**")
                with st.expander("🔴 KLIK UNTUK LIHAT DAFTAR PENAGIHAN & KONTAK WA", expanded=True):
                    # Menampilkan kolom lengkap: Sales, Klien, Tanggal, Sisa Uang, dan Catatan Kontak
                    df_ov_desk = overdue[[COL_MARKETING, COL_GROUP, COL_JATUH_TEMPO, COL_SISA_BAYAR, COL_CATATAN_BAYAR]].copy()
                    # Format angka jadi Rupiah
                    df_ov_desk[COL_SISA_BAYAR] = df_ov_desk[COL_SISA_BAYAR].apply(format_rupiah_display)
                    
                    st.dataframe(
                        df_ov_desk,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            COL_MARKETING: st.column_config.TextColumn("👤 Sales PIC", width="medium"),
                            COL_GROUP: st.column_config.TextColumn("🏢 Nama Klien", width="medium"),
                            COL_JATUH_TEMPO: st.column_config.DateColumn("📅 Jatuh Tempo", format="DD MMM YYYY"),
                            COL_SISA_BAYAR: st.column_config.TextColumn("💰 Sisa Tagihan", width="medium"),
                            COL_CATATAN_BAYAR: st.column_config.TextColumn("📞 Catatan / Kontak WA", width="large")
                        }
                    )

            if not due_soon.empty:
                st.warning(f"🔔 **REMINDER: {len(due_soon)} Tagihan akan jatuh tempo dalam 3 hari.**")
                with st.expander("🟡 LIHAT DAFTAR DUE SOON", expanded=True):
                    df_soon_desk = due_soon[[COL_MARKETING, COL_GROUP, COL_JATUH_TEMPO, COL_SISA_BAYAR, COL_CATATAN_BAYAR]].copy()
                    df_soon_desk[COL_SISA_BAYAR] = df_soon_desk[COL_SISA_BAYAR].apply(format_rupiah_display)
                    
                    st.dataframe(
                        df_soon_desk,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            COL_MARKETING: st.column_config.TextColumn("👤 Sales PIC", width="medium"),
                            COL_GROUP: st.column_config.TextColumn("🏢 Nama Klien", width="medium"),
                            COL_JATUH_TEMPO: st.column_config.DateColumn("📅 Jatuh Tempo", format="DD MMM YYYY"),
                            COL_SISA_BAYAR: st.column_config.TextColumn("💰 Sisa Tagihan", width="medium"),
                            COL_CATATAN_BAYAR: st.column_config.TextColumn("📞 Catatan / Kontak WA", width="large")
                        }
                    )

            st.divider()
            st.caption("Klik dua kali pada sel tabel utama di bawah ini untuk mengedit data pembayaran.")

            # --- 1. PROSES DATA SECARA DINAMIS (Normalisasi Tipe Data) ---
            df_ready = clean_df_types_dynamically(df_pay)
            
            # --- 2. GENERATE CONFIG SECARA OTOMATIS (Mencegah Error Tipe Data) ---
            auto_config = generate_dynamic_column_config(df_ready)
            
            # --- 3. LOGIKA PENGUNCIAN KOLOM OTOMATIS ---
            locked_keywords = ["timestamp", "updated by", "log", "pelaku", "waktu", "input"]
            disabled_list = [c for c in df_ready.columns if any(k in c.lower() for k in locked_keywords)]

            edited_pay = st.data_editor(
                df_ready,
                column_config=auto_config,
                # ... sisanya tetap
            )

            # --- 4. RENDER DATA EDITOR ---
            edited_pay = st.data_editor(
                df_ready,
                column_config=auto_config,
                disabled=disabled_list,
                hide_index=True,
                use_container_width=True,
                num_rows="dynamic",
                key="smart_payment_editor_v3"
            )

            # --- 5. LOGIKA SIMPAN PERUBAHAN ---
            if st.button("💾 Simpan Perubahan Riwayat", use_container_width=True):
                with st.spinner("Memproses audit log dan menyimpan data..."):
                    current_user = st.session_state.get("user_name", "Admin Desktop")
                    
                    # Bandingkan data lama (df_pay) dengan data yang diedit (edited_pay)
                    final_df = apply_audit_payments_changes(df_pay, edited_pay, actor=current_user)
                    
                    if save_pembayaran_dp(final_df):
                        st.success("✅ Perubahan database berhasil disimpan!")
                        st.cache_data.clear()
                        time.sleep(1.5)
                        st.rerun()
                    else:
                        st.error("Gagal menyimpan ke Google Sheets.")

            st.divider()

            # =========================================================
            # 3. FITUR TAMBAHAN: UPDATE FOTO BUKTI SUSULAN
            # =========================================================
            with st.expander("📎 Update Bukti Pembayaran (Susulan)", expanded=False):
                st.info("Gunakan fitur ini jika ingin menambahkan atau mengganti foto bukti transfer tanpa mengubah data lainnya.")
                df_pay_reset = df_pay.reset_index(drop=True)
                
                pay_options = [
                    f"{i+1}. {r[COL_MARKETING]} | {r[COL_GROUP]} | Sisa: {format_rupiah_display(r[COL_SISA_BAYAR])}" 
                    for i, r in df_pay_reset.iterrows()
                ]
                
                sel_idx_upd = st.selectbox("Pilih Record Pembayaran:", range(len(pay_options)), 
                                         format_func=lambda x: pay_options[x], key="desk_sel_susulan")
                
                file_susulan = st.file_uploader("Upload File Bukti Baru", key="desk_file_susulan")
                
                if st.button("⬆️ Upload Foto Sekarang", use_container_width=True):
                    if file_susulan:
                        mkt_name = df_pay_reset.iloc[sel_idx_upd][COL_MARKETING]
                        ok, msg = update_bukti_pembayaran_by_index(sel_idx_upd, file_susulan, mkt_name, actor="Admin")
                        if ok:
                            st.success("Foto bukti berhasil ditambahkan!")
                            st.cache_data.clear()
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error(msg)
                    else:
                        st.warning("Silakan pilih file terlebih dahulu.")

        render_section_watermark()

elif menu_nav == "📜 Global Audit Log":
    if IS_MOBILE:
        render_audit_mobile()
    else:
        # --- LOGIC DESKTOP ---
        st.markdown("## 📜 Global Audit Log")
        st.caption(
            "Rekaman jejak perubahan data. Transparansi data Admin & Manager.")

        # Load Data dari Service
        from audit_service import load_audit_log

        # Tombol Refresh
        if st.button("🔄 Refresh Log", use_container_width=True):
            st.cache_data.clear()
            st.rerun()

        with st.spinner("Memuat data log..."):
            df_raw = load_audit_log(spreadsheet)

        if not df_raw.empty:
            # 1. Jalankan Mapper Dinamis (Mengubah header GSheet lama/baru ke standar aplikasi)
            df_log = dynamic_column_mapper(df_raw)

            # 2. Pastikan kolom standar yang dibutuhkan UI tersedia (fallback "-" jika benar-benar tidak ketemu)
            standard_cols = ["Waktu", "User", "Status",
                             "Target Data", "Chat & Catatan", "Detail Perubahan"]
            for c in standard_cols:
                if c not in df_log.columns:
                    df_log[c] = "-"

            # 3. Urutkan Waktu (Terbaru di atas)
            try:
                df_log["Waktu"] = pd.to_datetime(
                    df_log["Waktu"], format="%d-%m-%Y %H:%M:%S", errors="coerce")
                df_log = df_log.sort_values(by="Waktu", ascending=False)
            except Exception:
                pass

            # --- FITUR FILTERING ---
            with st.expander("🔍 Filter Pencarian"):
                c1, c2 = st.columns(2)
                # Ambil list unik untuk filter
                all_users = df_log["User"].unique().tolist()
                all_sheets = df_log["Target Data"].unique().tolist()

                with c1:
                    filter_user = st.multiselect(
                        "Pilih Pelaku (User)", all_users)
                with c2:
                    filter_sheet = st.multiselect(
                        "Pilih Sheet/Data", all_sheets)

            # Terapkan Filter jika dipilih
            df_show = df_log.copy()
            if filter_user:
                df_show = df_show[df_show["User"].isin(filter_user)]
            if filter_sheet:
                df_show = df_show[df_show["Target Data"].isin(filter_sheet)]

            # --- TAMPILKAN DATA UI ---
            st.markdown(f"**Total Record:** {len(df_show)}")

            # 4. Render Dataframe (Pastikan Key Column Config sesuai hasil Mapping)
            st.dataframe(
                df_show[standard_cols],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Waktu": st.column_config.DatetimeColumn(
                        "🕒 Waktu",
                        format="D MMM YYYY, HH:mm",
                        width="small"
                    ),
                    "Target Data": st.column_config.TextColumn("Data"),
                    "Chat & Catatan": st.column_config.TextColumn("💬 Catatan / Chat", width="medium"),
                    "Detail Perubahan": st.column_config.TextColumn(
                        "📄 Detail Perubahan",
                        width="large",
                        help="Menampilkan detail perubahan data"
                    )
                }
            )

            # Download Button (Excel)
            if HAS_OPENPYXL:
                xb = df_to_excel_bytes(df_show, sheet_name="Audit_Log")
                if xb:
                    st.download_button(
                        "⬇️ Download Log (Excel)",
                        data=xb,
                        file_name="global_audit_log.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        else:
            st.info("Belum ada riwayat perubahan data.")

        # Watermark
        render_section_watermark()

elif menu_nav == "📊 Dashboard Admin":
    if IS_MOBILE:
        render_admin_mobile()
    else:
        st.markdown("## 📊 Dashboard Admin & Analytics")

        if not st.session_state.get("is_admin"):
            col_l1, col_l2, col_l3 = st.columns([1, 1, 1])
            with col_l2:
                with st.container(border=True):
                    st.markdown("### 🔐 Login Dashboard")
                    pwd_input = st.text_input("Masukkan Password Admin:", type="password", key="pwd_admin_desk")
                    if st.button("Masuk Ke Dashboard", use_container_width=True, type="primary"):
                        if verify_admin_password(pwd_input):
                            st.session_state["is_admin"] = True
                            st.rerun()
                        else:
                            st.error("Password salah. Akses ditolak.")
            st.stop()

        c_head1, c_head2 = st.columns([3, 1])
        with c_head1:
            st.info(f"Login sebagai: **{st.session_state.get('user_name')}** | Role: **{st.session_state.get('user_role').upper()}**")
        with c_head2:
            if st.button("🔓 Logout Admin", use_container_width=True):
                st.session_state["is_admin"] = False
                st.rerun()

        staff_list_global = get_daftar_staf_terbaru()
        df_all = load_all_reports(staff_list_global)

        if not df_all.empty:
            try:
                df_all[COL_TIMESTAMP] = pd.to_datetime(df_all[COL_TIMESTAMP], format="%d-%m-%Y %H:%M:%S", errors="coerce")
                df_all["Tanggal_Date"] = df_all[COL_TIMESTAMP].dt.date
                df_all["Kategori_Aktivitas"] = df_all[COL_TEMPAT].apply(
                    lambda x: "Digital/Kantor" if any(k in str(x) for k in ["Digital", "Marketing", "Ads", "Konten", "Telesales"])
                    else "Kunjungan Lapangan"
                )
            except Exception:
                pass

        st.markdown("### 🛠️ Data Controller (Mode Edit)")
        col_sel1, col_sel2 = st.columns([2, 1])
        with col_sel1:
            selected_staff_target = st.selectbox("👤 Pilih Target Staf (Untuk Mengedit Data):", ["-- Pilih Staf --"] + staff_list_global, key="adm_global_staff_sel")
        with col_sel2:
            st.caption("ℹ️ Pilih nama staf untuk mengaktifkan editor pada tab di bawah.")

        df_staff_current = pd.DataFrame()
        if selected_staff_target != "-- Pilih Staf --":
            ws_target = get_or_create_worksheet(selected_staff_target)
            if ws_target:
                df_staff_current = pd.DataFrame(ws_target.get_all_records())
                for col in NAMA_KOLOM_STANDAR:
                    if col not in df_staff_current.columns:
                        df_staff_current[col] = ""

        is_manager = (st.session_state.get("user_role") == "manager")
        
        tabs_labels = []
        if is_manager:
            tabs_labels.append("🔔 APPROVAL (ACC)")
        
        tabs_labels.extend([
            "📈 Produktivitas & AI",
            "🧲 Leads & Interest",
            "💬 Review & Feedback",
            "🖼️ Galeri Bukti",
            "📦 Master Data",
            "⚙️ Config Staff",
            "🗑️ Hapus Akun",
            "⚡ SUPER EDITOR"
        ])

        all_tabs = st.tabs(tabs_labels)
        tab_ptr = 0 

        if is_manager:
            with all_tabs[tab_ptr]:
                st.markdown("### 🔔 Pusat Persetujuan Manager")
                pending_data = get_pending_approvals()
                
                if not pending_data:
                    st.success("✅ Tidak ada data yang menunggu persetujuan.")
                else:
                    st.markdown(f"Menunggu persetujuan: **{len(pending_data)} data**")
                    for i, req in enumerate(pending_data):
                        with st.container(border=True):
                            c1, c2 = st.columns([3, 1])
                            with c1:
                                st.markdown(f"👤 **{req['Requestor']}** mengajukan perubahan pada `{req['Target Sheet']}`")
                                st.info(f"📝 Alasan: {req['Reason']}")
                            with c2:
                                st.caption(f"📅 {req['Timestamp']}")
                            
                            try:
                                new_d = json.loads(req.get("New Data JSON", "{}"))
                                old_d = json.loads(req.get("Old Data JSON", "{}"))
                                diff_list = [f"- {k}: `{old_d.get(k,'')}` ➡ **{v}**" for k, v in new_d.items() if str(v) != str(old_d.get(k,''))]
                                if diff_list:
                                    st.markdown("\n".join(diff_list))
                            except:
                                st.warning("Detail perubahan tidak dapat ditampilkan.")

                            ca, cb = st.columns(2)
                            if ca.button("✅ SETUJUI", key=f"acc_{i}", type="primary", use_container_width=True):
                                ok, m = execute_approval(i, "APPROVE", st.session_state["user_name"])
                                if ok:
                                    st.success(m)
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error(m)
                            
                            with cb.popover("❌ TOLAK REQUEST", use_container_width=True):
                                reason_rej = st.text_input("Alasan Penolakan:", key=f"rej_txt_{i}")
                                if st.button("Konfirmasi Tolak", key=f"rej_btn_{i}", type="primary", use_container_width=True):
                                    execute_approval(i, "REJECT", st.session_state["user_name"], reason_rej)
                                    st.warning("Request ditolak.")
                                    st.rerun()
            tab_ptr += 1

        with all_tabs[tab_ptr]:
            st.markdown("### 🚀 Analisa Kinerja & AI Insight")
            if not df_all.empty:
                d_opt = st.selectbox("Rentang Waktu Data:", [7, 14, 30, 90], index=2, key="d_opt_prod")
                cutoff = datetime.now(tz=TZ_JKT).date() - timedelta(days=d_opt)
                df_f = df_all[df_all["Tanggal_Date"] >= cutoff]

                col_m1, col_m2 = st.columns(2)
                with col_m1:
                    st.markdown("#### Total Laporan per Staf")
                    report_counts = df_f[COL_NAMA].value_counts()
                    st.bar_chart(report_counts)
                with col_m2:
                    if HAS_PLOTLY:
                        fig = px.pie(df_f, names="Kategori_Aktivitas", title="Proporsi Jenis Aktivitas", hole=0.3)
                        st.plotly_chart(fig, use_container_width=True)

                st.markdown("#### 🤖 AI / Machine Learning Management Insight")
                with st.spinner("Asisten Pak Nugroho sedang meninjau kinerja tim..."):
                    staf_stats_str = json.dumps(report_counts.to_dict(), indent=2)
                    full_prompt = f"""
                    [CONTEXT_DATA]
                    Nama Pemimpin: Pak Nugroho
                    Total Laporan Masuk: {len(df_f)}
                    Data Statistik Staf: {staf_stats_str}

                    [SYSTEM_INSTRUCTION]
                    Kamu adalah asisten kepercayaan Pak Nugroho. Gunakan bahasa Indonesia yang santun, cerdas, namun tetap membumi.
                    
                    PANDUAN PENULISAN:
                    1. Gunakan bahasa yang enak dibaca dan berwibawa.
                    2. JANGAN pernah menyebutkan target angka kunjungan mingguan.
                    3. Gunakan Logika Perbandingan Kompetitor.
                    4. Gunakan Teori Keunggulan Awal.
                    5. Jika laporan sedikit, gunakan sudut pandang Kualitas.
                    6. Berikan apresiasi kepada staf yang rajin.
                    7. JANGAN mengaku sebagai AI.

                    [TASK]
                    Berikan analisis kinerja tim Sales kepada Pak Nugroho secara naratif dan kreatif berdasarkan data laporan yang terkumpul hari ini.
                    """

                    ai_reply = ""
                    last_error = ""
                    for model_name in MODEL_FALLBACKS:
                        try:
                            if SDK == "new":
                                resp = client_ai.models.generate_content(model=model_name, contents=full_prompt)
                                ai_reply = resp.text
                            else:
                                model = genai_legacy.GenerativeModel(model_name)
                                resp = model.generate_content(full_prompt)
                                ai_reply = resp.text
                            if ai_reply: break 
                        except Exception as e:
                            last_error = str(e)
                            continue 
                    
                    if ai_reply:
                        st.info(ai_reply)
                    else:
                        # --- FALLBACK MECHANISM AKTIF ---
                        # Menggunakan data df_f yang sudah difilter berdasarkan tanggal
                        fallback_msg = generate_smart_insight_fallback(df_f, len(df_f))
                        
                        st.warning(
                            "⚠️ **AI Insight Limit Reached (Quota Exceeded)**\n"
                            "Sistem secara otomatis beralih ke analisis statistik data internal (Fallback Mode) "
                            "agar Anda tetap mendapatkan ringkasan kinerja."
                        )
                        st.success(fallback_msg) # Tampilkan sebagai success agar terlihat positif
            else:
                st.info("Belum ada data laporan masuk.")
            
            st.divider()
            
            st.markdown("### 🛠️ Editor Laporan Harian")
            if selected_staff_target == "-- Pilih Staf --":
                st.warning("👈 Silakan pilih nama staf di dropdown atas 'Data Controller' untuk mengedit data.")
            elif df_staff_current.empty:
                st.info(f"Data laporan untuk {selected_staff_target} masih kosong.")
            else:
                cols_prod = [COL_TIMESTAMP, COL_TEMPAT, COL_DESKRIPSI, COL_KESIMPULAN, COL_KENDALA]
                df_view = df_staff_current[cols_prod].copy()
                admin_smart_editor_ui(df_view, "prod_edit", selected_staff_target)
        tab_ptr += 1

        with all_tabs[tab_ptr]: # Tab Leads
            st.markdown("### 🧲 Leads Management")
            
            if not df_all.empty and COL_INTEREST in df_all.columns:
                st.markdown("#### 📥 Filter & Download Leads (Global)")
                
                # [FIX 1] Mapping Filter agar sinkron dengan data database
                # Pilihan UI -> Keyword Pencarian
                sel_in = st.selectbox("Pilih Tingkat Interest:", ["Under 50%", "50-75%", "75%-100%"], key="adm_leads_filter")
                
                # Ambil keyword utama saja (misal "50-75" dari "50-75% (B)")
                keyword_search = sel_in.split("%")[0].strip() 
                
                # [FIX 2] Gunakan .str.contains() agar pencarian lebih fleksibel
                # Ini akan mencocokkan "50-75" dengan "50-75% (B)" atau "50-75%"
                mask_leads = df_all[COL_INTEREST].astype(str).str.contains(keyword_search, case=False, na=False)
                df_leads_global = df_all[mask_leads].copy()
                
                # Tampilkan Data
                st.info(f"Ditemukan **{len(df_leads_global)}** leads potensial.")
                
                cols_view = [COL_TIMESTAMP, COL_NAMA, COL_NAMA_KLIEN, COL_KONTAK_KLIEN, COL_KESIMPULAN, COL_INTEREST]
                cols_final = [c for c in cols_view if c in df_leads_global.columns]
                
                st.dataframe(df_leads_global[cols_final], use_container_width=True, hide_index=True)

                # [FIX 3] Tambahkan Fitur Download Excel
                if HAS_OPENPYXL and not df_leads_global.empty:
                    xb = df_to_excel_bytes(df_leads_global[cols_final], sheet_name="Leads_Data")
                    if xb:
                        st.download_button(
                            label=f"⬇️ Download Excel ({sel_in})",
                            data=xb, 
                            file_name=f"Leads_{sel_in.replace(' ','_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
            else:
                st.info("Belum ada data leads yang masuk.")

            st.divider()
            
            # Bagian Editor Leads Per Staff (Tetap)
            st.markdown("#### 🛠️ Editor Leads (Per Staff)")
            if selected_staff_target == "-- Pilih Staf --":
                st.warning("👈 Silakan pilih nama staf di dropdown atas 'Data Controller' untuk mengedit.")
            elif not df_staff_current.empty:
                cols_leads = [COL_TIMESTAMP, COL_NAMA_KLIEN, COL_KONTAK_KLIEN, COL_INTEREST, COL_PENDING]
                cols_exist = [c for c in cols_leads if c in df_staff_current.columns]
                df_view = df_staff_current[cols_exist].copy()
                admin_smart_editor_ui(df_view, "leads_edit", selected_staff_target)
            else:
                st.info("Data leads untuk staf ini kosong.")
        tab_ptr += 1

        with all_tabs[tab_ptr]:
            st.markdown("### 💬 Review & Feedback")
            
            st.markdown("#### 📨 Kirim Feedback ke Staff")
            if not df_all.empty:
                for i, r in df_all.sort_values(by=COL_TIMESTAMP, ascending=False).head(5).iterrows():
                    with st.container(border=True):
                        st.markdown(f"**{r[COL_NAMA]}** | {r[COL_TIMESTAMP]} | 📍 {r[COL_TEMPAT]}")
                        st.write(f"📝 {r[COL_DESKRIPSI]}")
                        c_inp, c_btn = st.columns([3, 1])
                        with c_inp:
                            f_input = st.text_input("Feedback:", key=f"f_in_{i}", placeholder="Berikan masukan...")
                        with c_btn:
                            st.markdown("<br>", unsafe_allow_html=True) 
                            if st.button("Kirim", key=f"f_btn_{i}", use_container_width=True):
                                ok, m = kirim_feedback_admin(r[COL_NAMA], str(r[COL_TIMESTAMP]), f_input)
                                if ok:
                                    st.success("Terkirim!")
                                    st.rerun()
            
            st.divider()
            
            st.markdown("#### 🛠️ Koreksi Data Feedback")
            if selected_staff_target == "-- Pilih Staf --":
                st.warning("👈 Silakan pilih nama staf di dropdown atas untuk mengedit tabel feedback.")
            elif not df_staff_current.empty:
                cols_feed = [COL_TIMESTAMP, COL_DESKRIPSI, COL_FEEDBACK]
                if COL_FEEDBACK not in df_staff_current.columns:
                    df_staff_current[COL_FEEDBACK] = ""
                
                df_view = df_staff_current[cols_feed].copy()
                admin_smart_editor_ui(df_view, "feed_edit", selected_staff_target)
        tab_ptr += 1

        with all_tabs[tab_ptr]:
            st.markdown("### 🖼️ Galeri Foto Aktivitas")
            
            st.markdown("#### 📸 Tampilan Galeri")
            if not df_all.empty:
                df_img = df_all[df_all[COL_LINK_FOTO].str.contains("http", na=False)].head(12)
                c_gal = st.columns(3)
                for idx, row in enumerate(df_img.to_dict("records")):
                    with c_gal[idx % 3]:
                        img_clean = row[COL_LINK_FOTO].replace("www.dropbox.com", "dl.dropboxusercontent.com").replace("?dl=0", "")
                        st.image(img_clean, use_container_width=True, caption=f"{row[COL_NAMA]} @ {row[COL_TEMPAT]}")
            
            st.divider()
            
            with st.expander("🛠️ Klik di sini untuk mengedit Link Foto (Advanced)"):
                if selected_staff_target == "-- Pilih Staf --":
                    st.warning("Pilih staf di dropdown atas dahulu.")
                elif not df_staff_current.empty:
                    cols_img = [COL_TIMESTAMP, COL_TEMPAT, COL_LINK_FOTO]
                    df_view = df_staff_current[cols_img].copy()
                    admin_smart_editor_ui(df_view, "img_edit", selected_staff_target)
        tab_ptr += 1

        with all_tabs[tab_ptr]: # Tab Master Data
            st.markdown("### 📦 Master Data Editor")
            st.caption("Mengedit data global. Data diambil langsung dari database utama.")
            
            md_opt = st.radio("Pilih Data Master:", ["Closing Deal", "Pembayaran (DP/Termin)"], horizontal=True, key="adm_md_radio")
            
            # [PERBAIKAN] Gunakan Loader Standar agar format Rupiah/Tanggal terbaca otomatis
            df_md = pd.DataFrame()
            target_sheet_md = ""

            if md_opt == "Closing Deal":
                df_md = load_closing_deal() # Fungsi ini sudah otomatis parsing angka & tanggal
                target_sheet_md = SHEET_CLOSING_DEAL
            else:
                df_md = load_pembayaran_dp() # Fungsi ini sudah otomatis handle parsing Rupiah
                target_sheet_md = SHEET_PEMBAYARAN
            
            if not df_md.empty:
                # Tampilkan Editor dengan data yang sudah bersih
                # Note: admin_smart_editor_ui akan otomatis menangani diff checking
                admin_smart_editor_ui(df_md, f"md_editor_{md_opt}", target_sheet_md)
            else:
                st.info(f"Data Master '{md_opt}' masih kosong atau belum ada entri.")
        tab_ptr += 1

        with all_tabs[tab_ptr]: # Tab Config
            st.markdown("### ⚙️ Config Staff & Team")
            
            # Membagi Layout menjadi 2 Bagian: Individu & Team
            col_cfg_staf, col_cfg_team = st.columns(2)
            
            # --- BAGIAN KIRI: STAFF INDIVIDU ---
            with col_cfg_staf:
                st.markdown("#### 👤 Kelola Personel")
                with st.form("add_staf_form"):
                    st.caption("➕ Tambah Staf Baru")
                    new_nm = st.text_input("Nama Lengkap")
                    if st.form_submit_button("Simpan Staf"):
                        if new_nm.strip():
                            ok, m = tambah_staf_baru(new_nm)
                            if ok: 
                                st.success(m)
                                st.cache_data.clear()
                                time.sleep(1)
                                st.rerun()
                            else: st.error(m)
                        else: st.error("Nama tidak boleh kosong.")
                
                st.markdown("---")
                st.caption("🗑️ Hapus Akses Staf")
                del_nm = st.selectbox("Pilih Nama:", ["-"] + staff_list_global, key="del_st_cfg")
                if st.button("Hapus Akses", use_container_width=True):
                    if del_nm != "-":
                        ok, m = hapus_staf_by_name(del_nm)
                        if ok: st.success(m); st.cache_data.clear(); st.rerun()

            # --- BAGIAN KANAN: TEAM (FITUR BARU) ---
            with col_cfg_team:
                st.markdown("#### 🏆 Kelola Tim Sales")
                
                # [FIX] Form Tambah Team
                with st.expander("➕ Buat Team Baru", expanded=True):
                    with st.form("add_team_form"):
                        t_nama = st.text_input("Nama Team (Contoh: Tim Alpha)")
                        t_posisi = st.text_input("Posisi (Contoh: Sales Canvas)")
                        # Multiselect mengambil data dari daftar staf yang ada
                        t_anggota = st.multiselect("Pilih Anggota:", staff_list_global)
                        
                        if st.form_submit_button("Simpan Team"):
                            if t_nama and t_posisi and t_anggota:
                                ok, m = tambah_team_baru(t_nama, t_posisi, t_anggota)
                                if ok:
                                    st.success(m)
                                    st.cache_data.clear()
                                    time.sleep(1)
                                    st.rerun()
                                else: st.error(m)
                            else:
                                st.warning("Semua kolom wajib diisi.")
                
                st.divider()
                st.caption("📋 Data Team Saat Ini:")
                ws_tm = get_or_create_worksheet(SHEET_CONFIG_TEAM)
                df_tm = pd.DataFrame(ws_tm.get_all_records())
                if not df_tm.empty:
                    # Filter kolom agar rapi
                    valid_cols = [c for c in TEAM_COLUMNS if c in df_tm.columns]
                    admin_smart_editor_ui(df_tm[valid_cols], "team_cfg_view", SHEET_CONFIG_TEAM)
        tab_ptr += 1

        with all_tabs[tab_ptr]:
            st.markdown("### 🗑️ Hapus Akun Permanen")
            st.warning("Peringatan: Menghapus akun akan menghilangkan nama staf dari daftar pelapor.")
            target_del = st.selectbox("Pilih Akun:", ["-"] + staff_list_global, key="del_perm_sel")
            confirm_del = st.checkbox("Saya mengonfirmasi penghapusan ini.")
            
            if st.button("🔥 KONFIRMASI HAPUS", type="primary"):
                if target_del != "-" and confirm_del:
                    ok, m = hapus_staf_by_name(target_del)
                    if ok:
                        force_audit_log(st.session_state["user_name"], "DELETE USER", "Config_Users", f"Deleted {target_del}", "-")
                        st.success(f"Akun {target_del} berhasil dihapus.")
                        st.cache_data.clear()
                        time.sleep(1)
                        st.rerun()
                else:
                    st.error("Pilih nama staf dan centang konfirmasi terlebih dahulu.")
        tab_ptr += 1

        with all_tabs[tab_ptr]: # Tab Super Editor
            st.markdown("### ⚡ Super Editor (Advanced)")
            st.info("Mode Administrator: Mengedit langsung ke dalam Sheet.")
            
            se_type = st.selectbox("📂 Kategori Sheet:", ["Laporan Harian Staff", "Master Data Lainnya"], key="se_cat_fix")
            
            target_sheet_se = None
            # Variabel untuk menyimpan standar kolom
            forced_cols = None 
            
            if se_type == "Laporan Harian Staff":
                target_sheet_se = st.selectbox("👤 Pilih Nama Staff:", staff_list_global, key="se_st_sel")
                forced_cols = NAMA_KOLOM_STANDAR
            else:
                # Peta Nama Sheet -> Kolom Standar
                map_master = {
                    "Closing Deal": (SHEET_CLOSING_DEAL, CLOSING_COLUMNS),
                    "Pembayaran": (SHEET_PEMBAYARAN, PAYMENT_COLUMNS),
                    "Target Team": (SHEET_TARGET_TEAM, TEAM_CHECKLIST_COLUMNS),
                    "Target Individu": (SHEET_TARGET_INDIVIDU, INDIV_CHECKLIST_COLUMNS),
                    "Config Team": (SHEET_CONFIG_TEAM, TEAM_COLUMNS)
                }
                sel_master = st.selectbox("📄 Pilih Sheet Master:", list(map_master.keys()), key="se_ms_sel")
                target_sheet_se, forced_cols = map_master[sel_master]
            
            if target_sheet_se:
                st.divider()
                st.markdown(f"**Editing: `{target_sheet_se}`**")
                try:
                    ws_se = get_or_create_worksheet(target_sheet_se)
                    
                    # [PERBAIKAN KRUSIAL] Pastikan header sinkron sebelum load data
                    # Ini memperbaiki masalah data kosong pada Target Individu/Team
                    if forced_cols:
                        ensure_headers(ws_se, forced_cols)

                    # Load ulang setelah memastikan header
                    raw_data = ws_se.get_all_records()
                    df_se = pd.DataFrame(raw_data)
                    
                    # Jika dataframe kosong tapi kita punya header, buat dataframe kosong dengan kolom tersebut
                    if df_se.empty and forced_cols:
                        df_se = pd.DataFrame(columns=forced_cols)

                    # Jika ada data atau minimal kolom
                    if not df_se.empty or forced_cols:
                        # Terapkan pemaksaan urutan kolom (Reordering) & Normalisasi
                        if forced_cols:
                            for c in forced_cols:
                                if c not in df_se.columns: 
                                    df_se[c] = "" # Isi default string agar tidak error
                            df_se = df_se[forced_cols].copy()

                        # Bersihkan tipe data (Integer, Date) agar editor tidak error
                        df_se = clean_df_types_dynamically(df_se)
                        
                        admin_smart_editor_ui(df_se, f"super_edit_{target_sheet_se}", target_sheet_se)
                    else:
                        st.info(f"Sheet '{target_sheet_se}' benar-benar kosong.")
                except Exception as e:
                    st.error(f"Gagal memuat sheet: {e}")

        render_section_watermark()
